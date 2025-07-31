#!/usr/bin/env python3
"""
AWS Inventory Exporter

Collect metadata from multiple AWS services across one or more accounts
and write the results into an Excel workbook.
"""

# ───────────────────────────── imports ──────────────────────────────
from __future__ import annotations

import argparse
import json
import logging
import multiprocessing
import os
import random
import re
import sys
import threading
import time
from collections import OrderedDict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta, timezone
from functools import lru_cache
from itertools import islice
from typing import (
    Any,
    Callable,
    Dict,
    Iterator,
    List,
    Optional,
    Set,
    Tuple,
    Union,
)

import boto3
from botocore.client import BaseClient
from botocore.config import Config
from botocore.exceptions import (
    ClientError,
    EndpointConnectionError,
    NoCredentialsError,
)
from cron_descriptor import get_description as _cron_desc
from defusedxml import ElementTree as ET
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from tzlocal import get_localzone
from xml.sax.saxutils import unescape
from zoneinfo import ZoneInfo

# ─────────────────────────── configuration ──────────────────────────
CONNECT_TIMEOUT = int(os.getenv("AWS_CONNECT_TIMEOUT", "10"))
READ_TIMEOUT = int(os.getenv("AWS_READ_TIMEOUT", "60"))
MAX_ATTEMPTS = int(os.getenv("AWS_MAX_ATTEMPTS", "8"))

# -- parallelism -----------------------------------------------------
CPU_COUNT = multiprocessing.cpu_count()
MAX_REGIONS_IN_FLIGHT = min(int(os.getenv("MAX_PAR_REGION", "3")), 8)
MAX_TASKS_IN_REGION = min(int(os.getenv("MAX_PAR_TASK", "4")), 8)
_POOL_SIZE = MAX_REGIONS_IN_FLIGHT * MAX_TASKS_IN_REGION + 5  # HTTP-pool

RETRY_CONFIG = Config(
    retries={
        "mode": os.getenv("AWS_RETRY_MODE", "standard"),
        "max_attempts": MAX_ATTEMPTS,
    },
    connect_timeout=CONNECT_TIMEOUT,
    read_timeout=READ_TIMEOUT,
    max_pool_connections=_POOL_SIZE,
)

EXCEL_FILENAME = "aws_inventory.xlsx"
_MAX_EXCEL_NAME_LEN = 31

# cache-size defaults (env-overridable)
CLIENT_CACHE_MAX = int(os.getenv("AWS_CLIENT_CACHE_MAX", "128"))
FETCH_SPECS_CACHE_MAX = int(os.getenv("AWS_FETCH_SPECS_CACHE_MAX", "256"))
PARSE_VPN_CACHE_MAX = int(os.getenv("AWS_PARSE_VPN_CACHE_MAX", "128"))
BUCKET_REGION_CACHE_MAX = int(os.getenv("AWS_BUCKET_REGION_CACHE_MAX", "128"))
STORAGE_LENS_CACHE_MAX = int(os.getenv("AWS_STORAGE_LENS_CACHE_MAX", "128"))

# S3 / CloudWatch constants
_METRIC_PERIOD = 86_400  # seconds per day
_METRIC_OFFSET = 48  # hours of latency to tolerate
MAX_KEYS_FOR_FULL_SCAN = int(os.getenv("AWS_MAX_KEYS_FOR_FULL_SCAN", "20000"))

# misc
SECS_PER_YEAR = 31_536_000
DEFAULT_ROLE = "OrganizationAccountAccessRole"

REGION_TZ: Dict[str, str] = {
    "us-east-1": "America/New_York",
    "us-east-2": "America/Chicago",
    "us-west-1": "America/Los_Angeles",
    "us-west-2": "America/Los_Angeles",
    "ca-central-1": "America/Toronto",
    "sa-east-1": "America/Sao_Paulo",
    "eu-west-1": "Europe/Dublin",
    "eu-west-2": "Europe/London",
    "eu-west-3": "Europe/Paris",
    "eu-north-1": "Europe/Stockholm",
    "eu-central-1": "Europe/Berlin",
    "eu-central-2": "Europe/Zurich",
    "eu-south-1": "Europe/Rome",
    "eu-south-2": "Europe/Madrid",
    "ap-south-1": "Asia/Kolkata",
    "ap-south-2": "Asia/Kolkata",
    "ap-northeast-1": "Asia/Tokyo",
    "ap-northeast-2": "Asia/Seoul",
    "ap-northeast-3": "Asia/Tokyo",
    "ap-southeast-1": "Asia/Singapore",
    "ap-southeast-2": "Australia/Sydney",
    "ap-southeast-3": "Asia/Jakarta",
    "ap-southeast-4": "Australia/Melbourne",
    "ap-east-1": "Asia/Hong_Kong",
    "me-south-1": "Asia/Bahrain",
    "me-central-1": "Asia/Dubai",
    "af-south-1": "Africa/Johannesburg",
    "af-north-1": "Africa/Cairo",
    "il-central-1": "Asia/Jerusalem",
}

SERVICE_PREFIXES = {
    "ACM": "Certificate Manager",
    "AppFlow": "AppFlow",
    "AppMesh": "App Mesh",
    "AppRunner": "App Runner",
    "AppStream": "AppStream",
    "ApplicationDiscovery": "Application Discovery",
    "Artifact": "Artifact",
    "Athena": "Athena",
    "AutoScaling": "Auto Scaling",
    "Backup": "Backup",
    "Batch": "Batch",
    "Bedrock": "Bedrock",
    "Braket": "Braket",
    "Budgets": "Budgets",
    "CertificateManager": "Certificate Manager",
    "Chatbot": "Chatbot",
    "Chime": "Chime",
    "Cloud9": "Cloud9",
    "CloudFront": "CloudFront",
    "CloudFormation": "CloudFormation",
    "CloudSearch": "CloudSearch",
    "CloudTrail": "CloudTrail",
    "CloudWatch": "CloudWatch",
    "CodeArtifact": "CodeArtifact",
    "CodeBuild": "CodeBuild",
    "CodeCommit": "CodeCommit",
    "CodeDeploy": "CodeDeploy",
    "CodeGuru": "CodeGuru",
    "CodePipeline": "CodePipeline",
    "CodeStar": "CodeStar",
    "CodeWhisperer": "CodeWhisperer",
    "Comprehend": "Comprehend",
    "Config": "Config",
    "Console": "Console",
    "ControlTower": "Control Tower",
    "CostExplorer": "Cost Explorer",
    "DataBrew": "DataBrew",
    "DataPipeline": "Data Pipeline",
    "DataSync": "DataSync",
    "DatabaseMigration": "Database Migration Service",
    "DMS": "Database Migration Service",
    "DocumentDB": "DocumentDB",
    "DynamoDB": "DynamoDB",
    "EBS": "EBS",
    "EC2": "EC2",
    "ECR": "ECR",
    "ECS": "ECS",
    "EFS": "EFS",
    "EKS": "EKS",
    "Elemental": "Elemental",
    "ElastiCache": "ElastiCache",
    "ElasticBeanstalk": "Elastic Beanstalk",
    "ElasticLoadBalancing": "Elastic Load Balancing",
    "ElasticMapReduce": "EMR",
    "Elasticsearch": "Elasticsearch",
    "ELB": "Elastic Load Balancing",
    "EMR": "EMR",
    "EventBridge": "EventBridge",
    "Firewall": "Firewall Manager",
    "Forecast": "Forecast",
    "FraudDetector": "Fraud Detector",
    "FSx": "FSx",
    "GameLift": "GameLift",
    "GlobalAccelerator": "Global Accelerator",
    "Greengrass": "Greengrass",
    "GroundStation": "Ground Station",
    "GuardDuty": "GuardDuty",
    "Health": "Health Dashboard",
    "Honeycode": "Honeycode",
    "IAM": "IAM",
    "IdentityCenter": "IAM Identity Center",
    "Inspector": "Inspector",
    "IoT": "IoT",
    "IVS": "IVS",
    "Kendra": "Kendra",
    "Keyspaces": "Keyspaces",
    "KMS": "KMS",
    "Kinesis": "Kinesis",
    "LakeFormation": "Lake Formation",
    "Lex": "Lex",
    "LightSail": "LightSail",
    "License": "License Manager",
    "MachineLearning": "Machine Learning",
    "MediaConvert": "MediaConvert",
    "MediaLive": "MediaLive",
    "MediaPackage": "MediaPackage",
    "MediaStore": "MediaStore",
    "MediaTailor": "MediaTailor",
    "MigrationHub": "Migration Hub",
    "MQ": "MQ",
    "MSK": "MSK",
    "ManagedBlockchain": "Managed Blockchain",
    "ManagedStreamingKafka": "MSK",
    "Neptune": "Neptune",
    "NetworkManager": "Network Manager",
    "Organizations": "Organizations",
    "Pinpoint": "Pinpoint",
    "PinpointSMS": "Pinpoint SMS",
    "Polly": "Polly",
    "PrivateLink": "PrivateLink",
    "PrivateNetworks": "Private Networks",
    "QuantumLedger": "QLDB",
    "RAM": "Resource Access Manager",
    "RDS": "RDS",
    "Redshift": "Redshift",
    "Rekognition": "Rekognition",
    "RoboMaker": "RoboMaker",
    "S3": "S3",
    "SES": "SES",
    "SimpleEmail": "SES",
    "SimpleWorkflow": "SWF",
    "SMS": "Server Migration Service",
    "SNS": "SNS",
    "SQS": "SQS",
    "SSM": "Systems Manager",
    "SystemsManager": "Systems Manager",
    "SecurityHub": "Security Hub",
    "Serverless": "Serverless",
    "ServiceCatalog": "Service Catalog",
    "Sumerian": "Sumerian",
    "SWF": "SWF",
    "SnowFamily": "Snow Family",
    "StepFunctions": "Step Functions",
    "StorageGateway": "Storage Gateway",
    "StorageLens": "StorageLens",
    "STS": "Security Token Service",
    "Timestream": "Timestream",
    "Transcribe": "Transcribe",
    "Transfer": "Transfer Family",
    "Translate": "Translate",
    "VPC": "VPC",
    "Verified": "Verified Permissions",
    "WAF": "WAF",
    "WorkDocs": "WorkDocs",
    "WorkLink": "WorkLink",
    "WorkMail": "WorkMail",
    "WorkSpaces": "WorkSpaces",
    "XRay": "X-Ray",
}

# ───────────────────────────── logging ──────────────────────────────
LOG_FMT = "%(asctime)s %(levelname)-7s [%(account)s] %(name)s:%(lineno)d — %(message)s"
DATE_FMT = "%Y-%m-%dT%H:%M:%S%z"


class _TZFormatter(logging.Formatter):
    """Format %(asctime)s in the region’s local time-zone."""

    def __init__(self, fmt: str, datefmt: str):
        super().__init__(fmt, datefmt)
        region = (
            os.getenv("AWS_REGION") or os.getenv("AWS_DEFAULT_REGION") or "us-east-1"
        )
        self._tz = ZoneInfo(REGION_TZ.get(region, "UTC"))

    def formatTime(
        self, record: logging.LogRecord, datefmt: Optional[str] = None
    ) -> str:
        dt = datetime.fromtimestamp(record.created, tz=self._tz)
        return dt.strftime(datefmt or self.datefmt or "%Y-%m-%d %H:%M:%S%z")


class _AccountFilter(logging.Filter):
    """Guarantee `.account` exists on every log record."""

    def filter(self, record: logging.LogRecord) -> bool:
        record.account = getattr(record, "account", "-")
        return True


_handler = logging.StreamHandler(sys.stdout)
_handler.setFormatter(_TZFormatter(LOG_FMT, DATE_FMT))

logger = logging.getLogger("aws_inventory")
logger.setLevel(os.getenv("LOG_LEVEL", "INFO").upper())
logger.addHandler(_handler)
logger.addFilter(_AccountFilter())
logger.propagate = False


def log(level: str, msg: str, *args, account: str = "-", **kwargs) -> None:
    """Shortcut that injects `extra={'account': …}`."""
    extra = kwargs.pop("extra", {})
    extra.setdefault("account", account)
    getattr(logger, level)(msg, *args, extra=extra, **kwargs)


# ───────────────────────── helper utilities ─────────────────────────


def chunked(iterable, n: int):
    """Yield successive *n*-sized chunks from *iterable*."""
    it = iter(iterable)
    while piece := list(islice(it, n)):
        yield piece


_thread_local = threading.local()
_client_eviction_lock = threading.Lock()  # multi-thread cache eviction


def _lru_cache_per_thread() -> OrderedDict:
    if not hasattr(_thread_local, "client_cache"):
        _thread_local.client_cache = OrderedDict()
    return _thread_local.client_cache


def aws_client(
    service: str,
    region: str,
    session: Optional[boto3.Session] = None,
) -> BaseClient:
    """
    Thread-local LRU cache around ``boto3.Session.client``.
    boto3 clients are **not** thread-safe; building them on demand and
    caching per-thread avoids lock-contention in the AWS SDK.
    """
    if session is None:
        if not hasattr(_thread_local, "default_session"):
            _thread_local.default_session = boto3.Session()
        session = _thread_local.default_session

    cache: OrderedDict = _lru_cache_per_thread()
    key = (id(session), service, region)
    if key in cache:
        cache.move_to_end(key)
        return cache[key]

    client = session.client(service, region_name=region, config=RETRY_CONFIG)
    cache[key] = client

    # bounded cache
    if len(cache) > CLIENT_CACHE_MAX:
        with _client_eviction_lock:
            if len(cache) > CLIENT_CACHE_MAX:
                cache.popitem(last=False)

    return client


def retry_with_backoff(
    fn: Callable[..., Dict[str, Any]],
    *,
    max_attempts: int = MAX_ATTEMPTS,
    base_delay: float = 1.0,
    max_delay: float = 20.0,
    default: Optional[Dict[str, Any]] = None,
    account: str = "-",
    **kwargs: Any,
) -> Dict[str, Any]:
    """
    Retry throttled AWS calls with exponential back-off (+30 % jitter).
    """
    throttling_codes = {
        "Throttling",
        "ThrottlingException",
        "RequestLimitExceeded",
        "TooManyRequestsException",
        "SlowDown",
        "ProvisionedThroughputExceededException",
    }

    for attempt in range(1, max_attempts + 1):
        try:
            return fn(**kwargs)

        except ClientError as exc:
            code = exc.response.get("Error", {}).get("Code", "")
            if code not in throttling_codes and not code.startswith("Throttling"):
                raise  # not a throttle

            if attempt == max_attempts:
                log(
                    "error",
                    f"Max retries reached for {getattr(fn,'__name__',fn)}: {code}",
                    account=account,
                )
                return default or {}

            delay = min(base_delay * (2 ** (attempt - 1)), max_delay)
            delay += delay * 0.3 * random.random()  # ±30 % jitter (AWS guidance)
            log(
                "warning",
                f"Throttled ({code}) {attempt}/{max_attempts}; retrying in "
                f"{delay:.1f}s",
                account=account,
            )
            time.sleep(delay)

    return default or {}  # never reached – type guard


def _safe_aws_call(
    fn: Callable[..., Dict[str, Any]],
    *,
    default: Optional[Dict[str, Any]] = None,
    account: str = "-",
    retry: bool = True,
    **kwargs: Any,
) -> Dict[str, Any]:
    """
    Run an AWS SDK call and swallow *known-benign* faults.
    """
    kwargs.pop("service", None)
    kwargs.pop("operation", None)

    benign_s3_errors = {
        "NoSuchLifecycleConfiguration",
        "NoSuchTagSet",
        "NoSuchBucketPolicy",
        "NoSuchPublicAccessBlockConfiguration",
    }

    try:
        return (
            retry_with_backoff(fn, account=account, **kwargs) if retry else fn(**kwargs)
        )

    except ClientError as exc:
        code = exc.response.get("Error", {}).get("Code", "")
        client = getattr(fn, "__self__", None)
        is_s3 = (
            isinstance(client, BaseClient)
            and client.meta.service_model.service_name == "s3"
        )
        if is_s3 and code in benign_s3_errors:
            return default or {}

        log(
            "warning",
            f"AWS call {getattr(fn, '__name__', fn)} failed: {exc}",
            account=account,
        )
        return default or {}

    except EndpointConnectionError as exc:
        log("warning", f"Endpoint error: {exc}", account=account)
        return default or {}


def _safe_paginator(
    paginate_fn: Callable[..., Any],
    *,
    account: str = "-",
    **kwargs: Any,
) -> Iterator[Dict[str, Any]]:
    """
    Yield pages from a Boto3 paginator, swallowing network / client faults.
    """
    try:
        yield from paginate_fn(**kwargs)  # iterable → iterator (yield-from)
    except (ClientError, EndpointConnectionError) as exc:
        log("warning", f"Paginator aborted: {exc}", account=account)


# --------------------------------------------------------------------------
# Helper for strict paginator access
# --------------------------------------------------------------------------
def require_paginator(client: BaseClient, op: str) -> Paginator:
    """Return a paginator or raise if the operation can't be paginated."""
    if not client.can_paginate(op):
        raise RuntimeError(
            f"{client.meta.service_model.service_name} cannot paginate '{op}'"
        )
    return client.get_paginator(op)


# ─────────────────────── time / formatting helpers ───────────────────────

_TIMEZONE_OFFSET_RE = re.compile(r"([+-]\d{2})(\d{2})$")  # +0800 → +08:00


def _tz_for(region: str) -> str:
    if region in REGION_TZ:
        return REGION_TZ[region]
    if region.startswith(("us-gov-", "us-iso-")):
        return "America/Denver" if region.endswith("-iso-b-1") else "America/New_York"
    if region.startswith("cn-"):
        return "Asia/Shanghai"
    if region.startswith("il-"):
        return "Asia/Jerusalem"
    return "UTC"


def to_local(dt: Union[str, datetime, None], region: str) -> str:
    if not dt:
        return ""
    if isinstance(dt, str):
        iso = _TIMEZONE_OFFSET_RE.sub(r"\1:\2", dt.strip().replace("Z", "+00:00"))
        try:
            dt = datetime.fromisoformat(iso)
        except ValueError:
            return dt
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    loc = dt.astimezone(ZoneInfo(_tz_for(region)))
    return f"{loc:%Y-%m-%d %H:%M:%S} {_tz_for(region)}"


def human_size(size_bytes: Optional[int]) -> str:
    if size_bytes is None:
        return ""
    if size_bytes == 0:
        return "0 B"
    units = ["B", "KB", "MB", "GB", "TB", "PB"]
    size = float(size_bytes)
    for unit in units:
        if size < 1024 or unit == units[-1]:
            break
        size /= 1024
    if unit == "B":
        return f"{int(size)} B"
    return f"{size:.1f}".rstrip("0").rstrip(".") + f" {unit}"


def humanise_schedule(expr: str) -> Tuple[str, str]:
    if expr.startswith("rate("):
        val, unit = expr[5:-1].split()
        unit = unit.rstrip("s")
        return (f"Every {unit}" if val == "1" else f"Every {val} {unit}s"), ""
    if expr.startswith("at("):
        return "One-time", expr[3:-1]
    if expr.startswith("cron("):
        parts = expr[5:-1].split()
        if len(parts) == 6:
            m, h, dom, mon, dow, _ = parts
            if dom == "?" and mon == "*" and dow in ("?", "*"):
                return f"Daily at {h.zfill(2)}:{m.zfill(2)}", ""
            if dom.isdigit() and mon == "*" and dow in ("?", "*"):
                return f"Monthly on day {dom} at {h.zfill(2)}:{m.zfill(2)}", ""
        cron_expr = expr[5:-1]
        try:
            return _cron_desc(cron_expr, locale="en"), ""
        except TypeError:
            return _cron_desc(cron_expr), ""
    return expr, ""


def seconds_to_years(sec: Union[int, float, None]) -> str:
    if sec is None or sec < 0:
        return ""
    yrs = sec / SECS_PER_YEAR
    if yrs.is_integer():
        yrs_int = int(yrs)
        return f"{yrs_int} Year" + ("" if yrs_int == 1 else "s")
    return f"{yrs:.1f} Years"


def _plan_term_seconds(plan: dict) -> Optional[int]:
    if "termDurationInSeconds" in plan:
        return plan["termDurationInSeconds"]
    start = plan.get("startTime") or plan.get("start")
    end = plan.get("endTime") or plan.get("end")
    if isinstance(start, str):
        start = datetime.fromisoformat(start.replace("Z", "+00:00"))
    if isinstance(end, str):
        end = datetime.fromisoformat(end.replace("Z", "+00:00"))
    if start and end:
        return int((end - start).total_seconds())
    return None


# ─────────────────────────────────────────────────────────────────────
# IAM ROLE ASSUMPTION
# ─────────────────────────────────────────────────────────────────────


def assume_role(
    account_id: str,
    role_name: str = DEFAULT_ROLE,
    region: str = "us-east-1",
) -> Optional[boto3.Session]:
    sts = boto3.client("sts", region_name=region, config=RETRY_CONFIG)

    def _assume() -> Dict[str, str]:
        return sts.assume_role(
            RoleArn=f"arn:aws:iam::{account_id}:role/{role_name}",
            RoleSessionName="InventorySession",
        )["Credentials"]

    try:
        creds: Dict[str, str] = {}
        for attempt in range(1, MAX_ATTEMPTS + 1):
            try:
                creds = _assume()
                break
            except ClientError as exc:
                code = exc.response["Error"]["Code"]
                if (
                    code in {"AccessDenied", "AccessDeniedException"}
                    or attempt == MAX_ATTEMPTS
                ):
                    raise
                time.sleep(min(2**attempt, 20))
        if not creds:
            return None
    except (ClientError, NoCredentialsError) as exc:
        logger.error("AssumeRole failed: %s", exc, extra={"account": account_id})
        return None

    return boto3.Session(
        aws_access_key_id=creds["AccessKeyId"],
        aws_secret_access_key=creds["SecretAccessKey"],
        aws_session_token=creds["SessionToken"],
        region_name=region,
    )


# ─────────────────────────────────────────────────────────────────────
# BOUNDED CACHES
# ─────────────────────────────────────────────────────────────────────
@lru_cache(maxsize=FETCH_SPECS_CACHE_MAX)
def fetch_instance_type_specs(
    types: tuple[str, ...],
    region: str,
    session: boto3.Session,
) -> dict[str, Any]:
    """
    Return {instance_type: {vCPUs, Memory, NetworkPerformance}} for every
    *valid* type in *types*.

    • Works in ≤ 100-ID batches (API limit).
    • If the batch contains unknown / retired types, it removes them,
      logs a warning, and retries the remaining IDs - preventing
      `InvalidInstanceType` from killing the whole EC2 scan.
    """
    ec2 = aws_client("ec2", region, session)
    out: dict[str, Any] = {}

    # we may need to mutate the set, so copy into a list
    pending: list[str] = list(types)

    while pending:
        chunk = pending[:100]
        try:
            resp = ec2.describe_instance_types(InstanceTypes=chunk)
            for it in resp["InstanceTypes"]:
                mib = it["MemoryInfo"]["SizeInMiB"]
                gib = mib / 1024
                out[it["InstanceType"]] = {
                    "vCPUs": it["VCpuInfo"]["DefaultVCpus"],
                    "Memory": (
                        f"{gib/1024:.1f} TB" if gib >= 1024 else f"{round(gib)} GB"
                    ),
                    "NetworkPerformance": it["NetworkInfo"].get(
                        "NetworkPerformance", ""
                    ),
                }

            pending = pending[100:]  # finished this batch; move on

        except ClientError as exc:
            code = exc.response["Error"]["Code"]
            if code != "InvalidInstanceType":
                raise  # real error – re-raise

            # Parse “… do not exist: [g3.4xlarge, r3.8xlarge]”
            msg = exc.response["Error"]["Message"]
            bad = [t.strip() for t in msg.split(":")[-1].strip(" []").split(",")]
            log(
                "warning",
                f"Region {region}: unknown / retired instance types skipped: {', '.join(bad)}",
                account="-",
            )
            # remove bad IDs from *pending* and retry the reduced batch
            pending = [t for t in pending if t not in bad]

    return out


@lru_cache(maxsize=PARSE_VPN_CACHE_MAX)
def parse_vpn_config(xml: str) -> Tuple[str, List[str]]:
    root = ET.fromstring(unescape(xml).encode())
    cg = root.findtext(".//customer_gateway/tunnel_outside_address/ip_address") or "N/A"
    ips = {
        ip.text
        for ip in root.findall(
            ".//ipsec_tunnel/vpn_gateway/tunnel_outside_address/ip_address"
        )
        if ip.text
    }
    return cg, sorted(ips) or ["N/A"]


@lru_cache(maxsize=BUCKET_REGION_CACHE_MAX)
def bucket_region(
    bucket: str,
    region_hint: str,
    session: Optional[boto3.Session] = None,
) -> Optional[str]:
    s3 = aws_client("s3", region_hint, session)
    try:
        loc = s3.get_bucket_location(Bucket=bucket)["LocationConstraint"]
        if loc is None:
            if region_hint.startswith("us-gov"):
                return "us-gov-west-1"
            if region_hint.startswith(("us-iso-", "us-isob-")):
                return "us-iso-east-1"
            if region_hint.startswith("cn-"):
                return "cn-north-1"
            return "us-east-1"
        return loc
    except s3.exceptions.NoSuchBucket:
        logger.warning("Bucket %s no longer exists", bucket)
    except s3.exceptions.ClientError as exc:
        if exc.response["Error"].get("Code") == "PermanentRedirect":
            return exc.response["Error"]["BucketRegion"]
        logger.debug("bucket_region(%s) failed: %s", bucket, exc)
    return None


@lru_cache(maxsize=STORAGE_LENS_CACHE_MAX)
def storage_lens_metrics(
    bucket: str,
    region: str,
    session: Optional[boto3.Session] = None,
) -> Tuple[Optional[int], Optional[int]]:
    try:
        sl = aws_client("s3control", region, session)
        acct = aws_client("sts", region, session).get_caller_identity()["Account"]
        cfgs = []
        nxt = None
        while True:
            if nxt:
                resp = sl.list_storage_lens_configurations(AccountId=acct, NextToken=nxt)
            else:
                resp = sl.list_storage_lens_configurations(AccountId=acct)
            cfgs.extend(resp.get("StorageLensConfigurationList", []))
            nxt = resp.get("NextToken")
            if not nxt:
                break
        if not cfgs:
            return None, None
        cfg = sl.get_storage_lens_configuration(
            AccountId=acct,
            ConfigId=cfgs[0]["Id"],
        )["StorageLensConfiguration"]
        metrics = cfg["AccountLevel"]["BucketLevel"]["AdvancedMetrics"]
        return metrics.get("TotalStorageBytes"), metrics.get("TotalObjectCount")
    except Exception as exc:
        logger.debug("StorageLens lookup failed for %s: %s", bucket, exc)
        return None, None


@lru_cache(maxsize=1)
def get_lightsail_supported_regions(session: boto3.Session, account: str) -> set[str]:
    """
    Return a set of region names where Lightsail is supported.
    This is cached so it only runs once per account.
    """
    try:
        ls_client = aws_client("lightsail", "us-east-1", session)
        regions_resp = ls_client.get_regions(includeAvailabilityZones=False)
        return {r["name"] for r in regions_resp.get("regions", [])}
    except Exception as e:
        log(
            "warning",
            f"Could not query for available Lightsail regions: {e}",
            account=account,
        )
        return set()


# --------------------------------------------------------------------------
# VPN COLLECTOR
# --------------------------------------------------------------------------


def get_vpn_details(ec2_client: BaseClient, alias: str) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    resp = _safe_aws_call(
        ec2_client.describe_vpn_connections,
        default={"VpnConnections": []},
        account=alias,
    )
    for conn in resp.get("VpnConnections", []):
        xml = conn.get("CustomerGatewayConfiguration", "")
        cg, tns = parse_vpn_config(xml) if xml else ("N/A", ["N/A"])
        name = next(
            (t["Value"] for t in conn.get("Tags", []) if t["Key"].lower() == "name"),
            "",
        )
        out.append(
            {
                "VpnConnectionId": conn.get("VpnConnectionId", ""),
                "Name": name,
                "State": conn.get("State", ""),
                "CustomerGateway": conn.get("CustomerGatewayId", ""),
                "CustomerGatewaySource": cg,
                "TunnelOutsideIps": ", ".join(tns),
                "Region": ec2_client.meta.region_name,
                "AccountAlias": alias,
            }
        )
    return out


# --------------------------------------------------------------------------
# S3 HELPERS & COLLECTOR
# --------------------------------------------------------------------------


def _s3_select_sum_and_count(
    s3: BaseClient,
    bucket: str,
    key: str,
    *,
    account: str = "-",
) -> tuple[int, int]:
    sql = (
        'SELECT CAST(sum(cast(s."Size" AS int)) AS bigint), '
        "       CAST(count(1) AS bigint) "
        "FROM S3Object s"
    )
    resp = _safe_aws_call(
        s3.select_object_content,
        account=account,
        retry=True,
        Bucket=bucket,
        Key=key,
        ExpressionType="SQL",
        Expression=sql,
        InputSerialization={
            "CSV": {"FileHeaderInfo": "USE"},
            "CompressionType": "NONE",
        },
        OutputSerialization={"JSON": {"RecordDelimiter": "\n"}},
    )
    buf = b""
    for ev in resp.get("Payload", []):
        if rec := ev.get("Records"):
            buf += rec["Payload"]
    return tuple(json.loads(buf.decode()) if buf else (0, 0))


@lru_cache(maxsize=STORAGE_LENS_CACHE_MAX)
def inventory_metrics(
    bucket: str,
    region_hint: str,
    session: Optional[boto3.Session] = None,
) -> tuple[Optional[int], Optional[int]]:
    try:
        inv = aws_client("s3", region_hint, session)
        cfgs = inv.list_bucket_inventory_configurations(Bucket=bucket).get(
            "InventoryConfigurationList", []
        )
        if not cfgs:
            return None, None
        cfg = cfgs[0]
        inv_id = cfg["Id"]
        prefix = cfg["Destination"]["S3BucketDestination"].get("Prefix", "").rstrip("/")
        date_part = (datetime.now(timezone.utc) - timedelta(days=1)).strftime(
            "%Y-%m-%d"
        )
        manifest_key = f"{prefix or inv_id}/{date_part}/manifest.json"
        dest = cfg["Destination"]["S3BucketDestination"]["Bucket"].split(":::")[-1]
        dest_region = bucket_region(dest, region_hint, session)
        s3_dest = aws_client("s3", dest_region, session)
        manifest = json.loads(
            s3_dest.get_object(Bucket=dest, Key=manifest_key)["Body"].read()
        )
        total_bytes = total_objects = 0
        with ThreadPoolExecutor(max_workers=8) as pool:
            futures = [
                pool.submit(
                    _s3_select_sum_and_count,
                    s3_dest,
                    dest,
                    file_["key"],
                    account=bucket,
                )
                for file_ in manifest["files"]
            ]
            for fut in as_completed(futures):
                b, o = fut.result()
                total_bytes += b
                total_objects += o
        return total_bytes, total_objects
    except Exception as exc:
        logger.warning("Inventory metrics (S3 Select) failed for %s: %s", bucket, exc)
        return None, None


def get_s3_details(
    s3_client: BaseClient,
    all_buckets: List[Dict[str, Any]],
    alias: str,
    region: str,
    session: boto3.Session,
) -> List[Dict[str, Any]]:
    now = datetime.now(timezone.utc)
    start = now - timedelta(hours=_METRIC_OFFSET + 2)

    buckets = [
        b for b in all_buckets if bucket_region(b["Name"], region, session) == region
    ]
    if not buckets:
        return []

    cw = aws_client("cloudwatch", region, session)
    queries, id_map = [], {}
    for idx, b in enumerate(buckets):
        name = b["Name"]
        for metric, stype, key in [
            ("BucketSizeBytes", "StandardStorage", "size"),
            ("NumberOfObjects", "AllStorageTypes", "count"),
        ]:
            qid = f"b{idx:05}{key[0]}"
            id_map[qid] = (name, key)
            queries.append(
                {
                    "Id": qid,
                    "MetricStat": {
                        "Metric": {
                            "Namespace": "AWS/S3",
                            "MetricName": metric,
                            "Dimensions": [
                                {"Name": "BucketName", "Value": name},
                                {"Name": "StorageType", "Value": stype},
                            ],
                        },
                        "Period": _METRIC_PERIOD,
                        "Stat": "Average",
                    },
                    "ReturnData": True,
                }
            )

    metric_results: List[Dict[str, Any]] = []
    for chunk in (queries[i : i + 500] for i in range(0, len(queries), 500)):
        token = None
        while True:
            resp = _safe_aws_call(
                cw.get_metric_data,
                default={},
                account=alias,
                MetricDataQueries=chunk,
                StartTime=start,
                EndTime=now,
                ScanBy="TimestampDescending",
                **({"NextToken": token} if token else {}),
            )
            if not resp:
                break
            metric_results.extend(resp.get("MetricDataResults", []))
            token = resp.get("NextToken")
            if not token:
                break

    cw_data = {b["Name"]: {"size": None, "count": None} for b in buckets}
    for r in metric_results:
        vals = r.get("Values")
        if vals:
            name, key = id_map[r["Id"]]
            cw_data[name][key] = int(vals[0])

    rows: List[Dict[str, Any]] = []
    for b in buckets:
        name = b["Name"]
        created = b["CreationDate"]
        size = cw_data[name]["size"]
        count = cw_data[name]["count"]

        method = None
        if size is not None and count is not None:
            method = "CloudWatch"
        elif size is not None:
            count = 0
            method = "CloudWatch-SizeOnly"
        elif count is not None:
            size = 0
            method = "CloudWatch-CountOnly"
        else:
            size, count = storage_lens_metrics(name, region, session)
            if size is not None or count is not None:
                method = "StorageLens"
            else:
                size, count = inventory_metrics(name, region, session)
                if size is not None or count is not None:
                    method = "Inventory"

        if method is None:
            size = count = 0
            for i, page in enumerate(
                s3_client.get_paginator("list_objects_v2").paginate(
                    Bucket=name, PaginationConfig={"PageSize": 250}
                ),
                start=1,
            ):
                for obj in page.get("Contents", []):
                    size += obj.get("Size", 0)
                    count += 1
                    if count >= MAX_KEYS_FOR_FULL_SCAN:
                        size = count = None
                        break
                if count is None or i >= 5:
                    method = "TooLargeToScan" if count is None else "ListObjects"
                    break
            if method is None:
                method = "ListObjects"

        ver = _safe_aws_call(
            s3_client.get_bucket_versioning, default={}, account=alias, Bucket=name
        ).get("Status", "Disabled")

        enc_cfg = _safe_aws_call(
            s3_client.get_bucket_encryption, default={}, account=alias, Bucket=name
        )
        enc = (
            "Enabled"
            if "ServerSideEncryptionConfiguration" in enc_cfg
            else "Not enabled"
        )

        try:
            pab_fn = s3_client.get_public_access_block
        except AttributeError:  # pre-GA preview (rare)
            pab_fn = s3_client.get_bucket_public_access_block

        pab_cfg = _safe_aws_call(pab_fn, default={}, account=alias, Bucket=name).get(
            "PublicAccessBlockConfiguration", {}
        )
        public = "Blocked" if pab_cfg and all(pab_cfg.values()) else "Not fully blocked"

        policy_flag = _safe_aws_call(
            s3_client.get_bucket_policy_status,
            default={"PolicyStatus": {"IsPublic": False}},
            account=alias,
            Bucket=name,
        )["PolicyStatus"].get("IsPublic", False)
        policy = "Public" if policy_flag else "Not Public"

        lifecycle = _safe_aws_call(
            s3_client.get_bucket_lifecycle_configuration,
            default={"Rules": []},
            account=alias,
            Bucket=name,
        ).get("Rules", [])

        tagset = _safe_aws_call(
            s3_client.get_bucket_tagging,
            default={"TagSet": []},
            account=alias,
            Bucket=name,
        )["TagSet"]
        tags = {t["Key"]: t["Value"] for t in tagset}

        rows.append(
            {
                "BucketName": name,
                "CreationDate": to_local(created, region),
                "Region": region,
                "Size": human_size(size),
                "ObjectCount": "" if count is None else count,
                "LastMetricsUpdate": (
                    "" if (size is None and count is None) else to_local(now, region)
                ),
                "MetricsCalculationMethod": method,
                "Versioning": ver,
                "Encryption": enc,
                "PublicAccess": public,
                "PolicyStatus": policy,
                "LifecycleRules": lifecycle,
                "Tags": tags,
                "AccountAlias": alias,
            }
        )
    return rows


# --------------------------------------------------------------------------
# EC2 COLLECTOR
# --------------------------------------------------------------------------
def get_ec2_details(
    ec2_client: BaseClient,
    backup_client: BaseClient,
    alias: str,
    session: boto3.Session,
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []

    instances = [
        inst
        for page in _safe_paginator(
            ec2_client.get_paginator("describe_instances").paginate, account=alias
        )
        for r in page.get("Reservations", [])
        for inst in r.get("Instances", [])
    ]
    if not instances:
        return out

    inst_ids = {i["InstanceId"] for i in instances}
    volume_info: Dict[str, Dict[str, Any]] = {}
    for chunk in chunked(sorted(inst_ids), 500):
        resp = _safe_aws_call(
            ec2_client.describe_volumes,
            default={"Volumes": []},
            account=alias,
            Filters=[{"Name": "attachment.instance-id", "Values": list(chunk)}],
        )
        for v in resp.get("Volumes", []):
            volume_info[v["VolumeId"]] = {
                "Size": v.get("Size"),
                "Type": v.get("VolumeType"),
            }

    types = {i["InstanceType"] for i in instances}
    specs: Dict[str, Any] = {}
    for chunk in chunked(tuple(types), 200):
        specs.update(
            fetch_instance_type_specs(
                tuple(chunk), ec2_client.meta.region_name, session
            )
        )

    eip_resp = _safe_aws_call(
        ec2_client.describe_addresses,
        default={"Addresses": []},
        account=alias,
    )
    eips = {
        addr.get("PublicIp")
        for addr in eip_resp.get("Addresses", [])
        if addr.get("PublicIp")
    }

    prot_i, prot_v, plan_i, plan_v = set(), set(), set(), set()
    all_i = all_v = False
    for page in _safe_paginator(
        backup_client.get_paginator("list_protected_resources").paginate, account=alias
    ):
        for r in page.get("Results", []):
            arn = r.get("ResourceArn", "")
            if arn.endswith("/instance/*"):
                all_i = True
            elif arn.endswith("/volume/*"):
                all_v = True
            elif ":instance/" in arn:
                prot_i.add(arn.rsplit("/", 1)[-1])
            elif ":volume/" in arn:
                prot_v.add(arn.rsplit("/", 1)[-1])

    plans = _safe_aws_call(
        backup_client.list_backup_plans, default={"BackupPlansList": []}, account=alias
    )["BackupPlansList"]
    for p in plans:
        for sel in _safe_aws_call(
            backup_client.list_backup_selections,
            default={"BackupSelectionsList": []},
            account=alias,
            BackupPlanId=p["BackupPlanId"],
        )["BackupSelectionsList"]:
            cfg = _safe_aws_call(
                backup_client.get_backup_selection,
                default={"BackupSelection": {}},
                account=alias,
                BackupPlanId=p["BackupPlanId"],
                SelectionId=sel["SelectionId"],
            )["BackupSelection"]
            for arn in cfg.get("Resources", []):
                if arn.endswith("/instance/*"):
                    all_i = True
                elif arn.endswith("/volume/*"):
                    all_v = True
                elif ":instance/" in arn:
                    plan_i.add(arn.rsplit("/", 1)[-1])
                elif ":volume/" in arn:
                    plan_v.add(arn.rsplit("/", 1)[-1])

    for inst in instances:
        iid = inst["InstanceId"]
        attached = [
            bd["Ebs"]["VolumeId"]
            for bd in inst.get("BlockDeviceMappings", [])
            if bd.get("Ebs", {}).get("VolumeId")
        ]
        covered = (
            all_i
            or iid in prot_i
            or iid in plan_i
            or any(v in prot_v or v in plan_v for v in attached)
            or (all_v and attached)
        )
        spec = specs.get(inst["InstanceType"], {})
        pub = inst.get("PublicIpAddress", "")
        ipt = "Elastic" if pub in eips else ("Ephemeral" if pub else "")
        out.append(
            {
                "Name": next(
                    (
                        t["Value"]
                        for t in inst.get("Tags", [])
                        if t["Key"].lower() == "name"
                    ),
                    "",
                ),
                "InstanceId": iid,
                "InstanceType": inst["InstanceType"],
                "KeyPair": inst.get("KeyName", ""),
                "vCPUs": spec.get("vCPUs"),
                "Memory": spec.get("Memory"),
                "NetworkPerformance": spec.get("NetworkPerformance"),
                "AvailabilityZone": inst.get("Placement", {}).get(
                    "AvailabilityZone", ""
                ),
                "PublicIP": pub,
                "PrivateIP": inst.get("PrivateIpAddress", ""),
                "IPType": ipt,
                "EBSVolumes": ", ".join(
                    f"{v}:{volume_info.get(v,{}).get('Size','?')}GB" for v in attached
                ),
                "OS": inst.get("PlatformDetails", "Linux/UNIX"),
                "AWSBackup": "Covered" if covered else "Not Covered",
                "Region": ec2_client.meta.region_name,
                "AccountAlias": alias,
            }
        )
    return out


# ------------------------------------------------------------------
# EC2 RESERVED-INSTANCES COLLECTOR
# ------------------------------------------------------------------
def get_ec2_reserved_instances(
    ec2_client: BaseClient,
    alias: str,
    _session: boto3.Session,
) -> list[dict[str, Any]]:
    """
    Return every active EC2 Reserved Instance in *ec2_client.region*.

    Note: AWS now recommends **regional** RIs, so AvailabilityZone is
    usually blank – we omit that column entirely to avoid empty cells.
    """
    out: list[dict[str, Any]] = []

    # `describe_reserved_instances` isn’t paginated, so a single call is fine
    resp = _safe_aws_call(
        ec2_client.describe_reserved_instances,
        default={"ReservedInstances": []},
        account=alias,
        Filters=[{"Name": "state", "Values": ["active"]}],
    )

    for ri in resp.get("ReservedInstances", []):
        out.append(
            {
                "ReservedInstancesId": ri["ReservedInstancesId"],
                "InstanceType": ri["InstanceType"],
                "Scope": ri.get("Scope", ""),  # Region | AvailabilityZone
                "InstanceCount": ri.get("InstanceCount", 0),
                "OfferingType": ri.get("OfferingType", ""),
                "ProductDescription": ri.get("ProductDescription", ""),
                "Duration": seconds_to_years(ri["Duration"]),
                "FixedPrice": ri["FixedPrice"],
                "UsagePrice": ri["UsagePrice"],
                "CurrencyCode": ri["CurrencyCode"],
                "StartTime": to_local(ri.get("Start"), ec2_client.meta.region_name),
                "State": ri["State"],
                "Region": ec2_client.meta.region_name,
                "AccountAlias": alias,
            }
        )

    return out


# ------------------------------------------------------------------
# SAVINGS-PLANS COLLECTOR
# ------------------------------------------------------------------


def get_savings_plan_details(
    sp_client: BaseClient, alias: str, _session: boto3.Session
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    next_token: Optional[str] = None
    while True:
        params: Dict[str, Any] = {"nextToken": next_token} if next_token else {}
        resp = _safe_aws_call(
            sp_client.describe_savings_plans, default={}, account=alias, **params
        )
        if not resp:
            break
        for plan in resp.get("savingsPlans", []):
            start = plan.get("startTime") or plan.get("start")
            end = plan.get("endTime") or plan.get("end")
            out.append(
                {
                    "SavingsPlanId": plan.get("savingsPlanId", ""),
                    "SavingsPlanArn": plan.get("savingsPlanArn", ""),
                    "State": plan.get("state", ""),
                    "Start": to_local(start, sp_client.meta.region_name),
                    "End": to_local(end, sp_client.meta.region_name),
                    "Term": seconds_to_years(_plan_term_seconds(plan)),
                    "PaymentOption": plan.get("paymentOption", ""),
                    "PlanType": plan.get("savingsPlanType", ""),
                    "Region": "global",
                    "AccountAlias": alias,
                }
            )
        next_token = resp.get("nextToken")
        if not next_token:
            break
    return out


# --------------------------------------------------------------------------
# RDS COLLECTORS
# --------------------------------------------------------------------------


def get_rds_details(
    rds_client: BaseClient, alias: str, session: boto3.Session
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    all_dbs: List[Dict[str, Any]] = []
    for page in _safe_paginator(
        rds_client.get_paginator("describe_db_instances").paginate, account=alias
    ):
        all_dbs.extend(page.get("DBInstances", []))
    if not all_dbs:
        return out
    classes = {db["DBInstanceClass"].removeprefix("db.") for db in all_dbs}
    specs: Dict[str, Any] = {}
    for chunk in chunked(sorted(classes), 20):
        specs.update(
            fetch_instance_type_specs(
                tuple(chunk), rds_client.meta.region_name, session
            )
        )
    for db in all_dbs:
        arn = db["DBInstanceArn"]
        tags = {
            t["Key"]: t["Value"]
            for t in _safe_aws_call(
                rds_client.list_tags_for_resource,
                default={"TagList": []},
                account=alias,
                ResourceName=arn,
            )["TagList"]
        }
        cls = db["DBInstanceClass"]
        spec = specs.get(cls.removeprefix("db."), {})
        subnet_grp = db.get("DBSubnetGroup", {}).get("DBSubnetGroupName", "")
        vpc_sg_ids = [
            sg.get("VpcSecurityGroupId", "") for sg in db.get("VpcSecurityGroups", [])
        ]
        pub_bool = db.get("PubliclyAccessible", False)
        visibility = "Public" if pub_bool else "Private"
        out.append(
            {
                "DBInstanceIdentifier": db.get("DBInstanceIdentifier", ""),
                "Engine": db.get("Engine", ""),
                "DBInstanceClass": cls,
                "vCPUs": spec.get("vCPUs"),
                "Memory": spec.get("Memory"),
                "MultiAZ": db.get("MultiAZ", False),
                "PubliclyAccessible": pub_bool,
                "EndpointVisibility": visibility,
                "StorageType": db.get("StorageType", ""),
                "AllocatedStorage": db.get("AllocatedStorage", 0),
                "EndpointAddress": db.get("Endpoint", {}).get("Address", ""),
                "EndpointPort": db.get("Endpoint", {}).get("Port", ""),
                "InstanceCreateTime": to_local(
                    db.get("InstanceCreateTime"), rds_client.meta.region_name
                ),
                "LicenseModel": db.get("LicenseModel", ""),
                "VpcSecurityGroupIds": vpc_sg_ids,
                "DBSubnetGroup": subnet_grp,
                "Tags": tags,
                "Region": rds_client.meta.region_name,
                "AccountAlias": alias,
            }
        )
    return out


def get_rds_reserved_instances(
    rds_client: BaseClient, alias: str, session: boto3.Session
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for page in _safe_paginator(
        rds_client.get_paginator("describe_reserved_db_instances").paginate,
        account=alias,
    ):
        for ri in page.get("ReservedDBInstances", []):
            out.append(
                {
                    "ReservedDBInstanceId": ri["ReservedDBInstanceId"],
                    "DBInstanceClass": ri["DBInstanceClass"],
                    "Duration": seconds_to_years(ri["Duration"]),
                    "FixedPrice": ri["FixedPrice"],
                    "UsagePrice": ri["UsagePrice"],
                    "CurrencyCode": ri["CurrencyCode"],
                    "StartTime": to_local(
                        ri.get("StartTime"), rds_client.meta.region_name
                    ),
                    "State": ri["State"],
                    "MultiAZ": ri["MultiAZ"],
                    "OfferingType": ri["OfferingType"],
                    "ProductDescription": ri["ProductDescription"],
                    "Region": rds_client.meta.region_name,
                    "AccountAlias": alias,
                }
            )
    return out


# --------------------------------------------------------------------------
# SES, SNS & ROUTE53 COLLECTORS
# --------------------------------------------------------------------------
def get_ses_details(ses_client: BaseClient, alias: str) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for page in _safe_paginator(
        ses_client.get_paginator("list_identities").paginate, account=alias
    ):
        for ident in page.get("Identities", []):
            attrs = _safe_aws_call(
                ses_client.get_identity_verification_attributes,
                default={"VerificationAttributes": {}},
                account=alias,
                Identities=[ident],
            )["VerificationAttributes"]
            status = attrs.get(ident, {}).get("VerificationStatus", "")
            out.append(
                {
                    "Identity": ident,
                    "VerificationStatus": status,
                    "Region": ses_client.meta.region_name,
                    "AccountAlias": alias,
                }
            )
    return out


def get_sns_details(sns_client: BaseClient, alias: str) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for page in _safe_paginator(
        sns_client.get_paginator("list_topics").paginate, account=alias
    ):
        for topic in page.get("Topics", []):
            arn = topic["TopicArn"]

            subs: List[str] = []
            for sp in _safe_paginator(
                sns_client.get_paginator("list_subscriptions_by_topic").paginate,
                account=alias,
                TopicArn=arn,
            ):
                subs.extend(
                    f"{s.get('Protocol','')}://{s.get('Endpoint','')}"
                    for s in sp.get("Subscriptions", [])
                )

            out.append(
                {
                    "TopicArn": arn,
                    "TopicName": arn.split(":")[-1],
                    "SubscriptionCount": len(subs),
                    "Subscriptions": subs,  # <- now populated
                    "Region": sns_client.meta.region_name,
                    "AccountAlias": alias,
                }
            )
    return out


def get_route53_details(r53_client: BaseClient, alias: str) -> List[Dict[str, Any]]:
    """
    Return one row per hosted zone that the caller can see.

    *RecordTypes* → unique RR-set types in the zone
    *HealthChecks* → every health-check ID referenced by a record set
    """
    out: List[Dict[str, Any]] = []

    zones = _safe_aws_call(
        r53_client.list_hosted_zones,
        default={"HostedZones": []},
        account=alias,
    )["HostedZones"]

    for z in zones:
        zid = z["Id"].split("/")[-1]

        # ── zone details -------------------------------------------------
        det = _safe_aws_call(
            r53_client.get_hosted_zone,
            default={},
            account=alias,
            Id=zid,
        )
        vpcs = [
            {"VPCRegion": v["VPCRegion"], "VPCId": v["VPCId"]}
            for v in det.get("VPCs", [])
        ]
        ns = det.get("DelegationSet", {}).get("NameServers", [])

        sec = _safe_aws_call(
            r53_client.get_dnssec,
            default={},
            account=alias,
            HostedZoneId=zid,
        ).get("Status", {})
        dnssec = f'{sec.get("Status","")}/{sec.get("ServeSignature","")}'.strip("/")

        tags = {
            t["Key"]: t["Value"]
            for t in _safe_aws_call(
                r53_client.list_tags_for_resource,
                default={"ResourceTagSet": {"Tags": []}},
                account=alias,
                ResourceType="hostedzone",
                ResourceId=zid,
            )["ResourceTagSet"]["Tags"]
        }

        # ── NEW: scan RR-sets once for types & health-checks -------------
        record_types: set[str] = set()
        health_checks: set[str] = set()

        for page in _safe_paginator(
            r53_client.get_paginator("list_resource_record_sets").paginate,
            account=alias,
            HostedZoneId=zid,
        ):
            for rr in page.get("ResourceRecordSets", []):

                record_types.add(rr.get("Type", ""))
                if "HealthCheckId" in rr:
                    health_checks.add(rr["HealthCheckId"])

        out.append(
            {
                "Name": z.get("Name", "").rstrip("."),
                "Id": z["Id"],
                "Config": z.get("Config", {}),
                "ResourceRecordSetCount": z.get("ResourceRecordSetCount", 0),
                "RecordTypes": sorted(record_types),
                "Tags": tags,
                "VPCAssociations": vpcs,
                "HealthChecks": sorted(health_checks),
                "DNSSECStatus": dnssec,
                "DelegationSet": ns,
                "Region": "global",
                "AccountAlias": alias,
            }
        )

    return out


# --------------------------------------------------------------------------
# KMS COLLECTOR
# --------------------------------------------------------------------------
def _collect_key_aliases(kms: BaseClient, key_id: str, alias: str) -> List[str]:
    aliases: List[str] = []
    for page in _safe_paginator(
        kms.get_paginator("list_aliases").paginate, account=alias, KeyId=key_id
    ):
        aliases.extend(
            a["AliasName"]
            for a in page.get("Aliases", [])
            if a.get("AliasName", "").startswith("alias/")
        )
    return aliases


def get_kms_details(kms_client: BaseClient, alias: str) -> List[Dict[str, Any]]:
    """
    Return metadata for every KMS key the caller can see in *kms_client.region*.

    • For **AWS-managed keys** (`KeyManager == "AWS"`) we skip the calls that
      always fail (`ListResourceTags`, `GetKeyRotationStatus`, `ListGrants`),
      so no AccessDenied warnings are emitted.
    """
    out: List[Dict[str, Any]] = []

    for page in _safe_paginator(
        kms_client.get_paginator("list_keys").paginate, account=alias
    ):
        for key in page.get("Keys", []):
            key_id = key["KeyId"]

            meta = _safe_aws_call(
                kms_client.describe_key,
                default={},
                account=alias,
                KeyId=key_id,
            ).get("KeyMetadata", {})

            if not meta:
                continue

            aws_managed = meta.get("KeyManager") == "AWS"

            # ── optional look-ups (skip on AWS-managed keys) --------------
            tags = {}  # type: Dict[str, str]
            rotation = False  # type: bool
            grants = 0  # type: int

            if not aws_managed:
                tags = {
                    t["TagKey"]: t["TagValue"]
                    for page in _safe_paginator(
                        kms_client.get_paginator("list_resource_tags").paginate,
                        account=alias,
                        KeyId=key_id,
                    )
                    for t in page.get("Tags", [])
                }

                rotation = _safe_aws_call(
                    kms_client.get_key_rotation_status,
                    default={"KeyRotationEnabled": False},
                    account=alias,
                    KeyId=key_id,
                ).get("KeyRotationEnabled", False)

                grants = sum(
                    len(pg.get("Grants", []))
                    for pg in _safe_paginator(
                        kms_client.get_paginator("list_grants").paginate,
                        account=alias,
                        KeyId=key_id,
                    )
                )

            out.append(
                {
                    "KeyId": key_id,
                    "Description": meta.get("Description", ""),
                    "Enabled": meta.get("Enabled", False),
                    "KeyState": meta.get("KeyState", ""),
                    "KeyManager": meta.get("KeyManager", ""),
                    "KeySpec": meta.get("KeySpec", ""),
                    "KeyUsage": meta.get("KeyUsage", ""),
                    "Origin": meta.get("Origin", ""),
                    "CreationDate": to_local(
                        meta.get("CreationDate"), kms_client.meta.region_name
                    ),
                    "DeletionDate": to_local(
                        meta.get("DeletionDate"), kms_client.meta.region_name
                    ),
                    "ValidTo": to_local(
                        meta.get("ValidTo"), kms_client.meta.region_name
                    ),
                    "MultiRegion": meta.get("MultiRegion", False),
                    "PendingDeletion": meta.get("PendingDeletion", False),
                    "PendingWindowInDays": meta.get("PendingWindowInDays", 0),
                    "AliasNames": _collect_key_aliases(kms_client, key_id, alias),
                    "Tags": tags,
                    "RotationEnabled": rotation,
                    "GrantsCount": grants,
                    "Region": kms_client.meta.region_name,
                    "AccountAlias": alias,
                }
            )
    return out


# --------------------------------------------------------------------------
# ACM COLLECTOR
# --------------------------------------------------------------------------
def get_acm_details(acm_client: BaseClient, alias: str) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for page in _safe_paginator(
        acm_client.get_paginator("list_certificates").paginate, account=alias
    ):
        for cs in page.get("CertificateSummaryList", []):
            arn = cs["CertificateArn"]
            cert = _safe_aws_call(
                acm_client.describe_certificate,
                default={},
                account=alias,
                CertificateArn=arn,
            ).get("Certificate", {})
            if not cert:
                continue
            out.append(
                {
                    "DomainName": cert.get("DomainName", ""),
                    "CertificateArn": arn,
                    "Status": cert.get("Status", ""),
                    "Type": cert.get("Type", ""),
                    "InUse": bool(cert.get("InUseBy", [])),
                    "Issued": to_local(
                        cert.get("IssuedAt"), acm_client.meta.region_name
                    ),
                    "Expires": to_local(
                        cert.get("NotAfter"), acm_client.meta.region_name
                    ),
                    "RenewalEligibility": cert.get("RenewalEligibility", ""),
                    "SubjectAlternativeNames": cert.get("SubjectAlternativeNames", []),
                    "Region": acm_client.meta.region_name,
                    "AccountAlias": alias,
                }
            )
    return out


# --------------------------------------------------------------------------
# EventBridge COLLECTOR
# --------------------------------------------------------------------------
def get_eventbridge_details(events: BaseClient, alias: str) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    buses = [
        b
        for page in _safe_paginator(
            require_paginator(events, "list_event_buses").paginate,
            account=alias,
        )
        for b in page.get("EventBuses", [])
    ]
    for bus in buses:
        name = bus.get("Name", "")
        for page in _safe_paginator(
            events.get_paginator("list_rules").paginate,
            account=alias,
            EventBusName=name,
        ):
            for rule in page.get("Rules", []):
                expr = rule.get("ScheduleExpression", "")
                if not expr:
                    continue
                freq, det = humanise_schedule(expr)
                targ = [
                    t
                    for page in _safe_paginator(
                        require_paginator(events, "list_targets_by_rule").paginate,
                        account=alias,
                        Rule=rule["Name"],
                        EventBusName=name,
                    )
                    for t in page.get("Targets", [])
                ]
                first = targ[0] if targ else {}
                out.append(
                    {
                        "ScheduleName": rule["Name"],
                        "GroupName": name,
                        "State": rule.get("State", ""),
                        "Expression": expr,
                        "Timezone": _tz_for(events.meta.region_name),
                        "Frequency": freq,
                        "Details": det,
                        "TargetArn": first.get("Arn", ""),
                        "Input": first.get("Input", ""),
                        "Region": events.meta.region_name,
                        "AccountAlias": alias,
                    }
                )
    return out


def get_eventbridge_scheduler_details(
    scheduler: BaseClient, alias: str
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    groups = {"default"} | {
        g["Name"]
        for page in _safe_paginator(
            scheduler.get_paginator("list_schedule_groups").paginate, account=alias
        )
        for g in page.get("ScheduleGroups", [])
    }
    for grp in groups:
        for page in _safe_paginator(
            scheduler.get_paginator("list_schedules").paginate,
            account=alias,
            GroupName=grp,
        ):
            for s in page.get("Schedules", []):
                name = s.get("Name", "")
                det = _safe_aws_call(
                    scheduler.get_schedule,
                    default={},
                    account=alias,
                    Name=name,
                    GroupName=grp,
                )
                expr = det.get("ScheduleExpression", "")
                freq, desc = humanise_schedule(expr)
                tz = det.get(
                    "ScheduleExpressionTimezone",
                    REGION_TZ.get(scheduler.meta.region_name, "UTC"),
                )
                tgt = det.get("Target", {})
                out.append(
                    {
                        "ScheduleName": name,
                        "GroupName": grp,
                        "State": det.get("State", ""),
                        "Expression": expr,
                        "Timezone": tz,
                        "Frequency": freq,
                        "Details": desc,
                        "TargetArn": tgt.get("Arn", ""),
                        "Input": tgt.get("Input", ""),
                        "Region": scheduler.meta.region_name,
                        "AccountAlias": alias,
                    }
                )
    return out


# --------------------------------------------------------------------------
# AWS Backup COLLECTOR
# --------------------------------------------------------------------------
def get_backup_details(backup_client: BaseClient, alias: str) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    plans = _safe_aws_call(
        backup_client.list_backup_plans, default={"BackupPlansList": []}, account=alias
    )["BackupPlansList"]
    for summary in plans:
        pid, vid = summary["BackupPlanId"], summary["VersionId"]
        created = to_local(summary.get("CreationDate"), backup_client.meta.region_name)
        plan = _safe_aws_call(
            backup_client.get_backup_plan,
            default={"BackupPlan": {}},
            account=alias,
            BackupPlanId=pid,
            VersionId=vid,
        )["BackupPlan"]
        if not plan:
            continue

        sels = _safe_aws_call(
            backup_client.list_backup_selections,
            default={"BackupSelectionsList": []},
            account=alias,
            BackupPlanId=pid,
        )["BackupSelectionsList"]
        sel_map: Dict[str, Any] = {}
        for sel in sels:
            sid = sel["SelectionId"]
            cfg = _safe_aws_call(
                backup_client.get_backup_selection,
                default={"BackupSelection": {}},
                account=alias,
                BackupPlanId=pid,
                SelectionId=sid,
            )["BackupSelection"]
            sel_map[sid] = cfg

        for rule in plan.get("Rules", []):
            expr = rule.get("ScheduleExpression", "")
            freq, det = humanise_schedule(expr)
            tz = rule.get(
                "ScheduleExpressionTimezone",
                REGION_TZ.get(backup_client.meta.region_name, "UTC"),
            )
            vault = rule.get("TargetBackupVaultName", "")
            jobs = _safe_aws_call(
                backup_client.list_backup_jobs,
                default={"BackupJobs": []},
                account=alias,
                ByBackupVaultName=vault,
                ByState="COMPLETED",
                MaxResults=1,
            )["BackupJobs"]
            last = jobs[0] if jobs else {}
            last_exec = (
                to_local(
                    last.get("CompletionDate") or last.get("CreationDate"),
                    backup_client.meta.region_name,
                )
                if last
                else ""
            )

            for cfg in sel_map.values():
                out.append(
                    {
                        "PlanName": plan.get("BackupPlanName", ""),
                        "PlanId": pid,
                        "PlanArn": summary.get("BackupPlanArn", ""),
                        "PlanCreationDate": created,
                        "RuleName": rule.get("RuleName", ""),
                        "Schedule": freq,
                        "Details": det,
                        "Timezone": tz,
                        "LastExecutionDate": last_exec,
                        "VaultName": vault,
                        "SelectionName": cfg.get("SelectionName", ""),
                        "IamRole": cfg.get("IamRoleArn", ""),
                        "Resources": cfg.get("Resources", []),
                        "ResourceTags": cfg.get("ListOfTags", []),
                        "Region": backup_client.meta.region_name,
                        "AccountAlias": alias,
                    }
                )
    return out


# --------------------------------------------------------------------------
# WAFv2 & CLASSIC COLLECTORS
# --------------------------------------------------------------------------
def get_waf_v2_details(wafv2_client: BaseClient, alias: str) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    params = {"Scope": "REGIONAL"}
    # manual pagination
    while True:
        resp = _safe_aws_call(
            wafv2_client.list_web_acls, default={"WebACLs": []}, account=alias, **params
        )
        acls = resp.get("WebACLs", [])
        for acl in acls:
            full = _safe_aws_call(
                wafv2_client.get_web_acl,
                default={"WebACL": {}},
                account=alias,
                Name=acl["Name"],
                Scope="REGIONAL",
                Id=acl["Id"],
            )["WebACL"]
            rules = full.get("Rules", [])
            out.append(
                {
                    "Name": acl["Name"],
                    "WebACLId": acl["Id"],
                    "RuleCount": len(rules),
                    "Rules": [r.get("Name") for r in rules],
                    "Region": wafv2_client.meta.region_name,
                    "AccountAlias": alias,
                }
            )
        marker = resp.get("NextMarker")
        if not marker:
            break
        params["NextMarker"] = marker
    return out


def get_waf_classic_details(
    wafc_client: BaseClient, alias: str
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    params: Dict[str, Any] = {}
    while True:
        resp = _safe_aws_call(
            wafc_client.list_web_acls, default={"WebACLs": []}, account=alias, **params
        )
        for acl in resp.get("WebACLs", []):
            det = _safe_aws_call(
                wafc_client.get_web_acl,
                default={"WebACL": {}},
                account=alias,
                WebACLId=acl["WebACLId"],
            )["WebACL"]
            rules = det.get("Rules", [])
            out.append(
                {
                    "Name": det.get("Name", ""),
                    "WebACLId": acl["WebACLId"],
                    "RuleCount": len(rules),
                    "Rules": [r.get("RuleId", "") for r in rules],
                    "Region": wafc_client.meta.region_name,
                    "AccountAlias": alias,
                }
            )
        marker = resp.get("NextMarker")
        if not marker:
            break
        params["NextMarker"] = marker
    return out


# --------------------------------------------------------------------------
# ALB / NLB COLLECTOR
# --------------------------------------------------------------------------
def get_alb_details(elbv2_client: BaseClient, alias: str) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    lbs = [
        lb
        for page in _safe_paginator(
            elbv2_client.get_paginator("describe_load_balancers").paginate,
            account=alias,
        )
        for lb in page.get("LoadBalancers", [])
    ]
    for chunk in (lbs[i:i + 20] for i in range(0, len(lbs), 20)):
        arns = [lb["LoadBalancerArn"] for lb in chunk]
        tags = _safe_aws_call(
            elbv2_client.describe_tags,
            default={"TagDescriptions": []},
            account=alias,
            ResourceArns=arns,
        )["TagDescriptions"]
        tag_map = {
            td["ResourceArn"]: {t["Key"]: t["Value"] for t in td.get("Tags", [])}
            for td in tags
        }

        for lb in chunk:
            arn = lb["LoadBalancerArn"]
            listeners = [
                {"Port": listener["Port"], "Protocol": listener["Protocol"]}
                for page in _safe_paginator(
                    elbv2_client.get_paginator("describe_listeners").paginate,
                    account=alias,
                    LoadBalancerArn=arn,
                )
                for listener in page.get("Listeners", [])
            ]
            tgs = _safe_aws_call(
                elbv2_client.describe_target_groups,
                default={"TargetGroups": []},
                account=alias,
                LoadBalancerArn=arn,
            )["TargetGroups"]
            out.append(
                {
                    "LoadBalancerName": lb["LoadBalancerName"],
                    "DNSName": lb.get("DNSName", ""),
                    "Type": lb.get("Type", ""),
                    "Scheme": lb.get("Scheme", ""),
                    "VpcId": lb.get("VpcId", ""),
                    "State": lb.get("State", {}).get("Code", ""),
                    "AvailabilityZones": [
                        az.get("ZoneName", "") for az in lb.get("AvailabilityZones", [])
                    ],
                    "SecurityGroups": lb.get("SecurityGroups", []),
                    "Tags": tag_map.get(arn, {}),
                    "Listeners": listeners,
                    "TargetGroups": [tg["TargetGroupName"] for tg in tgs],
                    "Region": elbv2_client.meta.region_name,
                    "AccountAlias": alias,
                }
            )
    return out


# --------------------------------------------------------------------------
# LAMBDA COLLECTOR
# --------------------------------------------------------------------------
def get_lambda_details(lambda_client: BaseClient, alias: str) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []

    for page in _safe_paginator(
        lambda_client.get_paginator("list_functions").paginate, account=alias
    ):
        for fn in page.get("Functions", []):
            name, arn = fn["FunctionName"], fn["FunctionArn"]

            cfg = _safe_aws_call(
                lambda_client.get_function_configuration,
                default={},
                account=alias,
                FunctionName=name,
            )

            state = cfg.get("State", "")
            lastupd = cfg.get("LastUpdateStatus", "")
            kms = cfg.get("KMSKeyArn", "")
            hr_size = human_size(fn.get("CodeSize", 0))

            # one (non-paginated) call for tags is fine
            tag_resp = _safe_aws_call(
                lambda_client.list_tags,
                default={"Tags": {}},
                account=alias,
                Resource=arn,
            )
            tags = tag_resp.get("Tags", {})

            vpc = fn.get("VpcConfig", {})
            env = fn.get("Environment", {}).get("Variables", {})

            out.append(
                {
                    "FunctionName": name,
                    "FunctionArn": arn,
                    "Runtime": fn.get("Runtime", ""),
                    "Handler": fn.get("Handler", ""),
                    "Role": fn.get("Role", ""),
                    "Description": fn.get("Description", ""),
                    "MemorySize": fn.get("MemorySize", 0),
                    "Timeout": fn.get("Timeout", 0),
                    "PackageType": fn.get("PackageType", ""),
                    "Architectures": fn.get("Architectures", []),
                    "TracingMode": cfg.get("TracingConfig", {}).get("Mode", ""),
                    "State": state,
                    "LastUpdateStatus": lastupd,
                    "LastModified": to_local(
                        fn.get("LastModified"),
                        lambda_client.meta.region_name,
                    ),
                    "KMSKeyArn": kms,
                    "CodeSize": hr_size,
                    "VpcSecurityGroupIds": vpc.get("SecurityGroupIds", []),
                    "VpcSubnetIds": vpc.get("SubnetIds", []),
                    "EnvironmentVars": env,
                    "Tags": tags,
                    "Region": lambda_client.meta.region_name,
                    "AccountAlias": alias,
                }
            )
    return out


# --------------------------------------------------------------------------
# Lightsail COLLECTOR (Unified)
# --------------------------------------------------------------------------
def get_lightsail_inventory(
    ls_client: BaseClient,
    alias: str,
    session: boto3.Session,
    is_primary_region: bool,
    supported_regions: set[str],
) -> List[Dict[str, Any]]:
    """
    Return a unified list of all regional and global Lightsail resources,
    including detailed bundle specs and firewall rules for instances.
    """
    out: List[Dict[str, Any]] = []
    current_region = ls_client.meta.region_name

    if supported_regions and current_region not in supported_regions:
        return []

    # log("info", f"Running Lightsail collector in supported region: {current_region}", account=alias)

    # --- Pre-fetch Bundle Specifications for the region ---
    bundle_specs = {}
    try:
        # Fetch Instance Bundles
        for page in _safe_paginator(
            ls_client.get_paginator("get_bundles").paginate, account=alias
        ):
            for bundle in page.get("bundles", []):
                bundle_specs[bundle["bundleId"]] = {
                    "vCPUs": bundle.get("cpuCount"),
                    "MemoryInGB": bundle.get("ramSizeInGb"),
                    "DiskSizeGB": bundle.get("diskSizeInGb"),
                    "DataTransferGB": bundle.get("transferPerMonthInGb"),
                }
        # Fetch Database Bundles
        for page in _safe_paginator(
            ls_client.get_paginator("get_relational_database_bundles").paginate,
            account=alias,
        ):
            for bundle in page.get("bundles", []):
                bundle_specs[bundle["bundleId"]] = {
                    "vCPUs": bundle.get("cpuCount"),
                    "MemoryInGB": bundle.get("ramSizeInGb"),
                    "DiskSizeGB": bundle.get("diskSizeInGb"),
                }
    except Exception as e:
        log(
            "warning",
            f"Could not fetch all Lightsail bundle specs in {current_region}: {e}",
            account=alias,
        )

    # --- Regional Resources ---
    # Instances
    for page in _safe_paginator(
        ls_client.get_paginator("get_instances").paginate, account=alias
    ):
        for inst in page.get("instances", []):
            bundle_id = inst.get("bundleId")
            specs = bundle_specs.get(bundle_id, {})
            firewall_rules = []
            try:
                port_states = ls_client.get_instance_port_states(
                    instanceName=inst["name"]
                ).get("portStates", [])
                firewall_rules = port_states
            except Exception as e:
                log(
                    "warning",
                    f"Could not fetch firewall rules for instance {inst.get('name')}: {e}",
                    account=alias,
                )

            out.append(
                {
                    "Name": inst.get("name"),
                    "ResourceType": "Instance",
                    "Arn": inst.get("arn"),
                    "State": inst.get("state", {}).get("name"),
                    "Region": current_region,
                    "Location": inst.get("location", {}).get("availabilityZone"),
                    "BlueprintOrEngine": inst.get("blueprintName"),
                    "BundleId": bundle_id,
                    "PublicIpOrDnsName": inst.get("publicIpAddress"),
                    "SshKeyName": inst.get("sshKeyName"),
                    "Username": inst.get("username"),
                    "IpAddressType": inst.get("ipAddressType"),
                    "vCPUs": specs.get("vCPUs"),
                    "MemoryInGB": specs.get("MemoryInGB"),
                    "DiskSizeGB": specs.get("DiskSizeGB"),
                    "DataTransferGB": specs.get("DataTransferGB"),
                    "CreatedAt": to_local(inst.get("createdAt"), current_region),
                    "FirewallRules": firewall_rules,
                    "AccountAlias": alias,
                }
            )

    # Databases
    for page in _safe_paginator(
        ls_client.get_paginator("get_relational_databases").paginate, account=alias
    ):
        for db in page.get("relationalDatabases", []):
            bundle_id = db.get("relationalDatabaseBundleId")
            specs = bundle_specs.get(bundle_id, {})
            out.append(
                {
                    "Name": db.get("name"),
                    "ResourceType": "Database",
                    "Arn": db.get("arn"),
                    "State": db.get("state"),
                    "Region": current_region,
                    "Location": db.get("location", {}).get("availabilityZone"),
                    "BlueprintOrEngine": db.get("relationalDatabaseBlueprintId"),
                    "BundleId": bundle_id,
                    "vCPUs": specs.get("vCPUs"),
                    "MemoryInGB": specs.get("MemoryInGB"),
                    "DiskSizeGB": specs.get("DiskSizeGB"),
                    "CreatedAt": to_local(db.get("createdAt"), current_region),
                    "AccountAlias": alias,
                }
            )

    # Disks (Block Storage)
    for page in _safe_paginator(
        ls_client.get_paginator("get_disks").paginate, account=alias
    ):
        for disk in page.get("disks", []):
            out.append(
                {
                    "Name": disk.get("name"),
                    "ResourceType": "Disk",
                    "Arn": disk.get("arn"),
                    "State": disk.get("state"),
                    "Region": current_region,
                    "Location": disk.get("location", {}).get("availabilityZone"),
                    "SizeInGb": disk.get("sizeInGb"),
                    "AttachedTo": disk.get("attachedTo"),
                    "CreatedAt": to_local(disk.get("createdAt"), current_region),
                    "AccountAlias": alias,
                }
            )

    # Static IPs
    for page in _safe_paginator(
        ls_client.get_paginator("get_static_ips").paginate, account=alias
    ):
        for ip in page.get("staticIps", []):
            out.append(
                {
                    "Name": ip.get("name"),
                    "ResourceType": "StaticIp",
                    "Arn": ip.get("arn"),
                    "State": "N/A",
                    "Region": current_region,
                    "Location": ip.get("location", {}).get("availabilityZone"),
                    "PublicIpOrDnsName": ip.get("ipAddress"),
                    "AttachedTo": ip.get("attachedTo"),
                    "CreatedAt": to_local(ip.get("createdAt"), current_region),
                    "AccountAlias": alias,
                }
            )

    # Certificates
    cert_resp = _safe_aws_call(ls_client.get_certificates, account=alias, default={})
    for cert in cert_resp.get("certificates", []):
        out.append(
            {
                "Name": cert.get("name"),
                "ResourceType": "Certificate",
                "Arn": cert.get("arn"),
                "State": cert.get("status"),
                "Region": current_region,
                "PublicIpOrDnsName": cert.get("domainName"),
                "CreatedAt": to_local(cert.get("createdAt"), current_region),
                "ExpiresAt": to_local(cert.get("notAfter"), current_region),
                "AccountAlias": alias,
            }
        )

    # --- Global Resources (fetched only once from the primary region) ---
    if is_primary_region:
        # log("info", "Fetching global Lightsail Domains from us-east-1", account=alias)
        try:
            global_ls_client = aws_client("lightsail", "us-east-1", session)
            for page in _safe_paginator(
                global_ls_client.get_paginator("get_domains").paginate, account=alias
            ):
                for domain in page.get("domains", []):
                    out.append(
                        {
                            "Name": domain.get("name"),
                            "ResourceType": "Domain",
                            "Arn": domain.get("arn"),
                            "Region": "global",
                            "CreatedAt": to_local(domain.get("createdAt"), "us-east-1"),
                            "AccountAlias": alias,
                        }
                    )
        except Exception as e:
            log("error", f"Failed to fetch Lightsail domains: {e}", account=alias)

    # if not out and not is_primary_region:
    #     log(
    #         "info",
    #         f"No regional Lightsail resources found in {current_region}",
    #         account=alias,
    #     )

    return out


# --------------------------------------------------------------------------
# CloudWatch Logs & Alarms COLLECTOR
# --------------------------------------------------------------------------
def get_cloudwatch_logs_details(
    logs_client: BaseClient, alias: str
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for g in [
        g
        for page in _safe_paginator(
            logs_client.get_paginator("describe_log_groups").paginate, account=alias
        )
        for g in page.get("logGroups", [])
    ]:
        created = datetime.fromtimestamp(g.get("creationTime", 0) / 1000, timezone.utc)
        mfilters = _safe_aws_call(
            logs_client.describe_metric_filters,
            default={"metricFilters": []},
            account=alias,
            logGroupName=g["logGroupName"],
        )["metricFilters"]
        out.append(
            {
                "LogGroupName": g["logGroupName"],
                "RetentionInDays": g.get("retentionInDays", ""),
                "Size": human_size(g.get("storedBytes")),
                "CreationTime": to_local(created, logs_client.meta.region_name),
                "MetricFilterCount": len(mfilters),
                "Region": logs_client.meta.region_name,
                "AccountAlias": alias,
            }
        )
    return out


def get_cloudwatch_alarms_details(
    cw_client: BaseClient, alias: str
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for page in _safe_paginator(
        cw_client.get_paginator("describe_alarms").paginate, account=alias
    ):
        for a in page.get("MetricAlarms", []):
            out.append(
                {
                    "AlarmName": a["AlarmName"],
                    "AlarmType": "Metric",
                    "MetricName": a.get("MetricName", ""),
                    "Namespace": a.get("Namespace", ""),
                    "Statistic": a.get("Statistic", ""),
                    "Period": a.get("Period", ""),
                    "Threshold": a.get("Threshold", ""),
                    "ComparisonOperator": a.get("ComparisonOperator", ""),
                    "EvaluationPeriods": a.get("EvaluationPeriods", ""),
                    "DatapointsToAlarm": a.get("DatapointsToAlarm", ""),
                    "ActionsEnabled": a.get("ActionsEnabled", False),
                    "AlarmActions": a.get("AlarmActions", []),
                    "InsufficientDataActions": a.get("InsufficientDataActions", []),
                    "OKActions": a.get("OKActions", []),
                    "Region": cw_client.meta.region_name,
                    "AccountAlias": alias,
                }
            )
        for c in page.get("CompositeAlarms", []):
            out.append(
                {
                    "AlarmName": c["AlarmName"],
                    "AlarmType": "Composite",
                    "MetricName": "",
                    "Namespace": "",
                    "Statistic": "",
                    "Period": "",
                    "Threshold": "",
                    "ComparisonOperator": "",
                    "EvaluationPeriods": "",
                    "DatapointsToAlarm": "",
                    "ActionsEnabled": c.get("ActionsEnabled", False),
                    "AlarmActions": c.get("AlarmActions", []),
                    "InsufficientDataActions": c.get("InsufficientDataActions", []),
                    "OKActions": c.get("OKActions", []),
                    "Region": cw_client.meta.region_name,
                    "AccountAlias": alias,
                }
            )
    return out


# --------------------------------------------------------------------------
# IAM COLLECTOR (Unified)
# --------------------------------------------------------------------------
def get_iam_details(
    iam_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """
    Return a dictionary containing details for all IAM resources (Users,
    Roles, Groups, and Policies) for the account.
    """
    iam_resources = {
        "IAMUsers": [],
        "IAMRoles": [],
        "IAMGroups": [],
        "IAMPolicies": [],
    }

    # log("info", "Fetching all IAM resources", account=alias)

    # --- IAM Users ---
    for page in _safe_paginator(
        iam_client.get_paginator("list_users").paginate, account=alias
    ):
        for user in page.get("Users", []):
            user_name = user["UserName"]
            user_summary = {
                "UserName": user_name,
                "UserId": user["UserId"],
                "Arn": user["Arn"],
                "CreateDate": to_local(
                    user.get("CreateDate"), iam_client.meta.region_name
                ),
                "PasswordLastUsed": to_local(
                    user.get("PasswordLastUsed"), iam_client.meta.region_name
                ),
                "Groups": [
                    g["GroupName"]
                    for g in iam_client.list_groups_for_user(UserName=user_name).get(
                        "Groups", []
                    )
                ],
                "AttachedPolicies": [
                    p["PolicyArn"]
                    for p in iam_client.list_attached_user_policies(
                        UserName=user_name
                    ).get("AttachedPolicies", [])
                ],
                "InlinePolicies": sorted(
                    set(
                        p
                        for page in _safe_paginator(
                            require_paginator(iam_client, "list_user_policies").paginate,
                            account=alias,
                            UserName=user_name,
                        )
                        for p in page.get("PolicyNames", [])
                    )
                ),
                "AccountAlias": alias,
            }
            iam_resources["IAMUsers"].append(user_summary)

    # --- IAM Roles ---
    for page in _safe_paginator(
        iam_client.get_paginator("list_roles").paginate, account=alias
    ):
        for role in page.get("Roles", []):
            role_name = role["RoleName"]
            trust_policy = role.get("AssumeRolePolicyDocument", {})
            service_principals: set[str] = set()
            account_principals: set[str] = set()
            federated_principals: set[str] = set()
            if trust_policy:
                for stmt in trust_policy.get("Statement", []):
                    principal = stmt.get("Principal", {})
                    if principal == "*":
                        account_principals.add("*")
                        continue
                    if isinstance(principal, str):
                        account_principals.add(principal)
                        continue
                    if "Service" in principal:
                        services = principal["Service"]
                        service_principals.update(
                            services if isinstance(services, list) else [services]
                        )
                    if "AWS" in principal:
                        aws_accounts = principal["AWS"]
                        account_principals.update(
                            aws_accounts if isinstance(aws_accounts, list) else [aws_accounts]
                        )
                    if "Federated" in principal:
                        federated = principal["Federated"]
                        federated_principals.update(
                            federated if isinstance(federated, list) else [federated]
                        )

            all_principals = service_principals | account_principals | federated_principals

            role_summary = {
                "RoleName": role_name,
                "RoleId": role["RoleId"],
                "Arn": role["Arn"],
                "CreateDate": to_local(
                    role.get("CreateDate"), iam_client.meta.region_name
                ),
                "ServicePrincipals": sorted(all_principals),
                "AccountPrincipals": sorted(account_principals),
                "FederatedPrincipals": sorted(federated_principals),
                "AttachedPolicies": [
                    p["PolicyArn"]
                    for p in iam_client.list_attached_role_policies(
                        RoleName=role_name
                    ).get("AttachedPolicies", [])
                ],
                "InlinePolicies": sorted(
                    set(
                        p
                        for page in _safe_paginator(
                            require_paginator(iam_client, "list_role_policies").paginate,
                            account=alias,
                            RoleName=role_name,
                        )
                        for p in page.get("PolicyNames", [])
                    )
                ),
                "AccountAlias": alias,
            }
            iam_resources["IAMRoles"].append(role_summary)

    # --- IAM Groups ---
    for page in _safe_paginator(
        iam_client.get_paginator("list_groups").paginate, account=alias
    ):
        for group in page.get("Groups", []):
            group_name = group["GroupName"]
            group_summary = {
                "GroupName": group_name,
                "GroupId": group["GroupId"],
                "Arn": group["Arn"],
                "CreateDate": to_local(
                    group.get("CreateDate"), iam_client.meta.region_name
                ),
                "AttachedPolicies": [
                    p["PolicyArn"]
                    for p in iam_client.list_attached_group_policies(
                        GroupName=group_name
                    ).get("AttachedPolicies", [])
                ],
                "InlinePolicies": sorted(
                    set(
                        p
                        for page in _safe_paginator(
                            require_paginator(iam_client, "list_group_policies").paginate,
                            account=alias,
                            GroupName=group_name,
                        )
                        for p in page.get("PolicyNames", [])
                    )
                ),
                "Members": [
                    u["UserName"]
                    for u in iam_client.get_group(GroupName=group_name).get("Users", [])
                ],
                "AccountAlias": alias,
            }
            iam_resources["IAMGroups"].append(group_summary)

    # --- IAM Policies ---
    paginator = iam_client.get_paginator("list_policies")
    for scope in ["Local", "AWS"]:
        policy_type = "Customer Managed" if scope == "Local" else "AWS Managed"
        for page in paginator.paginate(Scope=scope):
            for policy in page.get("Policies", []):
                if scope == "AWS" and policy.get("AttachmentCount", 0) == 0:
                    continue

                service_category = ""
                if scope == "AWS":
                    name = policy["PolicyName"]
                    if name.startswith("Amazon"):
                        name = name[6:]
                    elif name.startswith("AWS"):
                        name = name[3:]
                    for prefix, service_name in SERVICE_PREFIXES.items():
                        if name.startswith(prefix):
                            service_category = service_name
                            break
                    if not service_category:
                        arn_parts = policy["Arn"].split("/")
                        if len(arn_parts) > 1 and "service-role" in arn_parts:
                            role_part = arn_parts[-1]
                            for prefix, service_name in SERVICE_PREFIXES.items():
                                if prefix in role_part:
                                    service_category = service_name
                                    break

                policy_summary = {
                    "PolicyName": policy["PolicyName"],
                    "PolicyId": policy["PolicyId"],
                    "Arn": policy["Arn"],
                    "AttachmentCount": policy.get("AttachmentCount", 0),
                    "DefaultVersionId": policy["DefaultVersionId"],
                    "PolicyType": policy_type,
                    "ServiceCategory": service_category,
                    "AttachmentEntities": [],
                    "CreateDate": to_local(
                        policy.get("CreateDate"), iam_client.meta.region_name
                    ),
                    "UpdateDate": to_local(
                        policy.get("UpdateDate"), iam_client.meta.region_name
                    ),
                    "PolicyDocument": {},
                    "AccountAlias": alias,
                }

                try:
                    ver = iam_client.get_policy_version(
                        PolicyArn=policy["Arn"], VersionId=policy["DefaultVersionId"]
                    )
                    if scope == "AWS":
                        actions = []
                        for stmt in ver["PolicyVersion"]["Document"].get(
                            "Statement", []
                        ):
                            if "Action" in stmt:
                                actions.extend(
                                    stmt["Action"]
                                    if isinstance(stmt["Action"], list)
                                    else [stmt["Action"]]
                                )
                        policy_summary["ActionPatterns"] = list(set(actions))
                    else:
                        policy_summary["PolicyDocument"] = ver["PolicyVersion"][
                            "Document"
                        ]
                except Exception:
                    pass

                attached_entities = iam_client.list_entities_for_policy(
                    PolicyArn=policy["Arn"]
                )
                for entity_type_plural in (
                    "PolicyGroups",
                    "PolicyUsers",
                    "PolicyRoles",
                ):
                    entity_type_singular = entity_type_plural.replace("Policy", "")[:-1]
                    name_key = f"{entity_type_singular}Name"
                    for entity in attached_entities.get(entity_type_plural, []):
                        policy_summary["AttachmentEntities"].append(
                            {"Type": entity_type_singular, "Name": entity[name_key]}
                        )

                iam_resources["IAMPolicies"].append(policy_summary)

    return iam_resources


# --------------------------------------------------------------------------
# VPC INVENTORY COLLECTOR (Unified)
# --------------------------------------------------------------------------
def get_vpc_inventory(
    ec2_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """
    Return a dictionary containing details for all core VPC resources.
    """
    vpc_resources = {
        "VPCs": [],
        "Subnets": [],
        "RouteTables": [],
        "SecurityGroups": [],
    }
    region = ec2_client.meta.region_name
    # log("info", f"Fetching all VPC resources in {region}", account=alias)

    # --- VPCs ---
    for page in _safe_paginator(
        ec2_client.get_paginator("describe_vpcs").paginate, account=alias
    ):
        for vpc in page.get("Vpcs", []):
            tags = {t["Key"]: t["Value"] for t in vpc.get("Tags", [])}
            vpc_resources["VPCs"].append(
                {
                    "VpcId": vpc.get("VpcId"),
                    "State": vpc.get("State"),
                    "IsDefault": vpc.get("IsDefault"),
                    "CidrBlock": vpc.get("CidrBlock"),
                    "Tags": tags,
                    "Region": region,
                    "AccountAlias": alias,
                }
            )

    # --- Subnets ---
    for page in _safe_paginator(
        ec2_client.get_paginator("describe_subnets").paginate, account=alias
    ):
        for subnet in page.get("Subnets", []):
            tags = {t["Key"]: t["Value"] for t in subnet.get("Tags", [])}
            vpc_resources["Subnets"].append(
                {
                    "SubnetId": subnet.get("SubnetId"),
                    "VpcId": subnet.get("VpcId"),
                    "State": subnet.get("State"),
                    "AvailabilityZone": subnet.get("AvailabilityZone"),
                    "CidrBlock": subnet.get("CidrBlock"),
                    "AvailableIpAddressCount": subnet.get("AvailableIpAddressCount"),
                    "MapPublicIpOnLaunch": subnet.get("MapPublicIpOnLaunch"),
                    "Tags": tags,
                    "Region": region,
                    "AccountAlias": alias,
                }
            )

    # --- Route Tables ---
    for page in _safe_paginator(
        ec2_client.get_paginator("describe_route_tables").paginate, account=alias
    ):
        for rt in page.get("RouteTables", []):
            tags = {t["Key"]: t["Value"] for t in rt.get("Tags", [])}
            is_main = any(a.get("Main", False) for a in rt.get("Associations", []))
            vpc_resources["RouteTables"].append(
                {
                    "RouteTableId": rt.get("RouteTableId"),
                    "VpcId": rt.get("VpcId"),
                    "IsMain": is_main,
                    "Routes": rt.get("Routes"),
                    "Associations": rt.get("Associations"),
                    "Tags": tags,
                    "Region": region,
                    "AccountAlias": alias,
                }
            )

    # --- Security Groups ---
    for page in _safe_paginator(
        ec2_client.get_paginator("describe_security_groups").paginate, account=alias
    ):
        for sg in page.get("SecurityGroups", []):
            tags = {t["Key"]: t["Value"] for t in sg.get("Tags", [])}
            vpc_resources["SecurityGroups"].append(
                {
                    "GroupId": sg.get("GroupId"),
                    "GroupName": sg.get("GroupName"),
                    "VpcId": sg.get("VpcId"),
                    "Description": sg.get("Description"),
                    "IngressRules": sg.get("IpPermissions"),
                    "EgressRules": sg.get("IpPermissionsEgress"),
                    "Tags": tags,
                    "Region": region,
                    "AccountAlias": alias,
                }
            )

    return vpc_resources


# --------------------------------------------------------------------------
# GOVERNANCE COLLECTOR
# --------------------------------------------------------------------------
def get_governance_details(
    sess: boto3.Session, alias: str, region: str
) -> List[Dict[str, Any]]:
    """Checks the status of key governance and security services."""
    out: List[Dict[str, Any]] = []

    # --- Security Hub ---
    try:
        sh_client = aws_client("securityhub", region, sess)
        sh_client.describe_hub()
        out.append(
            {
                "Service": "Security Hub",
                "Status": "Enabled",
                "Details": "Security Hub is enabled for this region.",
            }
        )
    except ClientError as e:
        if e.response["Error"]["Code"] == "InvalidAccessException":
            out.append(
                {"Service": "Security Hub", "Status": "Not Enabled", "Details": ""}
            )
        else:
            out.append(
                {"Service": "Security Hub", "Status": "Error", "Details": str(e)}
            )
    except Exception as e:
        out.append(
            {
                "Service": "Security Hub",
                "Status": "Error",
                "Details": f"Region may not be supported or other error: {e}",
            }
        )

    # --- GuardDuty ---
    try:
        gd_client = aws_client("guardduty", region, sess)
        detectors = gd_client.list_detectors().get("DetectorIds", [])
        if detectors:
            status = gd_client.get_detector(DetectorId=detectors[0]).get(
                "Status", "UNKNOWN"
            )
            out.append(
                {
                    "Service": "GuardDuty",
                    "Status": status.title(),
                    "Details": f"DetectorId: {detectors[0]}",
                }
            )
        else:
            out.append(
                {
                    "Service": "GuardDuty",
                    "Status": "Not Found",
                    "Details": "No GuardDuty detector found in this region.",
                }
            )
    except Exception as e:
        out.append({"Service": "GuardDuty", "Status": "Error", "Details": str(e)})

    # --- AWS Config ---
    try:
        config_client = aws_client("config", region, sess)
        recorders = config_client.describe_configuration_recorder_status().get(
            "ConfigurationRecordersStatus", []
        )
        if not recorders:
            out.append(
                {
                    "Service": "AWS Config",
                    "Status": "Not Enabled",
                    "Details": "No configuration recorder found.",
                }
            )
        else:
            for recorder in recorders:
                status = "Running" if recorder.get("recording") else "Stopped"
                name = recorder.get("name")
                out.append(
                    {
                        "Service": "AWS Config",
                        "Status": status,
                        "Details": f"Recorder: {name}",
                    }
                )
    except Exception as e:
        out.append({"Service": "AWS Config", "Status": "Error", "Details": str(e)})

    # --- Amazon Inspector ---
    try:
        inspector_client = aws_client("inspector2", region, sess)
        status = inspector_client.get_status().get("status", "UNKNOWN")
        out.append(
            {
                "Service": "Inspector",
                "Status": status.replace("_", " ").title(),
                "Details": "Checks for EC2, ECR, and Lambda vulnerabilities.",
            }
        )
    except Exception as e:
        out.append({"Service": "Inspector", "Status": "Error", "Details": str(e)})

    # --- Global Service Checks (run only once per account in us-east-1) ---
    if region == "us-east-1":
        # --- CloudTrail ---
        try:
            ct_client = aws_client("cloudtrail", region, sess)
            trails = ct_client.describe_trails().get("trailList", [])
            if not trails:
                out.append(
                    {
                        "Service": "CloudTrail",
                        "Status": "No Trails Found",
                        "Details": "A multi-region trail is a security best practice.",
                    }
                )
            else:
                multi_region_trail_exists = False
                for trail in trails:
                    if trail.get("IsMultiRegionTrail"):
                        multi_region_trail_exists = True
                        status = "Enabled (Multi-Region)"
                        out.append(
                            {
                                "Service": "CloudTrail",
                                "Status": status,
                                "Details": f"Trail: {trail.get('Name')}",
                            }
                        )
                if not multi_region_trail_exists:
                    out.append(
                        {
                            "Service": "CloudTrail",
                            "Status": "Enabled (Single Region Only)",
                            "Details": "No multi-region trail found.",
                        }
                    )
        except Exception as e:
            out.append({"Service": "CloudTrail", "Status": "Error", "Details": str(e)})

        # --- S3 Block Public Access ---
        try:
            s3_control_client = aws_client("s3control", region, sess)
            acct_id = sess.client("sts").get_caller_identity()["Account"]
            config = s3_control_client.get_public_access_block(AccountId=acct_id).get(
                "PublicAccessBlockConfiguration", {}
            )
            if all(config.values()):
                out.append(
                    {
                        "Service": "S3 Block Public Access",
                        "Status": "Enabled",
                        "Details": "All public access is blocked at the account level.",
                    }
                )
            else:
                out.append(
                    {
                        "Service": "S3 Block Public Access",
                        "Status": "Not Fully Enabled",
                        "Details": "One or more public access settings are disabled.",
                    }
                )
        except Exception as e:
            out.append(
                {
                    "Service": "S3 Block Public Access",
                    "Status": "Error",
                    "Details": str(e),
                }
            )

        # --- AWS Trusted Advisor (via Support Plan) ---
        try:
            support_client = aws_client("support", region, sess)
            severity_levels = support_client.describe_severity_levels().get(
                "severityLevels", []
            )
            if any(level["code"] == "urgent" for level in severity_levels):
                out.append(
                    {
                        "Service": "Trusted Advisor (Support Plan)",
                        "Status": "Business or Enterprise",
                        "Details": "Full Trusted Advisor checks are available.",
                    }
                )
            else:
                out.append(
                    {
                        "Service": "Trusted Advisor (Support Plan)",
                        "Status": "Developer or Basic",
                        "Details": "Limited Trusted Advisor checks are available.",
                    }
                )
        except Exception as e:
            out.append(
                {
                    "Service": "Trusted Advisor (Support Plan)",
                    "Status": "Error",
                    "Details": str(e),
                }
            )

    for item in out:
        item.update({"AccountAlias": alias, "Region": region})
    return out


# --------------------------------------------------------------------------
# COST-OPTIMIZATION COLLECTOR
# --------------------------------------------------------------------------
def get_cost_opportunities(
    ec2_client: BaseClient,
    s3_client: BaseClient,
    elbv2_client: BaseClient,
    cw_client: BaseClient,
    lambda_client: BaseClient,
    rds_client: BaseClient,
    alias: str,
) -> List[Dict[str, Any]]:
    """Identifies potential cost savings opportunities based on AWS best practices."""
    out: List[Dict[str, Any]] = []
    region = ec2_client.meta.region_name
    # log("info", f"Scanning for cost opportunities in {region}", account=alias)

    fourteen_days_ago = datetime.now(timezone.utc) - timedelta(days=14)
    seven_days_ago = datetime.now(timezone.utc) - timedelta(days=7)

    # --- 1. Unattached EBS Volumes ---
    try:
        for page in _safe_paginator(
            ec2_client.get_paginator("describe_volumes").paginate,
            account=alias,
            Filters=[{"Name": "status", "Values": ["available"]}],
        ):
            for vol in page.get("Volumes", []):
                out.append(
                    {
                        "ResourceType": "EBS Volume",
                        "ResourceId": vol["VolumeId"],
                    "Reason": "Unattached (Available)",
                    "Details": f"Size: {vol['Size']} GiB, Type: {vol['VolumeType']}, Created: {vol['CreateTime'].strftime('%Y-%m-%d')}",
                }
            )
    except Exception as e:
        log(
            "warning",
            f"Could not check for unattached EBS volumes in {region}: {e}",
            account=alias,
        )

    # --- 2. Unassociated Elastic IPs ---
    try:
        addresses = ec2_client.describe_addresses().get("Addresses", [])
        for addr in addresses:
            if "AssociationId" not in addr:
                out.append(
                    {
                        "ResourceType": "Elastic IP",
                        "ResourceId": addr["PublicIp"],
                        "Reason": "Unassociated",
                        "Details": f"AllocationId: {addr['AllocationId']}",
                    }
                )
    except Exception as e:
        log(
            "warning",
            f"Could not check for unassociated Elastic IPs in {region}: {e}",
            account=alias,
        )

    # --- 3. Idle Load Balancers ---
    try:
        lbs = elbv2_client.describe_load_balancers().get("LoadBalancers", [])
        for lb in lbs:
            lb_arn = lb["LoadBalancerArn"]
            target_groups = elbv2_client.describe_target_groups(
                LoadBalancerArn=lb_arn
            ).get("TargetGroups", [])
            if not target_groups:
                out.append(
                    {
                        "ResourceType": "Load Balancer",
                        "ResourceId": lb["DNSName"],
                        "Reason": "Idle (No Target Groups)",
                        "Details": f"Name: {lb['LoadBalancerName']}",
                    }
                )
                continue
            is_idle = True
            for tg in target_groups:
                health = elbv2_client.describe_target_health(
                    TargetGroupArn=tg["TargetGroupArn"]
                )
                if any(
                    t.get("TargetHealth", {}).get("State") == "healthy"
                    for t in health.get("TargetHealthDescriptions", [])
                ):
                    is_idle = False
                    break
            if is_idle:
                out.append(
                    {
                        "ResourceType": "Load Balancer",
                        "ResourceId": lb["DNSName"],
                        "Reason": "Idle (No Healthy Targets)",
                        "Details": f"Name: {lb['LoadBalancerName']}",
                    }
                )
    except Exception as e:
        log(
            "warning",
            f"Could not check for idle Load Balancers in {region}: {e}",
            account=alias,
        )

    # --- 4. Old EBS Snapshots ---
    try:
        ninety_days_ago = datetime.now(timezone.utc) - timedelta(days=90)
        paginator = ec2_client.get_paginator("describe_snapshots")
        for page in paginator.paginate(OwnerIds=["self"]):
            for snap in page.get("Snapshots", []):
                if snap["StartTime"] < ninety_days_ago:
                    out.append(
                        {
                            "ResourceType": "EBS Snapshot",
                            "ResourceId": snap["SnapshotId"],
                            "Reason": "Old Snapshot (>90 days)",
                            "Details": f"Created: {snap['StartTime'].strftime('%Y-%m-%d')}, Size: {snap['VolumeSize']} GiB",
                            "Description": snap.get("Description", ""),
                            "VolumeId": snap.get("VolumeId", ""),
                            "Encrypted": snap.get("Encrypted", False),
                            "StorageTier": snap.get("StorageTier", "standard"),
                            "Tags": {
                                t["Key"]: t["Value"] for t in snap.get("Tags", [])
                            },
                        }
                    )
    except Exception as e:
        log(
            "warning",
            f"Could not check for old EBS snapshots in {region}: {e}",
            account=alias,
        )

    # --- 5. S3 Buckets without Effective Lifecycle Policies & Incomplete Multipart Uploads ---
    try:
        buckets = s3_client.list_buckets().get("Buckets", [])
        for bucket in buckets:
            bucket_name = bucket["Name"]
            has_effective_policy = False
            try:
                lifecycle = s3_client.get_bucket_lifecycle_configuration(
                    Bucket=bucket_name
                )
                for rule in lifecycle.get("Rules", []):
                    if rule.get("Status") == "Enabled":
                        if (
                            "Transition" in rule
                            or "NoncurrentVersionTransition" in rule
                            or "AbortIncompleteMultipartUpload" in rule
                        ):
                            has_effective_policy = True
                            break
            except ClientError as e:
                if e.response["Error"]["Code"] != "NoSuchLifecycleConfiguration":
                    raise

            if not has_effective_policy:
                out.append(
                    {
                        "ResourceType": "S3 Bucket",
                        "ResourceId": bucket_name,
                        "Reason": "No Lifecycle Policy",
                        "Details": "Consider adding a lifecycle policy to transition or expire objects.",
                    }
                )

            # Check for old incomplete multipart uploads
            mpu_paginator = s3_client.get_paginator("list_multipart_uploads")
            for mpu_page in mpu_paginator.paginate(Bucket=bucket_name):
                for upload in mpu_page.get("Uploads", []):
                    if upload["Initiated"] < seven_days_ago:
                        out.append(
                            {
                                "ResourceType": "S3 Incomplete Upload",
                                "ResourceId": f"{bucket_name}/{upload['Key']}",
                                "Reason": "Incomplete Multipart Upload (>7 days)",
                                "Details": f"UploadId: {upload['UploadId']}, Initiated: {upload['Initiated'].strftime('%Y-%m-%d')}",
                            }
                        )
    except Exception as e:
        log(
            "warning",
            f"Could not check for S3 opportunities in {region}: {e}",
            account=alias,
        )

    # --- 6. Underutilized EC2 Instances ---
    try:
        for page in _safe_paginator(
            ec2_client.get_paginator("describe_instances").paginate,
            account=alias,
            Filters=[{"Name": "instance-state-name", "Values": ["running"]}],
        ):
            for reservation in page.get("Reservations", []):
                for instance in reservation.get("Instances", []):
                    instance_id = instance["InstanceId"]
                    metrics = cw_client.get_metric_statistics(
                        Namespace="AWS/EC2",
                    MetricName="CPUUtilization",
                    Dimensions=[{"Name": "InstanceId", "Value": instance_id}],
                    StartTime=fourteen_days_ago,
                    EndTime=datetime.now(timezone.utc),
                    Period=86400,
                    Statistics=["Maximum"],
                )
                if metrics["Datapoints"]:
                    max_cpu = max(dp["Maximum"] for dp in metrics["Datapoints"])
                    if max_cpu < 5:  # Less than 5% max utilization
                        out.append(
                            {
                                "ResourceType": "EC2 Instance",
                                "ResourceId": instance_id,
                                "Reason": "Underutilized (CPU < 5%)",
                                "Details": f"Type: {instance.get('InstanceType')}, Max CPU in 14 days: {max_cpu:.2f}%",
                            }
                        )
    except Exception as e:
        log(
            "warning",
            f"Could not check for underutilized EC2 instances in {region}: {e}",
            account=alias,
        )

    # --- 7. Underutilized RDS Instances ---
    try:
        dbs = rds_client.describe_db_instances().get("DBInstances", [])
        for db in dbs:
            db_id = db["DBInstanceIdentifier"]
            cpu_metrics = cw_client.get_metric_statistics(
                Namespace="AWS/RDS",
                MetricName="CPUUtilization",
                Dimensions=[{"Name": "DBInstanceIdentifier", "Value": db_id}],
                StartTime=fourteen_days_ago,
                EndTime=datetime.now(timezone.utc),
                Period=86400,
                Statistics=["Maximum"],
            )
            conn_metrics = cw_client.get_metric_statistics(
                Namespace="AWS/RDS",
                MetricName="DatabaseConnections",
                Dimensions=[{"Name": "DBInstanceIdentifier", "Value": db_id}],
                StartTime=fourteen_days_ago,
                EndTime=datetime.now(timezone.utc),
                Period=86400,
                Statistics=["Maximum"],
            )
            max_cpu = (
                max(dp["Maximum"] for dp in cpu_metrics.get("Datapoints", []))
                if cpu_metrics.get("Datapoints")
                else 0
            )
            max_conn = (
                max(dp["Maximum"] for dp in conn_metrics.get("Datapoints", []))
                if conn_metrics.get("Datapoints")
                else 0
            )
            if max_cpu < 5 and max_conn < 5:
                out.append(
                    {
                        "ResourceType": "RDS Instance",
                        "ResourceId": db_id,
                        "Reason": "Underutilized (CPU < 5% and < 5 connections)",
                        "Details": f"Type: {db.get('DBInstanceClass')}, Max CPU: {max_cpu:.2f}%, Max Connections: {max_conn}",
                    }
                )
    except Exception as e:
        log(
            "warning",
            f"Could not check for underutilized RDS instances in {region}: {e}",
            account=alias,
        )

    # --- 8. Lambda Functions with Low Invocation or High Error Rates ---
    try:
        functions = lambda_client.list_functions().get("Functions", [])
        for func in functions:
            func_name = func["FunctionName"]
            invocations_resp = cw_client.get_metric_statistics(
                Namespace="AWS/Lambda",
                MetricName="Invocations",
                Dimensions=[{"Name": "FunctionName", "Value": func_name}],
                StartTime=fourteen_days_ago,
                EndTime=datetime.now(timezone.utc),
                Period=1209600,
                Statistics=["Sum"],
            )
            total_invocations = sum(
                dp["Sum"] for dp in invocations_resp.get("Datapoints", [])
            )
            if total_invocations < 10:
                out.append(
                    {
                        "ResourceType": "Lambda Function",
                        "ResourceId": func_name,
                        "Reason": "Low Invocations (< 10 in 14 days)",
                        "Details": f"Total invocations: {total_invocations}",
                    }
                )

            errors_resp = cw_client.get_metric_statistics(
                Namespace="AWS/Lambda",
                MetricName="Errors",
                Dimensions=[{"Name": "FunctionName", "Value": func_name}],
                StartTime=fourteen_days_ago,
                EndTime=datetime.now(timezone.utc),
                Period=1209600,
                Statistics=["Sum"],
            )
            total_errors = sum(dp["Sum"] for dp in errors_resp.get("Datapoints", []))
            if total_invocations > 0 and (total_errors / total_invocations) > 0.1:
                error_rate = (total_errors / total_invocations) * 100
                out.append(
                    {
                        "ResourceType": "Lambda Function",
                        "ResourceId": func_name,
                        "Reason": "High Error Rate (>10%)",
                        "Details": f"Error Rate: {error_rate:.2f}% ({int(total_errors)} errors / {int(total_invocations)} invocations)",
                    }
                )

    except Exception as e:
        log(
            "warning",
            f"Could not check for Lambda opportunities in {region}: {e}",
            account=alias,
        )

    for item in out:
        item.update({"AccountAlias": alias, "Region": region})
    return out


# --------------------------------------------------------------------------
# UNIFIED COLLECTOR WRAPPERS
# --------------------------------------------------------------------------
def get_ec2_inventory(
    ec2_client: BaseClient,
    backup_client: BaseClient,
    alias: str,
    session: boto3.Session,
) -> Dict[str, List[Dict[str, Any]]]:
    """Unified collector for all EC2-related resources."""
    # The INFO log is silenced here for cleaner output during normal runs.
    # log("info", f"Fetching all EC2 resources in {ec2_client.meta.region_name}", account=alias)
    return {
        "EC2": get_ec2_details(ec2_client, backup_client, alias, session),
        "EC2ReservedInstances": get_ec2_reserved_instances(ec2_client, alias, session),
    }


def get_rds_inventory(
    rds_client: BaseClient, alias: str, session: boto3.Session
) -> Dict[str, List[Dict[str, Any]]]:
    """Unified collector for all RDS-related resources."""
    # log("info", f"Fetching all RDS resources in {rds_client.meta.region_name}", account=alias)
    return {
        "RDS": get_rds_details(rds_client, alias, session),
        "RDSReservedInstances": get_rds_reserved_instances(rds_client, alias, session),
    }


def get_eventbridge_inventory(
    events_client: BaseClient, scheduler_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """Unified collector for all EventBridge-related resources."""
    # log("info", f"Fetching all EventBridge resources in {events_client.meta.region_name}", account=alias)
    return {
        "EventBridge": get_eventbridge_details(events_client, alias),
        "EventBridgeScheduler": get_eventbridge_scheduler_details(
            scheduler_client, alias
        ),
    }


# --------------------------------------------------------------------------
# GLOBAL SERVICE REGISTRY
# --------------------------------------------------------------------------
GLOBAL_SERVICES = {
    "Route53": ("route53", get_route53_details),
    "SES": ("ses", get_ses_details),
    "SNS": ("sns", get_sns_details),
    "IAM": ("iam", get_iam_details),
}

# --------------------------------------------------------------------------
# EXCEL COLUMNS & UNIQUE KEYS
# --------------------------------------------------------------------------
SERVICE_COLUMNS: Dict[str, List[str]] = {
    "Summary": ["AccountId", "AccountAlias", "TotalResources"],
    "ACM": [
        "DomainName",
        "CertificateArn",
        "Status",
        "Type",
        "InUse",
        "Issued",
        "Expires",
        "RenewalEligibility",
        "SubjectAlternativeNames",
        "Region",
        "AccountAlias",
    ],
    "ALB": [
        "LoadBalancerName",
        "DNSName",
        "State",
        "Type",
        "Scheme",
        "VpcId",
        "AvailabilityZones",
        "SecurityGroups",
        "Listeners",
        "TargetGroups",
        "Tags",
        "Region",
        "AccountAlias",
    ],
    "Backup": [
        "PlanName",
        "RuleName",
        "SelectionName",
        "IamRole",
        "VaultName",
        "Schedule",
        "LastExecutionDate",
        "Details",
        "Timezone",
        "PlanId",
        "PlanArn",
        "PlanCreationDate",
        "Resources",
        "ResourceTags",
        "Region",
        "AccountAlias",
    ],
    "CloudWatchAlarms": [
        "AlarmName",
        "AlarmType",
        "MetricName",
        "Namespace",
        "Statistic",
        "ComparisonOperator",
        "Threshold",
        "Period",
        "EvaluationPeriods",
        "DatapointsToAlarm",
        "ActionsEnabled",
        "AlarmActions",
        "InsufficientDataActions",
        "OKActions",
        "Region",
        "AccountAlias",
    ],
    "CloudWatchLogs": [
        "LogGroupName",
        "RetentionInDays",
        "Size",
        "CreationTime",
        "MetricFilterCount",
        "Region",
        "AccountAlias",
    ],
    "CostOpportunities": [
        "ResourceType",
        "ResourceId",
        "Reason",
        "Details",
        "Description",
        "VolumeId",
        "Encrypted",
        "StorageTier",
        "Tags",
        "Region",
        "AccountAlias",
    ],
    "EC2": [
        "Name",
        "InstanceId",
        "InstanceType",
        "vCPUs",
        "Memory",
        "OS",
        "State",
        "PublicIP",
        "PrivateIP",
        "IPType",
        "AvailabilityZone",
        "EBSVolumes",
        "KeyPair",
        "NetworkPerformance",
        "AWSBackup",
        "Region",
        "AccountAlias",
    ],
    "EC2ReservedInstances": [
        "ReservedInstancesId",
        "InstanceType",
        "State",
        "InstanceCount",
        "OfferingType",
        "Scope",
        "ProductDescription",
        "Duration",
        "StartTime",
        "FixedPrice",
        "UsagePrice",
        "CurrencyCode",
        "Region",
        "AccountAlias",
    ],
    "EventBridge": [
        "ScheduleName",
        "GroupName",
        "State",
        "Frequency",
        "Expression",
        "Details",
        "Timezone",
        "TargetArn",
        "Input",
        "Region",
        "AccountAlias",
    ],
    "EventBridgeScheduler": [
        "ScheduleName",
        "GroupName",
        "State",
        "Frequency",
        "Expression",
        "Details",
        "Timezone",
        "TargetArn",
        "Input",
        "Region",
        "AccountAlias",
    ],
    "Governance": [
        "Service",
        "Status",
        "Details",
        "Region",
        "AccountAlias",
    ],
    "IAMGroups": [
        "GroupName",
        "GroupId",
        "Arn",
        "CreateDate",
        "Members",
        "AttachedPolicies",
        "InlinePolicies",
        "AccountAlias",
    ],
    "IAMPolicies": [
        "PolicyName",
        "PolicyId",
        "Arn",
        "PolicyType",
        "ServiceCategory",
        "AttachmentCount",
        "CreateDate",
        "UpdateDate",
        "DefaultVersionId",
        "ActionPatterns",
        "AttachmentEntities",
        "PolicyDocument",
        "AccountAlias",
    ],
    "IAMRoles": [
        "RoleName",
        "RoleId",
        "Arn",
        "CreateDate",
        "ServicePrincipals",
        "AccountPrincipals",
        "FederatedPrincipals",
        "AttachedPolicies",
        "InlinePolicies",
        "AccountAlias",
    ],
    "IAMUsers": [
        "UserName",
        "UserId",
        "Arn",
        "CreateDate",
        "PasswordLastUsed",
        "Groups",
        "AttachedPolicies",
        "InlinePolicies",
        "AccountAlias",
    ],
    "KMS": [
        "KeyId",
        "AliasNames",
        "Description",
        "KeyState",
        "Enabled",
        "KeyManager",
        "KeySpec",
        "KeyUsage",
        "Origin",
        "RotationEnabled",
        "CreationDate",
        "DeletionDate",
        "ValidTo",
        "MultiRegion",
        "PendingDeletion",
        "PendingWindowInDays",
        "GrantsCount",
        "Tags",
        "Region",
        "AccountAlias",
    ],
    "Lambda": [
        "FunctionName",
        "FunctionArn",
        "State",
        "LastUpdateStatus",
        "Runtime",
        "Handler",
        "Role",
        "Description",
        "MemorySize",
        "Timeout",
        "PackageType",
        "Architectures",
        "TracingMode",
        "LastModified",
        "KMSKeyArn",
        "CodeSize",
        "VpcSecurityGroupIds",
        "VpcSubnetIds",
        "EnvironmentVars",
        "Tags",
        "Region",
        "AccountAlias",
    ],
    "Lightsail": [
        "Name",
        "ResourceType",
        "State",
        "Region",
        "Location",
        "IpOrDnsName",
        "BlueprintOrEngine",
        "BundleId",
        "vCPUs",
        "MemoryInGB",
        "DiskSizeGB",
        "DataTransferGB",
        "AttachedTo",
        "SshKeyName",
        "Username",
        "Arn",
        "IpAddressType",
        "CreatedAt",
        "ExpiresAt",
        "FirewallRules",
        "AccountAlias",
    ],
    "RDS": [
        "DBInstanceIdentifier",
        "Engine",
        "DBInstanceClass",
        "vCPUs",
        "Memory",
        "MultiAZ",
        "StorageType",
        "AllocatedStorage",
        "EndpointAddress",
        "EndpointPort",
        "PubliclyAccessible",
        "InstanceCreateTime",
        "LicenseModel",
        "VpcSecurityGroupIds",
        "DBSubnetGroup",
        "Tags",
        "Region",
        "AccountAlias",
    ],
    "RDSReservedInstances": [
        "ReservedDBInstanceId",
        "DBInstanceClass",
        "State",
        "MultiAZ",
        "OfferingType",
        "Duration",
        "StartTime",
        "FixedPrice",
        "UsagePrice",
        "CurrencyCode",
        "ProductDescription",
        "Region",
        "AccountAlias",
    ],
    "Route53": [
        "Name",
        "Id",
        "ResourceRecordSetCount",
        "Config",
        "RecordTypes",
        "HealthChecks",
        "DNSSECStatus",
        "VPCAssociations",
        "DelegationSet",
        "Tags",
        "Region",
        "AccountAlias",
    ],
    "RouteTables": [
        "RouteTableId",
        "VpcId",
        "IsMain",
        "Routes",
        "Associations",
        "Tags",
        "Region",
        "AccountAlias",
    ],
    "S3": [
        "BucketName",
        "Region",
        "CreationDate",
        "Size",
        "ObjectCount",
        "Versioning",
        "Encryption",
        "PublicAccess",
        "PolicyStatus",
        "LastMetricsUpdate",
        "MetricsCalculationMethod",
        "LifecycleRules",
        "Tags",
        "AccountAlias",
    ],
    "SavingsPlans": [
        "SavingsPlanId",
        "State",
        "PlanType",
        "PaymentOption",
        "Term",
        "Start",
        "End",
        "SavingsPlanArn",
        "Region",
        "AccountAlias",
    ],
    "SecurityGroups": [
        "GroupId",
        "GroupName",
        "Description",
        "VpcId",
        "IngressRules",
        "EgressRules",
        "Tags",
        "Region",
        "AccountAlias",
    ],
    "SES": [
        "Identity",
        "VerificationStatus",
        "Region",
        "AccountAlias",
    ],
    "SNS": [
        "TopicName",
        "TopicArn",
        "SubscriptionCount",
        "Subscriptions",
        "Region",
        "AccountAlias",
    ],
    "Subnets": [
        "SubnetId",
        "VpcId",
        "State",
        "AvailabilityZone",
        "CidrBlock",
        "AvailableIpAddressCount",
        "MapPublicIpOnLaunch",
        "Tags",
        "Region",
        "AccountAlias",
    ],
    "VPCs": [
        "VpcId",
        "State",
        "IsDefault",
        "CidrBlock",
        "Tags",
        "Region",
        "AccountAlias",
    ],
    "VPN": [
        "VpnConnectionId",
        "Name",
        "State",
        "CustomerGateway",
        "CustomerGatewaySource",
        "TunnelOutsideIps",
        "Region",
        "AccountAlias",
    ],
    "WAFClassic": [
        "Name",
        "WebACLId",
        "RuleCount",
        "Rules",
        "Region",
        "AccountAlias",
    ],
    "WAFv2": [
        "Name",
        "WebACLId",
        "RuleCount",
        "Rules",
        "Region",
        "AccountAlias",
    ],
}

UNIQUE_KEYS: Dict[str, Union[str, Tuple[str, ...]]] = {
    "ACM": "CertificateArn",
    "ALB": "LoadBalancerName",
    "Backup": ("PlanId", "RuleName", "SelectionName"),
    "CloudWatchAlarms": "AlarmName",
    "CloudWatchLogs": "LogGroupName",
    "CostOpportunities": "ResourceId",
    "EC2": "InstanceId",
    "EC2ReservedInstances": "ReservedInstancesId",
    "EventBridge": "ScheduleName",
    "EventBridgeScheduler": "ScheduleName",
    "Governance": ("Region", "Service"),
    "IAMGroups": "GroupId",
    "IAMPolicies": "PolicyId",
    "IAMRoles": "RoleId",
    "IAMUsers": "UserId",
    "KMS": "KeyId",
    "Lambda": "FunctionArn",
    "Lightsail": "Arn",
    "RDS": "DBInstanceIdentifier",
    "RDSReservedInstances": "ReservedDBInstanceId",
    "Route53": "Id",
    "RouteTables": "RouteTableId",
    "S3": "BucketName",
    "SavingsPlans": "SavingsPlanId",
    "SecurityGroups": "GroupId",
    "SES": "Identity",
    "SNS": "TopicArn",
    "Subnets": "SubnetId",
    "VPCs": "VpcId",
    "VPN": "VpnConnectionId",
    "WAFClassic": "WebACLId",
    "WAFv2": "WebACLId",
}

_REGION_SCOPED_SHEETS: set[str] = {
    "ACM",
    "CloudWatchAlarms",
    "CloudWatchLogs",
    "KMS",
    "Lambda",
}


# --------------------------------------------------------------------------
# EXCEL WRITER
# --------------------------------------------------------------------------
class StreamingExcelWriter:
    def __init__(self, filename: str, export_tz: str):
        self._acct_ids: dict[str, str] = {}
        self.filename = filename
        self.export_tz = export_tz
        self.wb = Workbook()
        # remove default sheet
        self.wb.remove(self.wb.active)
        self.sheets: Dict[str, Worksheet] = {}
        # seen[(account)][sheet] = set of (unique_key, region)
        self._seen: Dict[str, Dict[str, set[Tuple[str, str]]]] = {}
        self._table_names: set[str] = set()

    @staticmethod
    def _excel_str(cell: Any, *, limit: int = 120) -> str:
        """
        Serialise a cell value for Excel width calculation.

        * `None`   → empty string
        * >`limit` → truncated with ellipsis
        * else     → str(value)
        """
        if cell is None:
            return ""
        s = str(cell)
        return (s[: limit - 3] + "...") if len(s) > limit else s

    @staticmethod
    def _logical_len(items: Union[Set[Any], Tuple[Any, ...]]) -> int:
        """
        Return the count of *logical* resources in *items*.
        Elements may be raw strings or tuples such as (resource_id, region).
        """
        if not items:
            return 0
        items = set(items)
        return len({v[0] if isinstance(v, tuple) else v for v in items})

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------
    def record_account(self, account_id: str, alias: str) -> None:
        """Remember the AccountId → Alias mapping once per scan."""
        self._acct_ids[account_id] = alias

    def _safe_table_name(self, sheet_name: str) -> str:
        """
        Generate a unique, Excel-legal table name (≤ 31 chars).
        """
        stem = "".join(c if c.isalnum() else "_" for c in sheet_name)
        if stem and stem[0].isdigit():
            stem = "_" + stem
        stem = stem[:_MAX_EXCEL_NAME_LEN]

        existing = self._table_names.union(self.wb.sheetnames)
        name, idx = stem, 1
        while name in existing:
            suffix = f"_{idx}"
            name = f"{stem[: _MAX_EXCEL_NAME_LEN - len(suffix)]}{suffix}"
            idx += 1

        self._table_names.add(name)
        return name

    def _get_sheet(self, name: str) -> Worksheet:
        # create on first use, write header row
        if name not in self.sheets:
            ws = self.wb.create_sheet(name[:31])
            ws.append(SERVICE_COLUMNS[name])
            self.sheets[name] = ws
        return self.sheets[name]

    def _append_serialized(
        self,
        ws: Worksheet,
        sheet: str,
        row: Dict[str, Any],
    ) -> None:
        """
        JSON-encode complex cells and append the row.
        """
        serialised = [
            (
                json.dumps(row[col], separators=(",", ":"))
                if isinstance(row.get(col), (dict, list, tuple, set))
                else row.get(col, "")
            )
            for col in SERVICE_COLUMNS[sheet]
        ]
        ws.append(serialised)

    def write_row(self, sheet: str, row: Dict[str, Any]) -> None:
        """
        Append a single row, JSON-encoding complex objects and
        de-duping within an account.
        """
        acct_id = row.get("AccountId")
        uniq_key = UNIQUE_KEYS.get(sheet)
        region = row.get("Region")

        if not (uniq_key and acct_id):
            self._append_serialized(self._get_sheet(sheet), sheet, row)
            return

        base_id = (
            tuple(row.get(k) for k in uniq_key)
            if isinstance(uniq_key, (list, tuple))
            else row.get(uniq_key)
        )

        # ensure the identifier itself is hashable
        if sheet in _REGION_SCOPED_SHEETS:
            identifier = (base_id, region)
        else:
            identifier = base_id

        seen = self._seen.setdefault(acct_id, {}).setdefault(sheet, set())
        if identifier in seen:
            return

        seen.add(identifier)
        self._append_serialized(self._get_sheet(sheet), sheet, row)

    def close(self) -> None:
        ordered = sorted(self.sheets)
        for idx, title in enumerate(ordered):
            ws = self.sheets[title]
            current_idx = self.wb.sheetnames.index(title)
            offset = idx - current_idx
            if offset:
                self.wb.move_sheet(ws, offset)

        # format each data sheet
        for name in ordered:
            ws = self.sheets[name]
            if ws.max_row == 1:
                continue
            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            tbl = Table(displayName=self._safe_table_name(name), ref=ref)
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleLight9", showRowStripes=True
            )
            ws.add_table(tbl)
            ws.freeze_panes = "A2"

            for idx in range(1, ws.max_column + 1):
                letter = get_column_letter(idx)
                sample = (c.value for c in ws[letter][:250])
                ws.column_dimensions[letter].width = (
                    max(len(self._excel_str(c)) for c in sample) + 2
                )

        summary = self.wb.create_sheet("Summary", 0)
        summary.append(
            [
                "Exported At:",
                datetime.now(ZoneInfo(self.export_tz)).strftime("%Y-%m-%d %H:%M:%S %Z"),
            ]
        )
        summary.append([])
        header = ["AccountId", "AccountAlias", "TotalResources"] + ordered
        summary.append(header)

        for acct_id, svc_map in self._seen.items():
            alias = self._acct_ids.get(acct_id, acct_id)
            total = sum(self._logical_len(svc_map.get(s, ())) for s in ordered)
            summary.append(
                [acct_id, alias, total]
                + [self._logical_len(svc_map.get(s, ())) for s in ordered]
            )

        ref = f"A3:{get_column_letter(summary.max_column)}{summary.max_row}"
        tbl = Table(displayName="SummaryTable", ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showRowStripes=True
        )
        summary.add_table(tbl)
        summary.freeze_panes = "A4"

        for idx in range(1, summary.max_column + 1):
            letter = get_column_letter(idx)
            summary.column_dimensions[letter].width = (
                max(len(str(c.value or "")) for c in summary[letter]) + 2
            )

        try:
            self.wb.save(self.filename)
        except PermissionError as exc:
            logger.error(
                "Failed to save %s (%s). Is the file open?",
                self.filename,
                exc,
                extra={"account": "-"},
            )
            raise


def parse_cli() -> argparse.Namespace:
    """Parses command-line arguments for the AWS Inventory Exporter."""
    parser = argparse.ArgumentParser(
        description="AWS Inventory Exporter - Canary Version"
    )

    # --- Arguments for Controlling the Scan ---
    parser.add_argument(
        "--master",
        action="store_true",
        help="Run from a management/master account to scan member accounts.",
    )
    parser.add_argument(
        "--exclude-accounts",
        default="",
        help="Comma-separated list of account IDs to exclude from the scan (e.g., 111122223333,444455556666).",
    )
    parser.add_argument(
        "--include-accounts",
        default="",
        help="Comma-separated list of account IDs to exclusively scan (e.g., 123456789012,987654321098).",
    )
    parser.add_argument(
        "--role-name",
        default=DEFAULT_ROLE,
        help=f"IAM role to assume in member accounts (default: {DEFAULT_ROLE}).",
    )
    parser.add_argument(
        "-r",
        "--regions",
        help="Comma-separated list of regions to scan (e.g., us-east-1,us-west-2). Default: all enabled regions.",
    )
    parser.add_argument(
        "--no-excel",
        action="store_true",
        help="Disable the generation of the Excel report.",
    )
    parser.add_argument(
        "--include",
        default="",
        help="Comma-separated list of collectors to include (e.g., EC2,S3,IAM).",
    )
    parser.add_argument(
        "--exclude",
        default="",
        help="Comma-separated list of collectors to exclude (e.g., WAFClassic,ALB).",
    )
    parser.add_argument(
        "--scan-mode",
        choices=["inventory", "security", "cost"],
        default="inventory",
        help="Specify the scan mode: 'inventory' for a full resource scan, 'security' for a security-focused scan, or 'cost' for a cost-optimization scan.",
    )

    return parser.parse_args()


# --------------------------------------------------------------------------
# ACCOUNT SCAN
# --------------------------------------------------------------------------
def scan_account(
    acct_id: str,
    sess: boto3.Session,
    regions: list[str],
    include_collectors: Optional[set[str]] = None,
    exclude_collectors: Optional[set[str]] = None,
) -> tuple[dict[str, Any], str]:
    """
    Scans a single AWS account for resources across specified regions.
    """
    try:
        aliases = (
            sess.client("iam", region_name="us-east-1")
            .list_account_aliases()
            .get("AccountAliases", [])
        )
        alias = aliases[0] if aliases else acct_id
    except Exception:
        alias = acct_id

    s3_global = aws_client("s3", "us-east-1", sess)
    all_buckets = s3_global.list_buckets().get("Buckets", [])

    buckets_by_region: dict[str, list[dict[str, Any]]] = {r: [] for r in regions}
    for b in all_buckets:
        b_region = bucket_region(b["Name"], regions[0], sess)
        if b_region in buckets_by_region:
            buckets_by_region[b_region].append(b)

    # Fetch supported regions once per account scan.
    lightsail_supported_regions = get_lightsail_supported_regions(sess, acct_id)

    primary_region = regions[0] if regions else "us-east-1"

    global_block: dict[str, Any] = {}
    for sheet, (api, fn) in GLOBAL_SERVICES.items():
        if include_collectors and sheet not in include_collectors:
            continue
        if exclude_collectors and sheet in exclude_collectors:
            continue
        try:
            client = aws_client(api, "us-east-1", sess)
            result = fn(client, alias)

            if sheet == "IAM":
                global_block.update(result)
            else:
                global_block[sheet] = result

        except Exception as exc:
            log("error", f"Global collector {sheet} failed: {exc}", account=acct_id)
            if sheet == "IAM":
                global_block.update(
                    {"IAMUsers": [], "IAMRoles": [], "IAMGroups": [], "IAMPolicies": []}
                )
            else:
                global_block[sheet] = []

    if (not include_collectors or "SavingsPlans" in include_collectors) and (
        not exclude_collectors or "SavingsPlans" not in exclude_collectors
    ):
        try:
            sp_client = aws_client("savingsplans", "us-east-1", sess)
            global_block["SavingsPlans"] = get_savings_plan_details(
                sp_client, alias, sess
            )
        except Exception as exc:
            log(
                "error", f"Global collector SavingsPlans failed: {exc}", account=acct_id
            )
            global_block["SavingsPlans"] = []

    global_block = {k: v for k, v in global_block.items() if v is not None}

    # Update region_worker to accept the supported regions set directly.
    def region_worker(
        region: str, supported_ls_regions: set[str]
    ) -> tuple[str, dict[str, Any]]:
        is_primary = region == primary_region

        # --- Define unified collector helpers ---
        def s3_collector():
            return get_s3_details(
                aws_client("s3", region, sess),
                buckets_by_region.get(region, []),
                alias,
                region,
                sess,
            )

        def ec2_collector():
            return get_ec2_inventory(
                aws_client("ec2", region, sess),
                aws_client("backup", region, sess),
                alias,
                sess,
            )

        def rds_collector():
            return get_rds_inventory(aws_client("rds", region, sess), alias, sess)

        def eventbridge_collector():
            return get_eventbridge_inventory(
                aws_client("events", region, sess),
                aws_client("scheduler", region, sess),
                alias,
            )

        def vpc_collector():
            return get_vpc_inventory(aws_client("ec2", region, sess), alias)

        def cost_collector():
            return get_cost_opportunities(
                aws_client("ec2", region, sess),
                aws_client("s3", region, sess),
                aws_client("elbv2", region, sess),
                aws_client("cloudwatch", region, sess),
                aws_client("lambda", region, sess),
                aws_client("rds", region, sess),
                alias,
            )

        def governance_collector():
            return get_governance_details(sess, alias, region)

        collectors: dict[str, Callable[[], Any]] = {
            "ACM": lambda: get_acm_details(aws_client("acm", region, sess), alias),
            "ALB": lambda: get_alb_details(aws_client("elbv2", region, sess), alias),
            "Backup": lambda: get_backup_details(
                aws_client("backup", region, sess), alias
            ),
            "CloudWatchAlarms": lambda: get_cloudwatch_alarms_details(
                aws_client("cloudwatch", region, sess), alias
            ),
            "CloudWatchLogs": lambda: get_cloudwatch_logs_details(
                aws_client("logs", region, sess), alias
            ),
            "CostOpportunities": cost_collector,
            "EC2": ec2_collector,
            "EventBridge": eventbridge_collector,
            "Governance": governance_collector,
            "KMS": lambda: get_kms_details(aws_client("kms", region, sess), alias),
            "Lambda": lambda: get_lambda_details(
                aws_client("lambda", region, sess), alias
            ),
            "Lightsail": lambda: get_lightsail_inventory(
                aws_client("lightsail", region, sess),
                alias,
                sess,
                is_primary,
                supported_ls_regions,
            ),
            "RDS": rds_collector,
            "S3": s3_collector,
            "VPC": vpc_collector,
            "VPN": lambda: get_vpn_details(aws_client("ec2", region, sess), alias),
            "WAFClassic": lambda: get_waf_classic_details(
                aws_client("waf-regional", region, sess), alias
            ),
            "WAFv2": lambda: get_waf_v2_details(
                aws_client("wafv2", region, sess), alias
            ),
        }

        # --- Smartly handle --include and --exclude for unified collectors ---
        unified_map = {
            "EC2": {"EC2", "EC2ReservedInstances"},
            "RDS": {"RDS", "RDSReservedInstances"},
            "EventBridge": {"EventBridge", "EventBridgeScheduler"},
            "VPC": {"VPCs", "Subnets", "RouteTables", "SecurityGroups"},
            "IAM": {"IAMUsers", "IAMRoles", "IAMGroups", "IAMPolicies"},
        }

        collectors_to_run_keys = set(collectors.keys())
        if include_collectors:
            expanded_includes = set(include_collectors)
            for key, sub_services in unified_map.items():
                if key in include_collectors:
                    expanded_includes.update(sub_services)
            collectors_to_run_keys &= expanded_includes

        if exclude_collectors:
            expanded_excludes = set(exclude_collectors)
            for key, sub_services in unified_map.items():
                if key in exclude_collectors:
                    expanded_excludes.update(sub_services)
            collectors_to_run_keys -= expanded_excludes

        collectors_to_run = {
            k: collectors[k] for k in collectors_to_run_keys if k in collectors
        }

        # --- Execute selected collectors ---
        region_block: dict[str, Any] = {}
        with ThreadPoolExecutor(max_workers=MAX_TASKS_IN_REGION) as pool:
            futures = {pool.submit(fn): name for name, fn in collectors_to_run.items()}
            for fut in as_completed(futures):
                sheet_name = futures[fut]
                try:
                    result = fut.result()
                    # If the collector is unified, its result is a dictionary to be merged.
                    if sheet_name in unified_map:
                        region_block.update(result)
                    else:
                        region_block[sheet_name] = result
                except Exception as exc:
                    log(
                        "error",
                        f"Collector {sheet_name} in {region} failed: {exc}",
                        account=acct_id,
                    )
                    if sheet_name in unified_map:
                        region_block.update({k: [] for k in unified_map[sheet_name]})
                    else:
                        region_block[sheet_name] = []

        return region, {k: v for k, v in region_block.items() if v is not None}

    # --- Main execution loop for accounts and regions ---
    all_regions: dict[str, Any] = {}
    with ThreadPoolExecutor(
        max_workers=min(len(regions), MAX_REGIONS_IN_FLIGHT)
    ) as pool:
        # Pass the supported regions set to each worker thread.
        futures = {
            pool.submit(region_worker, r, lightsail_supported_regions): r
            for r in regions
        }
        for fut in as_completed(futures):
            region_name, svc_data = fut.result()
            all_regions[region_name] = svc_data

    all_regions["global"] = global_block
    return all_regions, alias


# ────────────────────────────── main CLI ─────────────────────────────
def main() -> None:
    args = parse_cli()
    try:
        frozen = boto3.Session().get_credentials().get_frozen_credentials()
        if not (frozen.access_key and frozen.secret_key):
            raise RuntimeError
    except Exception:
        sys.exit(
            "ERROR: No AWS credentials - set AWS_PROFILE or AWS_ACCESS_KEY_ID / AWS_SECRET_ACCESS_KEY"
        )

    base_session = boto3.Session()
    try:
        ec2 = base_session.client("ec2", config=RETRY_CONFIG)
        resp = ec2.describe_regions(
            Filters=[
                {"Name": "opt-in-status", "Values": ["opt-in-not-required", "opted-in"]}
            ]
        )
        enabled_regions = [r["RegionName"] for r in resp["Regions"]]
    except (ClientError, NoCredentialsError):
        enabled_regions = base_session.get_available_regions("ec2")

    regions = (
        [r.strip() for r in args.regions.split(",")]
        if args.regions
        else enabled_regions
    )
    unknown = set(regions) - set(enabled_regions)
    if unknown:
        logger.error(
            "Region(s) requested but not enabled in this account: %s",
            ", ".join(sorted(unknown)),
        )
        sys.exit(2)

    include_accounts = {
        a.strip() for a in args.include_accounts.split(",") if a.strip()
    }
    exclude_accounts = {
        a.strip() for a in args.exclude_accounts.split(",") if a.strip()
    }

    # --- Handle Scan Mode ---
    user_includes = {c.strip() for c in args.include.split(",") if c.strip()}

    if args.scan_mode == "security":
        log("info", "Running in SECURITY scan mode.")
        if not user_includes:
            include_collectors = {"IAM", "Governance", "SecurityGroups", "VPC"}
        else:
            include_collectors = user_includes
    elif args.scan_mode == "cost":
        log("info", "Running in COST scan mode.")
        if not user_includes:
            include_collectors = {"CostOpportunities", "EC2", "RDS", "SavingsPlans"}
        else:
            include_collectors = user_includes
    else:
        log("info", "Running in default INVENTORY scan mode.")
        all_collectors = {
            "ACM",
            "ALB",
            "Backup",
            "CloudWatchAlarms",
            "CloudWatchLogs",
            "EC2",
            "EventBridge",
            "KMS",
            "Lambda",
            "Lightsail",
            "RDS",
            "S3",
            "VPC",
            "VPN",
            "WAFClassic",
            "WAFv2",
            "IAM",
            "SavingsPlans",
        }
        include_collectors = all_collectors

        if "CostOpportunities" in user_includes:
            include_collectors.add("CostOpportunities")
        if "Governance" in user_includes:
            include_collectors.add("Governance")

    exclude_collectors = {c.strip() for c in args.exclude.split(",") if c.strip()}
    if exclude_collectors:
        include_collectors -= exclude_collectors

    sessions: Dict[str, boto3.Session] = {}
    if args.master:
        org = base_session.client("organizations", config=RETRY_CONFIG)
        accounts = (
            a
            for page in org.get_paginator("list_accounts").paginate()
            for a in page["Accounts"]
            if a["Status"] == "ACTIVE"
        )
        for acct in accounts:
            aid = acct["Id"]
            if include_accounts and aid not in include_accounts:
                continue
            if aid in exclude_accounts:
                continue
            if sess := assume_role(aid, args.role_name, regions[0]):
                sessions[aid] = sess
    else:
        me = base_session.client("sts", config=RETRY_CONFIG).get_caller_identity()[
            "Account"
        ]
        if include_accounts and me not in include_accounts:
            return
        if me in exclude_accounts:
            return
        sessions[me] = base_session

    export_tz = str(get_localzone()) if get_localzone else "UTC"
    writer = None if args.no_excel else StreamingExcelWriter(EXCEL_FILENAME, export_tz)
    for acct_id, sess in sessions.items():
        log("info", f"Scanning account {acct_id}", account=acct_id)
        region_data, alias = scan_account(
            acct_id,
            sess,
            regions,
            include_collectors=include_collectors,
            exclude_collectors=exclude_collectors,
        )
        if writer:
            writer.record_account(acct_id, alias)
            for region_name, svc_block in region_data.items():
                for sheet, rows in svc_block.items():
                    if rows is not None:
                        for row in rows:
                            row["AccountId"] = acct_id
                            writer.write_row(sheet, row)
    if writer:
        writer.close()
    logger.info("Done.", extra={"account": "-"})


# ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit("Interrupted.")
