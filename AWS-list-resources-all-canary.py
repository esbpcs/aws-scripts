#!/usr/bin/env python3
"""
AWS Scan Report Generator

Collects metadata from multiple AWS services across one or more accounts
based on a selected scan mode (inventory, security, or cost) and
writes the results into a structured Excel workbook.
"""

# =================================================================================================
# IMPORTS
# =================================================================================================

from __future__ import annotations

import argparse
import json
import logging
import multiprocessing
import os
import random
import re
import sys
import threading
import time
from collections import OrderedDict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta, timezone
from functools import lru_cache
from itertools import islice
from typing import (
    Any,
    Callable,
    Dict,
    Iterator,
    List,
    Optional,
    Set,
    Tuple,
    Union,
)
from xml.sax.saxutils import unescape

import boto3
from botocore.client import BaseClient
from botocore.paginate import Paginator
from botocore.config import Config
from botocore.exceptions import (
    ClientError,
    EndpointConnectionError,
    NoCredentialsError,
)
from cron_descriptor import get_description as _cron_desc
from defusedxml import ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from tzlocal import get_localzone
from zoneinfo import ZoneInfo

# =================================================================================================
# CONFIGURATION
# =================================================================================================

# --- AWS API Configuration ---
CONNECT_TIMEOUT = int(os.getenv("AWS_CONNECT_TIMEOUT", "10"))
READ_TIMEOUT = int(os.getenv("AWS_READ_TIMEOUT", "60"))
MAX_ATTEMPTS = int(os.getenv("AWS_MAX_ATTEMPTS", "8"))
AWS_RETRY_MODE = os.getenv("AWS_RETRY_MODE", "standard")

# --- Parallelism Configuration ---
CPU_COUNT = multiprocessing.cpu_count()
MAX_REGIONS_IN_FLIGHT = min(int(os.getenv("MAX_PAR_REGION", "3")), 8)
MAX_TASKS_IN_REGION = min(int(os.getenv("MAX_PAR_TASK", "4")), 8)
_POOL_SIZE = MAX_REGIONS_IN_FLIGHT * MAX_TASKS_IN_REGION + 5

# --- Boto3 Retry Configuration ---
RETRY_CONFIG = Config(
    retries={
        "mode": AWS_RETRY_MODE,
        "max_attempts": MAX_ATTEMPTS,
    },
    connect_timeout=CONNECT_TIMEOUT,
    read_timeout=READ_TIMEOUT,
    max_pool_connections=_POOL_SIZE,
)

# --- Excel Configuration ---
_MAX_EXCEL_NAME_LEN = 31

# --- Cache Size Configuration ---
CLIENT_CACHE_MAX = int(os.getenv("AWS_CLIENT_CACHE_MAX", "128"))
FETCH_SPECS_CACHE_MAX = int(os.getenv("AWS_FETCH_SPECS_CACHE_MAX", "256"))
PARSE_VPN_CACHE_MAX = int(os.getenv("AWS_PARSE_VPN_CACHE_MAX", "128"))
BUCKET_REGION_CACHE_MAX = int(os.getenv("AWS_BUCKET_REGION_CACHE_MAX", "128"))
STORAGE_LENS_CACHE_MAX = int(os.getenv("AWS_STORAGE_LENS_CACHE_MAX", "128"))

# --- S3/CloudWatch Constants ---
_METRIC_PERIOD = 86_400
_METRIC_OFFSET = 48
MAX_KEYS_FOR_FULL_SCAN = int(os.getenv("AWS_MAX_KEYS_FOR_FULL_SCAN", "20000"))

# --- Miscellaneous Constants ---
SECS_PER_YEAR = 31_536_000
DEFAULT_ROLE = "OrganizationAccountAccessRole"

# --- Timezone Mapping ---
REGION_TZ: Dict[str, str] = {
    "us-east-1": "America/New_York",
    "us-east-2": "America/Chicago",
    "us-west-1": "America/Los_Angeles",
    "us-west-2": "America/Los_Angeles",
    "ca-central-1": "America/Toronto",
    "sa-east-1": "America/Sao_Paulo",
    "eu-west-1": "Europe/Dublin",
    "eu-west-2": "Europe/London",
    "eu-west-3": "Europe/Paris",
    "eu-north-1": "Europe/Stockholm",
    "eu-central-1": "Europe/Berlin",
    "eu-central-2": "Europe/Zurich",
    "eu-south-1": "Europe/Rome",
    "eu-south-2": "Europe/Madrid",
    "ap-south-1": "Asia/Kolkata",
    "ap-south-2": "Asia/Kolkata",
    "ap-northeast-1": "Asia/Tokyo",
    "ap-northeast-2": "Asia/Seoul",
    "ap-northeast-3": "Asia/Tokyo",
    "ap-southeast-1": "Asia/Singapore",
    "ap-southeast-2": "Australia/Sydney",
    "ap-southeast-3": "Asia/Jakarta",
    "ap-southeast-4": "Australia/Melbourne",
    "ap-east-1": "Asia/Hong_Kong",
    "me-south-1": "Asia/Bahrain",
    "me-central-1": "Asia/Dubai",
    "af-south-1": "Africa/Johannesburg",
    "af-north-1": "Africa/Cairo",
    "il-central-1": "Asia/Jerusalem",
}

# --- Load External Configuration ---
try:
    with open("config.json") as f:
        config = json.load(f)
    SERVICE_COLUMNS = config["SERVICE_COLUMNS"]
    UNIQUE_KEYS = config["UNIQUE_KEYS"]
    _REGION_SCOPED_SHEETS = set(config["REGION_SCOPED_SHEETS"])
except (FileNotFoundError, KeyError) as e:
    sys.exit(f"ERROR: Could not load configuration from config.json: {e}")

# =================================================================================================
# LOGGING
# =================================================================================================

LOG_FMT = "%(asctime)s %(levelname)-7s [%(account)s] %(name)s:%(lineno)d — %(message)s"
DATE_FMT = "%Y-%m-%dT%H:%M:%S%z"


class _TZFormatter(logging.Formatter):
    """Formats %(asctime)s in the region’s local time-zone."""

    def __init__(self, fmt: str, datefmt: str):
        super().__init__(fmt, datefmt)
        region = (
            os.getenv("AWS_REGION") or os.getenv("AWS_DEFAULT_REGION") or "us-east-1"
        )
        self._tz = ZoneInfo(REGION_TZ.get(region, "UTC"))

    def formatTime(
        self, record: logging.LogRecord, datefmt: Optional[str] = None
    ) -> str:
        dt = datetime.fromtimestamp(record.created, tz=self._tz)
        return dt.strftime(datefmt or self.datefmt or "%Y-%m-%d %H:%M:%S%z")


class _AccountFilter(logging.Filter):
    """Ensures that a `.account` attribute exists on every log record."""

    def filter(self, record: logging.LogRecord) -> bool:
        record.account = getattr(record, "account", "-")
        return True


_handler = logging.StreamHandler(sys.stdout)
_handler.setFormatter(_TZFormatter(LOG_FMT, DATE_FMT))
logger = logging.getLogger("aws_scan_report")
logger.setLevel(os.getenv("LOG_LEVEL", "INFO").upper())
logger.addHandler(_handler)
logger.addFilter(_AccountFilter())
logger.propagate = False


def log(level: str, msg: str, *args, account: str = "-", **kwargs) -> None:
    """A logging shortcut that injects `extra={'account': ...}` into the log record."""
    extra = kwargs.pop("extra", {})
    extra.setdefault("account", account)
    getattr(logger, level)(msg, *args, extra=extra, **kwargs)


# =================================================================================================
# HELPER UTILITIES
# =================================================================================================


def chunked(iterable, n: int):
    """Yields successive n-sized chunks from an iterable."""
    it = iter(iterable)
    while piece := list(islice(it, n)):
        yield piece


_thread_local = threading.local()
_client_eviction_lock = threading.Lock()


def _lru_cache_per_thread() -> OrderedDict:
    if not hasattr(_thread_local, "client_cache"):
        _thread_local.client_cache = OrderedDict()
    return _thread_local.client_cache


def aws_client(
    service: str, region: str, session: Optional[boto3.Session] = None
) -> BaseClient:
    """Creates a thread-local LRU cache around `boto3.Session.client`."""
    if session is None:
        if not hasattr(_thread_local, "default_session"):
            _thread_local.default_session = boto3.Session()
        session = _thread_local.default_session
    cache: OrderedDict = _lru_cache_per_thread()
    key = (id(session), service, region)
    if key in cache:
        cache.move_to_end(key)
        return cache[key]
    client = session.client(service, region_name=region, config=RETRY_CONFIG)
    cache[key] = client
    if len(cache) > CLIENT_CACHE_MAX:
        with _client_eviction_lock:
            if len(cache) > CLIENT_CACHE_MAX:
                cache.popitem(last=False)
    return client


def retry_with_backoff(
    fn: Callable[..., Dict[str, Any]],
    *,
    max_attempts: int = MAX_ATTEMPTS,
    base_delay: float = 1.0,
    max_delay: float = 20.0,
    default: Optional[Dict[str, Any]] = None,
    account: str = "-",
    **kwargs: Any,
) -> Dict[str, Any]:
    """Retries throttled AWS API calls with exponential backoff and jitter."""
    throttling_codes = {
        "Throttling",
        "ThrottlingException",
        "RequestLimitExceeded",
        "TooManyRequestsException",
        "SlowDown",
        "ProvisionedThroughputExceededException",
    }
    for attempt in range(1, max_attempts + 1):
        try:
            return fn(**kwargs)
        except ClientError as exc:
            code = exc.response.get("Error", {}).get("Code", "")
            if code not in throttling_codes and not code.startswith("Throttling"):
                raise
            if attempt == max_attempts:
                log(
                    "error",
                    f"Max retries reached for {getattr(fn,'__name__',fn)}: {code}",
                    account=account,
                )
                return default or {}
            delay = min(base_delay * (2 ** (attempt - 1)), max_delay)
            delay += delay * 0.3 * random.random()
            log(
                "warning",
                f"Throttled ({code}) {attempt}/{max_attempts}; retrying in {delay:.1f}s",
                account=account,
            )
            time.sleep(delay)
    return default or {}


def _safe_aws_call(
    fn: Callable[..., Dict[str, Any]],
    *,
    default: Optional[Dict[str, Any]] = None,
    account: str = "-",
    retry: bool = True,
    **kwargs: Any,
) -> Dict[str, Any]:
    """Executes an AWS SDK call and handles known-benign exceptions."""
    benign_s3_errors = {
        "NoSuchLifecycleConfiguration",
        "NoSuchTagSet",
        "NoSuchBucketPolicy",
        "NoSuchPublicAccessBlockConfiguration",
    }
    try:
        return (
            retry_with_backoff(fn, account=account, **kwargs) if retry else fn(**kwargs)
        )
    except ClientError as exc:
        code = exc.response.get("Error", {}).get("Code", "")
        if (
            isinstance(getattr(fn, "__self__", None), BaseClient)
            and getattr(fn, "__self__").meta.service_model.service_name == "s3"
            and code in benign_s3_errors
        ):
            return default or {}
        log(
            "warning",
            f"AWS call {getattr(fn, '__name__', fn)} failed: {exc}",
            account=account,
        )
        return default or {}
    except EndpointConnectionError as exc:
        log("warning", f"Endpoint error: {exc}", account=account)
        return default or {}


def _safe_paginator(
    paginate_fn: Callable[..., Any], *, account: str = "-", **kwargs: Any
) -> Iterator[Dict[str, Any]]:
    """Yields pages from a Boto3 paginator, swallowing network and client faults."""
    try:
        yield from paginate_fn(**kwargs)
    except (ClientError, EndpointConnectionError) as exc:
        log("warning", f"Paginator aborted: {exc}", account=account)


def require_paginator(client: BaseClient, op: str) -> Paginator:
    """Return a paginator or raise if the operation can't be paginated."""
    if not client.can_paginate(op):
        raise RuntimeError(
            f"{client.meta.service_model.service_name} cannot paginate '{op}'"
        )
    return client.get_paginator(op)


_TIMEZONE_OFFSET_RE = re.compile(r"([+-]\d{2})(\d{2})$")


def _tz_for(region: str) -> str:
    if region in REGION_TZ:
        return REGION_TZ[region]
    if region.startswith(("us-gov-", "us-iso-")):
        return "America/Denver" if region.endswith("-iso-b-1") else "America/New_York"
    if region.startswith("cn-"):
        return "Asia/Shanghai"
    if region.startswith("il-"):
        return "Asia/Jerusalem"
    return "UTC"


def to_local(dt: Union[str, datetime, None], region: str) -> str:
    """Converts a datetime object or string to a localized string."""
    if not dt:
        return ""
    if isinstance(dt, str):
        iso = _TIMEZONE_OFFSET_RE.sub(r"\1:\2", dt.strip().replace("Z", "+00:00"))
        try:
            dt = datetime.fromisoformat(iso)
        except ValueError:
            return dt
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    loc = dt.astimezone(ZoneInfo(_tz_for(region)))
    return f"{loc:%Y-%m-%d %H:%M:%S} {_tz_for(region)}"


def human_size(size_bytes: Optional[int]) -> str:
    """Converts a byte count to a human-readable string."""
    if size_bytes is None:
        return ""
    if size_bytes == 0:
        return "0 B"
    units = ["B", "KB", "MB", "GB", "TB", "PB"]
    size = float(size_bytes)
    for unit in units:
        if size < 1024 or unit == units[-1]:
            break
        size /= 1024
    return (
        f"{int(size)} B"
        if unit == "B"
        else f"{size:.1f}".rstrip("0").rstrip(".") + f" {unit}"
    )


def humanise_schedule(expr: str) -> Tuple[str, str]:
    """Converts a schedule expression to a human-readable string."""
    if expr.startswith("rate("):
        val, unit = expr[5:-1].split()
        unit = unit.rstrip("s")
        return (f"Every {unit}" if val == "1" else f"Every {val} {unit}s"), ""
    if expr.startswith("at("):
        return "One-time", expr[3:-1]
    if expr.startswith("cron("):
        cron_expr = expr[5:-1]
        try:
            return _cron_desc(cron_expr, locale="en"), ""
        except TypeError:
            return _cron_desc(cron_expr), ""
    return expr, ""


def seconds_to_years(sec: Union[int, float, None]) -> str:
    """Converts a duration in seconds to a string in years."""
    if sec is None or sec < 0:
        return ""
    yrs = sec / SECS_PER_YEAR
    return (
        f"{int(yrs)} Year" + ("" if int(yrs) == 1 else "s")
        if yrs.is_integer()
        else f"{yrs:.1f} Years"
    )


def format_tags(tags_list: List[Dict[str, str]]) -> Dict[str, str]:
    """Converts a list of Boto3 tag dictionaries into a simple key-value map."""
    if not tags_list:
        return {}
    return {tag.get("Key"): tag.get("Value") for tag in tags_list if tag.get("Key")}


# =================================================================================================
# IAM ROLE ASSUMPTION AND CACHING
# =================================================================================================


def assume_role(
    account_id: str, role_name: str = DEFAULT_ROLE, region: str = "us-east-1"
) -> Optional[boto3.Session]:
    """Assumes an IAM role in the specified account and returns a Boto3 session."""
    sts = boto3.client("sts", region_name=region, config=RETRY_CONFIG)
    try:
        creds_resp = retry_with_backoff(
            sts.assume_role,
            RoleArn=f"arn:aws:iam::{account_id}:role/{role_name}",
            RoleSessionName="ScanSession",
        )
        if "Credentials" in creds_resp:
            creds = creds_resp["Credentials"]
            return boto3.Session(
                aws_access_key_id=creds["AccessKeyId"],
                aws_secret_access_key=creds["SecretAccessKey"],
                aws_session_token=creds["SessionToken"],
                region_name=region,
            )
    except (ClientError, NoCredentialsError) as exc:
        logger.error(
            "AssumeRole failed for account %s: %s",
            account_id,
            exc,
            extra={"account": account_id},
        )
    return None


@lru_cache(maxsize=FETCH_SPECS_CACHE_MAX)
def fetch_instance_type_specs(
    types: tuple[str, ...], region: str, session: boto3.Session
) -> dict[str, Any]:
    """Fetches detailed specifications for a list of EC2 instance types."""
    ec2 = aws_client("ec2", region, session)
    out: dict[str, Any] = {}
    pending = list(types)
    while pending:
        chunk = pending[:100]
        try:
            resp = ec2.describe_instance_types(InstanceTypes=chunk)
            for it in resp["InstanceTypes"]:
                mib = it["MemoryInfo"]["SizeInMiB"]
                gib = mib / 1024
                out[it["InstanceType"]] = {
                    "vCPUs": it["VCpuInfo"]["DefaultVCpus"],
                    "Memory": (
                        f"{gib/1024:.1f} TB" if gib >= 1024 else f"{round(gib)} GB"
                    ),
                    "NetworkPerformance": it["NetworkInfo"].get(
                        "NetworkPerformance", ""
                    ),
                }
            pending = pending[100:]
        except ClientError as exc:
            if exc.response["Error"]["Code"] != "InvalidInstanceType":
                raise
            bad = [
                t.strip()
                for t in exc.response["Error"]["Message"]
                .split(":")[-1]
                .strip(" []")
                .split(",")
            ]
            log(
                "warning",
                f"Region {region}: unknown/retired instance types skipped: {', '.join(bad)}",
                account="-",
            )
            pending = [t for t in pending if t not in bad]
    return out


# =================================================================================================
# GENERIC RESOURCE COLLECTOR
# =================================================================================================


def get_aws_resource_details(
    client: BaseClient,
    paginator_name: str,
    list_key: str,
    alias: str,
    **paginator_kwargs,
) -> List[Dict[str, Any]]:
    """A generic function to list AWS resources using a paginator."""
    out = []
    for page in _safe_paginator(
        require_paginator(client, paginator_name).paginate,
        account=alias,
        **paginator_kwargs,
    ):
        for item in page.get(list_key, []):
            item.update({"AccountAlias": alias, "Region": client.meta.region_name})
            out.append(item)
    return out


# =================================================================================================
# BOUNDED CACHES & S3/VPN HELPERS
# =================================================================================================


@lru_cache(maxsize=PARSE_VPN_CACHE_MAX)
def parse_vpn_config(xml: str) -> Tuple[str, List[str]]:
    """Parses the VPN connection configuration XML to extract tunnel IPs."""
    if not xml:
        return "N/A", ["N/A"]
    root = ET.fromstring(unescape(xml).encode())
    cg = root.findtext(".//customer_gateway/tunnel_outside_address/ip_address") or "N/A"
    ips = {
        ip.text
        for ip in root.findall(
            ".//ipsec_tunnel/vpn_gateway/tunnel_outside_address/ip_address"
        )
        if ip.text
    }
    return cg, sorted(ips) or ["N/A"]


@lru_cache(maxsize=BUCKET_REGION_CACHE_MAX)
def bucket_region(
    bucket: str, region_hint: str, session: Optional[boto3.Session] = None
) -> Optional[str]:
    """Determines the AWS region of an S3 bucket."""
    s3 = aws_client("s3", region_hint, session)
    try:
        loc = s3.get_bucket_location(Bucket=bucket)["LocationConstraint"]
        return loc if loc is not None else "us-east-1"
    except ClientError as exc:
        if exc.response["Error"].get("Code") == "PermanentRedirect":
            return exc.response["Error"].get("BucketRegion")
        log("warning", f"Could not get region for bucket {bucket}: {exc}")
    return None


def _s3_select_sum_and_count(
    s3: BaseClient, bucket: str, key: str, *, account: str = "-"
) -> tuple[int, int]:
    """Uses S3 Select to calculate total size and count from an inventory file."""
    sql = 'SELECT CAST(sum(cast(s."Size" AS int)) AS bigint), CAST(count(1) AS bigint) FROM S3Object s'
    resp = _safe_aws_call(
        s3.select_object_content,
        account=account,
        Bucket=bucket,
        Key=key,
        ExpressionType="SQL",
        Expression=sql,
        InputSerialization={
            "CSV": {"FileHeaderInfo": "USE"},
            "CompressionType": "GZIP",
        },
        OutputSerialization={"JSON": {}},
    )
    records = b"".join(
        event["Records"]["Payload"]
        for event in resp.get("Payload", [])
        if "Records" in event
    )
    return tuple(json.loads(records)) if records else (0, 0)


@lru_cache(maxsize=128)
def inventory_metrics(
    bucket: str, region_hint: str, alias: str, session: Optional[boto3.Session] = None
) -> Tuple[Optional[int], Optional[int]]:
    """Retrieves bucket metrics from the latest S3 inventory report."""
    try:
        s3_inv = aws_client("s3", region_hint, session)
        inv_configs = s3_inv.list_bucket_inventory_configurations(Bucket=bucket).get(
            "InventoryConfigurationList", []
        )
        if not inv_configs:
            return None, None
        config = inv_configs[0]
        dest_bucket = config["Destination"]["S3BucketDestination"]["Bucket"].split(
            ":::"
        )[-1]
        prefix = config["Destination"]["S3BucketDestination"].get("Prefix", "")
        dest_region = bucket_region(dest_bucket, region_hint, session)
        if not dest_region:
            return None, None
        s3_dest_client = aws_client("s3", dest_region, session)
        paginator = require_paginator(s3_dest_client, "list_objects_v2")
        manifests = []
        for page in _safe_paginator(
            paginator.paginate,
            account=alias,
            Bucket=dest_bucket,
            Prefix=f"{prefix}/{bucket}/",
        ):
            manifests.extend(page.get("Contents", []))
        latest_manifest = max(manifests, key=lambda x: x["LastModified"], default=None)
        if not latest_manifest or not latest_manifest["Key"].endswith("manifest.json"):
            return None, None
        manifest_data = json.loads(
            s3_dest_client.get_object(Bucket=dest_bucket, Key=latest_manifest["Key"])[
                "Body"
            ].read()
        )
        total_bytes, total_objects = 0, 0
        with ThreadPoolExecutor(max_workers=10) as executor:
            futures = [
                executor.submit(
                    _s3_select_sum_and_count,
                    s3_dest_client,
                    dest_bucket,
                    file_meta["key"],
                    account=alias,
                )
                for file_meta in manifest_data["files"]
            ]
            for future in as_completed(futures):
                byte_chunk, obj_chunk = future.result()
                total_bytes += byte_chunk
                total_objects += obj_chunk
        return total_bytes, total_objects
    except Exception as e:
        log("debug", f"S3 Inventory metrics for {bucket} failed: {e}")
    return None, None


@lru_cache(maxsize=STORAGE_LENS_CACHE_MAX)
def storage_lens_metrics(
    bucket: str, region: str, session: Optional[boto3.Session] = None
) -> Tuple[Optional[int], Optional[int]]:
    """
    Dynamically discovers and queries S3 Storage Lens exports to retrieve bucket metrics.
    Gracefully handles cases where Storage Lens is not configured for S3 export.
    """
    try:
        s3control = aws_client(
            "s3control", "us-east-1", session
        )  # S3 Storage Lens config is global
        account_id = aws_client("sts", "us-east-1", session).get_caller_identity()[
            "Account"
        ]

        # 1. List all Storage Lens configurations for the account
        configs = []
        next_token = None
        while True:
            if next_token:
                resp = s3control.list_storage_lens_configurations(
                    AccountId=account_id,
                    NextToken=next_token,
                )
            else:
                resp = s3control.list_storage_lens_configurations(AccountId=account_id)
            configs.extend(resp.get("StorageLensConfigurationList", []))
            next_token = resp.get("NextToken")
            if not next_token:
                break
        if not configs:
            log("debug", "No S3 Storage Lens configurations found in this account.")
            return None, None

        export_config = None
        # 2. Find a configuration that is enabled and exports to S3
        for config_summary in configs:
            config_id = config_summary.get("Id")
            full_config_wrapper = s3control.get_storage_lens_configuration(
                AccountId=account_id, ConfigId=config_id
            )
            full_config = full_config_wrapper.get("StorageLensConfiguration", {})

            if full_config.get("IsEnabled") and full_config.get("S3BucketDestination"):
                export_config = full_config["S3BucketDestination"]
                break  # Use the first valid export configuration found

        if not export_config:
            log(
                "debug",
                "No S3 Storage Lens configuration with an S3 export destination was found.",
            )
            return None, None

        # 3. Automatically discover the export bucket and prefix from the configuration
        export_bucket_arn = export_config.get("Arn")
        if not export_bucket_arn:
            return None, None

        export_bucket = export_bucket_arn.split(":")[-1]
        export_prefix = export_config.get("Prefix", "")
        export_account_id = export_config.get("AccountId")

        dest_region = bucket_region(export_bucket, region, session)
        if not dest_region:
            return None, None

        s3_dest_client = aws_client("s3", dest_region, session)

        # 4. Find the latest manifest file in the dynamically discovered location
        # The path is typically /Bucket/AccountId/DashboardId/YYYY-MM-DDTHH-MMZ/manifest.json
        manifest_prefix = f"{export_prefix}{export_account_id}/"
        paginator = require_paginator(s3_dest_client, "list_objects_v2")
        manifests = []
        for page in _safe_paginator(
            paginator.paginate,
            account=account_id,
            Bucket=export_bucket,
            Prefix=manifest_prefix,
        ):
            manifests.extend(page.get("Contents", []))

        latest_manifest = max(
            (m for m in manifests if m.get("Key", "").endswith("manifest.json")),
            key=lambda x: x["LastModified"],
            default=None,
        )

        if not latest_manifest:
            log(
                "debug",
                f"No recent Storage Lens manifest found in bucket {export_bucket}",
            )
            return None, None

        # 5. Download and parse the manifest to find the data file key
        manifest_data = json.loads(
            s3_dest_client.get_object(Bucket=export_bucket, Key=latest_manifest["Key"])[
                "Body"
            ].read()
        )
        data_file_key = manifest_data.get("files", [{}])[0].get("key")

        if not data_file_key:
            return None, None

        # 6. Use S3 Select to query the data file for our specific bucket
        sql = f'SELECT s."total_storage_bytes", s."object_count" FROM S3Object s WHERE s."bucket_name" = \'{bucket}\''
        resp = _safe_aws_call(
            s3_dest_client.select_object_content,
            Bucket=export_bucket,
            Key=data_file_key,
            ExpressionType="SQL",
            Expression=sql,
            InputSerialization={
                "CSV": {"FileHeaderInfo": "USE"},
                "CompressionType": "GZIP",
            },
            OutputSerialization={"JSON": {}},
        )

        records = b"".join(
            event["Records"]["Payload"]
            for event in resp.get("Payload", [])
            if "Records" in event
        )
        if records:
            data = json.loads(records)
            return int(data["_1"]), int(data["_2"])

    except ClientError as e:
        # Gracefully handle cases where the service is not subscribed
        if e.response["Error"]["Code"] == "NotSignedUp":
            log("debug", "Account is not signed up for S3 Storage Lens.")
        else:
            log("debug", f"StorageLens lookup for {bucket} failed: {e}")
    except Exception as e:
        log(
            "debug",
            f"An unexpected error occurred during StorageLens lookup for {bucket}: {e}",
        )

    return None, None


@lru_cache(maxsize=1)
def get_lightsail_supported_regions(session: boto3.Session, account: str) -> set[str]:
    """Returns a set of region names where Lightsail is supported."""
    try:
        ls_client = aws_client("lightsail", "us-east-1", session)
        return {
            r["name"]
            for r in ls_client.get_regions(includeAvailabilityZones=False).get(
                "regions", []
            )
        }
    except Exception as e:
        log(
            "warning",
            f"Could not query for available Lightsail regions: {e}",
            account=account,
        )
        return set()


# =================================================================================================
# COST CALCULATION HELPERS
# =================================================================================================
REGION_NAMES = {
    "us-east-1": "US East (N. Virginia)",
    "us-east-2": "US East (Ohio)",
    "us-west-1": "US West (N. California)",
    "us-west-2": "US West (Oregon)",
    "ca-central-1": "Canada (Central)",
    "eu-central-1": "EU (Frankfurt)",
    "eu-west-1": "EU (Ireland)",
    "eu-west-2": "EU (London)",
    "eu-west-3": "EU (Paris)",
    "eu-north-1": "EU (Stockholm)",
    "ap-northeast-1": "Asia Pacific (Tokyo)",
    "ap-northeast-2": "Asia Pacific (Seoul)",
    "ap-southeast-1": "Asia Pacific (Singapore)",
    "ap-southeast-2": "Asia Pacific (Sydney)",
    "ap-south-1": "Asia Pacific (Mumbai)",
    "sa-east-1": "South America (Sao Paulo)",
}


@lru_cache(maxsize=128)
def get_ebs_volume_cost(
    volume_type: str, volume_size: int, region: str, session: boto3.Session
) -> Optional[float]:
    """
    Calculates the estimated monthly cost of an EBS volume using the AWS Pricing API.
    """
    try:
        pricing_client = aws_client(
            "pricing", "us-east-1", session
        )  # Pricing API is only in us-east-1
        response = pricing_client.get_products(
            ServiceCode="AmazonEC2",
            Filters=[
                {"Type": "TERM_MATCH", "Field": "productFamily", "Value": "Storage"},
                {"Type": "TERM_MATCH", "Field": "volumeApiName", "Value": volume_type},
                {
                    "Type": "TERM_MATCH",
                    "Field": "location",
                    "Value": REGION_NAMES.get(region, region),
                },
            ],
        )

        for price_item_str in response.get("PriceList", []):
            price_item = json.loads(price_item_str)
            on_demand = price_item.get("terms", {}).get("OnDemand", {})
            price_dimensions = next(iter(on_demand.values()), {}).get(
                "priceDimensions", {}
            )
            price_per_gb_str = (
                next(iter(price_dimensions.values()), {})
                .get("pricePerUnit", {})
                .get("USD")
            )

            if price_per_gb_str:
                return float(price_per_gb_str) * volume_size
    except Exception as e:
        log(
            "warning",
            f"Could not calculate EBS cost for type {volume_type} in {region}: {e}",
        )
    return None


@lru_cache(maxsize=20)
def get_eip_cost(region: str, session: boto3.Session) -> Optional[float]:
    """
    Calculates the estimated monthly cost of an unassociated Elastic IP.
    """
    try:
        pricing_client = aws_client("pricing", "us-east-1", session)
        response = pricing_client.get_products(
            ServiceCode="AmazonEC2",
            Filters=[
                {"Type": "TERM_MATCH", "Field": "productFamily", "Value": "IP Address"},
                {
                    "Type": "TERM_MATCH",
                    "Field": "usagetype",
                    "Value": f"{region.upper()}-IdleAddress:Vpc",
                },
            ],
        )

        price_list = response.get("PriceList", [])
        if not price_list:
            return None

        price_item = json.loads(price_list[0])
        on_demand = price_item.get("terms", {}).get("OnDemand", {})
        price_dimensions = next(iter(on_demand.values()), {}).get("priceDimensions", {})
        price_per_hour_str = (
            next(iter(price_dimensions.values()), {}).get("pricePerUnit", {}).get("USD")
        )

        if price_per_hour_str:
            return float(price_per_hour_str) * 730

    except Exception as e:
        log("warning", f"Could not calculate EIP cost for region {region}: {e}")
    return None


def get_ec2_instance_cost(instance_id: str, ce_client: BaseClient) -> Optional[float]:
    """
    Gets the cost of a specific EC2 instance over the last 30 days using Cost Explorer.
    """
    try:
        end_date = datetime.now(timezone.utc)
        start_date = end_date - timedelta(days=30)

        response = ce_client.get_cost_and_usage_with_resources(
            TimePeriod={
                "Start": start_date.strftime("%Y-%m-%d"),
                "End": end_date.strftime("%Y-%m-%d"),
            },
            Granularity="MONTHLY",
            Metrics=["UnblendedCost"],
            Filter={"Dimensions": {"Key": "RESOURCE_ID", "Values": [instance_id]}},
        )

        results = response.get("ResultsByTime", [])
        if results and results[0].get("Groups"):
            cost_str = results[0]["Groups"][0]["Metrics"]["UnblendedCost"]["Amount"]
            return float(cost_str)
    except Exception as e:
        log(
            "warning",
            f"Could not get Cost Explorer data for instance {instance_id}: {e}",
        )
    return None


# =================================================================================================
# SERVICE-SPECIFIC COLLECTORS
# =================================================================================================
def get_acm_details(
    acm_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """Collects details for ACM certificates."""
    out = []
    # Use the generic collector to list summaries, then describe each one.
    for cert_summary in get_aws_resource_details(
        client=acm_client,
        paginator_name="list_certificates",
        list_key="CertificateSummaryList",
        alias=alias,
    ):
        details = _safe_aws_call(
            acm_client.describe_certificate,
            default={},
            account=alias,
            CertificateArn=cert_summary["CertificateArn"],
        ).get("Certificate", {})

        if details:
            # Manually add common fields and transform data
            details["AccountAlias"] = alias
            details["Region"] = acm_client.meta.region_name
            details["InUse"] = bool(details.get("InUseBy"))
            details["Issued"] = to_local(
                details.get("IssuedAt"), acm_client.meta.region_name
            )
            details["Expires"] = to_local(
                details.get("NotAfter"), acm_client.meta.region_name
            )
            out.append(details)
    return {"ACM": out}


def get_alb_details(
    elbv2_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """Collects details for Application and Network Load Balancers."""
    out: List[Dict[str, Any]] = []

    # First, list all load balancers in the region.
    lbs = get_aws_resource_details(
        client=elbv2_client,
        paginator_name="describe_load_balancers",
        list_key="LoadBalancers",
        alias=alias,
    )

    # Process each load balancer individually to gather its associated details.
    for lb in lbs:
        arn = lb.get("LoadBalancerArn")
        if not arn:
            continue

        # Get Tags for the load balancer.
        tags_desc = _safe_aws_call(
            elbv2_client.describe_tags,
            default={"TagDescriptions": []},
            account=alias,
            ResourceArns=[arn],
        )["TagDescriptions"]

        # Use the standardized format_tags helper
        tags = format_tags(tags_desc[0].get("Tags", [])) if tags_desc else {}

        # Get Listeners associated with the load balancer.
        listeners = [
            {"Port": listener.get("Port"), "Protocol": listener.get("Protocol")}
            for page in _safe_paginator(
                require_paginator(elbv2_client, "describe_listeners").paginate,
                account=alias,
                LoadBalancerArn=arn,
            )
            for listener in page.get("Listeners", [])
        ]

        # Get Target Groups associated with the load balancer.
        target_groups = [
            tg.get("TargetGroupName")
            for tg in _safe_aws_call(
                elbv2_client.describe_target_groups,
                default={"TargetGroups": []},
                account=alias,
                LoadBalancerArn=arn,
            ).get("TargetGroups", [])
        ]

        # Assemble the final record for the report.
        out.append(
            {
                "LoadBalancerName": lb.get("LoadBalancerName"),
                "DNSName": lb.get("DNSName"),
                "State": lb.get("State", {}).get("Code"),
                "Type": lb.get("Type"),
                "Scheme": lb.get("Scheme"),
                "VpcId": lb.get("VpcId"),
                "AvailabilityZones": [
                    az.get("ZoneName") for az in lb.get("AvailabilityZones", [])
                ],
                "SecurityGroups": lb.get("SecurityGroups", []),
                "Tags": tags,
                "Listeners": listeners,
                "TargetGroups": target_groups,
                "Region": elbv2_client.meta.region_name,
                "AccountAlias": alias,
            }
        )

    return {"ALB": out}


def get_backup_details(
    backup_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """Collects details for AWS Backup plans, rules, and selections."""
    out: List[Dict[str, Any]] = []

    # List all backup plans in the region.
    plans = get_aws_resource_details(
        client=backup_client,
        paginator_name="list_backup_plans",
        list_key="BackupPlansList",
        alias=alias,
    )

    for summary in plans:
        pid, vid = summary["BackupPlanId"], summary["VersionId"]

        # Get the full details for each backup plan.
        plan = _safe_aws_call(
            backup_client.get_backup_plan,
            default={"BackupPlan": {}},
            account=alias,
            BackupPlanId=pid,
            VersionId=vid,
        )["BackupPlan"]
        if not plan:
            continue

        # Get all backup selections associated with the plan.
        selections = _safe_aws_call(
            backup_client.list_backup_selections,
            default={"BackupSelectionsList": []},
            account=alias,
            BackupPlanId=pid,
        )["BackupSelectionsList"]

        # Process each rule within the backup plan.
        for rule in plan.get("Rules", []):
            freq, det = humanise_schedule(rule.get("ScheduleExpression", ""))
            vault = rule.get("TargetBackupVaultName", "")

            # Get the last completion time for the backup vault associated with the rule.
            jobs = _safe_aws_call(
                backup_client.list_backup_jobs,
                default={"BackupJobs": []},
                account=alias,
                ByBackupVaultName=vault,
                ByState="COMPLETED",
                MaxResults=1,
            )["BackupJobs"]
            last_exec = to_local(
                (
                    (jobs[0].get("CompletionDate") or jobs[0].get("CreationDate"))
                    if jobs
                    else None
                ),
                backup_client.meta.region_name,
            )

            # If there are no selections, still report the plan and rule.
            if not selections:
                out.append(
                    {
                        "PlanName": plan.get("BackupPlanName", ""),
                        "RuleName": rule.get("RuleName", ""),
                        "SelectionName": "N/A",
                        "IamRole": "N/A",
                        "VaultName": vault,
                        "Schedule": freq,
                        "LastExecutionDate": last_exec,
                        "Details": det,
                        "Timezone": rule.get(
                            "ScheduleExpressionTimezone",
                            REGION_TZ.get(backup_client.meta.region_name, "UTC"),
                        ),
                        "PlanId": pid,
                        "PlanArn": summary.get("BackupPlanArn", ""),
                        "PlanCreationDate": to_local(
                            summary.get("CreationDate"), backup_client.meta.region_name
                        ),
                        "Resources": [],
                        "ResourceTags": [],
                        "Region": backup_client.meta.region_name,
                        "AccountAlias": alias,
                    }
                )
            else:
                # Associate each selection with the current rule's details.
                for sel in selections:
                    selection_details = _safe_aws_call(
                        backup_client.get_backup_selection,
                        default={"BackupSelection": {}},
                        account=alias,
                        BackupPlanId=pid,
                        SelectionId=sel["SelectionId"],
                    )["BackupSelection"]

                    out.append(
                        {
                            "PlanName": plan.get("BackupPlanName", ""),
                            "RuleName": rule.get("RuleName", ""),
                            "SelectionName": selection_details.get("SelectionName", ""),
                            "IamRole": selection_details.get("IamRoleArn", ""),
                            "VaultName": vault,
                            "Schedule": freq,
                            "LastExecutionDate": last_exec,
                            "Details": det,
                            "Timezone": rule.get(
                                "ScheduleExpressionTimezone",
                                REGION_TZ.get(backup_client.meta.region_name, "UTC"),
                            ),
                            "PlanId": pid,
                            "PlanArn": summary.get("BackupPlanArn", ""),
                            "PlanCreationDate": to_local(
                                summary.get("CreationDate"),
                                backup_client.meta.region_name,
                            ),
                            "Resources": selection_details.get("Resources", []),
                            "ResourceTags": selection_details.get("ListOfTags", []),
                            "Region": backup_client.meta.region_name,
                            "AccountAlias": alias,
                        }
                    )
    return {"Backup": out}


def get_cloudwatch_alarms_details(
    cw_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """Collects details for both Metric and Composite CloudWatch Alarms."""
    out: List[Dict[str, Any]] = []

    # The describe_alarms paginator efficiently returns both alarm types.
    for page in _safe_paginator(
        require_paginator(cw_client, "describe_alarms").paginate, account=alias
    ):
        # --- Process Metric Alarms ---
        for alarm in page.get("MetricAlarms", []):
            out.append(
                {
                    "AlarmName": alarm.get("AlarmName"),
                    "AlarmType": "Metric",
                    "MetricName": alarm.get("MetricName"),
                    "Namespace": alarm.get("Namespace"),
                    "Statistic": alarm.get("Statistic"),
                    "ComparisonOperator": alarm.get("ComparisonOperator"),
                    "Threshold": alarm.get("Threshold"),
                    "Period": alarm.get("Period"),
                    "EvaluationPeriods": alarm.get("EvaluationPeriods"),
                    "DatapointsToAlarm": alarm.get("DatapointsToAlarm"),
                    "ActionsEnabled": alarm.get("ActionsEnabled"),
                    "AlarmActions": alarm.get("AlarmActions"),
                    "InsufficientDataActions": alarm.get("InsufficientDataActions"),
                    "OKActions": alarm.get("OKActions"),
                    "Details": alarm.get("AlarmDescription", ""),
                    "Region": cw_client.meta.region_name,
                    "AccountAlias": alias,
                }
            )

        # --- Process Composite Alarms with Readable Details  ---
        for alarm in page.get("CompositeAlarms", []):
            out.append(
                {
                    "AlarmName": alarm.get("AlarmName"),
                    "AlarmType": "Composite",
                    "MetricName": "N/A (Composite)",
                    "Namespace": "N/A (Composite)",
                    "Statistic": "N/A",
                    "ComparisonOperator": "N/A",
                    "Threshold": "N/A",  # No threshold for composite alarms
                    "Period": "N/A",
                    "EvaluationPeriods": "N/A",
                    "DatapointsToAlarm": "N/A",
                    "ActionsEnabled": alarm.get("ActionsEnabled"),
                    "AlarmActions": alarm.get("AlarmActions"),
                    "InsufficientDataActions": alarm.get("InsufficientDataActions"),
                    "OKActions": alarm.get("OKActions"),
                    "Details": alarm.get("AlarmRule", "No rule found"),
                    "Region": cw_client.meta.region_name,
                    "AccountAlias": alias,
                }
            )

    return {"CloudWatchAlarms": out}


def get_cloudwatch_logs_details(
    logs_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """Collects details for CloudWatch Log Groups."""
    out = []

    # Use the generic helper to list all log groups.
    log_groups = get_aws_resource_details(
        client=logs_client,
        paginator_name="describe_log_groups",
        list_key="logGroups",
        alias=alias,
    )

    for group in log_groups:
        log_group_name = group.get("logGroupName")
        if not log_group_name:
            continue

        # For each log group, collect all associated metric filters.
        metric_filters: List[Dict[str, Any]] = []
        for page in _safe_paginator(
            require_paginator(logs_client, "describe_metric_filters").paginate,
            account=alias,
            logGroupName=log_group_name,
        ):
            metric_filters.extend(page.get("metricFilters", []))

        # Build a new dictionary safely to prevent KeyErrors and ensure data integrity.
        out.append(
            {
                "LogGroupName": log_group_name,
                "RetentionInDays": group.get("retentionInDays", "Never"),
                "Size": human_size(group.get("storedBytes")),
                "CreationTime": to_local(
                    (
                        datetime.fromtimestamp(
                            group.get("creationTime", 0) / 1000, timezone.utc
                        )
                        if group.get("creationTime")
                        else None
                    ),
                    logs_client.meta.region_name,
                ),
                "MetricFilterCount": len(metric_filters),
                "Region": logs_client.meta.region_name,
                "AccountAlias": alias,
            }
        )

    return {"CloudWatchLogs": out}


def get_cost_opportunities(
    ec2_client: BaseClient,
    s3_client: BaseClient,
    elbv2_client: BaseClient,
    cw_client: BaseClient,
    ce_client: BaseClient,
    alias: str,
    session: boto3.Session,
) -> Dict[str, List[Dict[str, Any]]]:
    """Identifies potential cost savings opportunities and estimates the savings."""
    out: List[Dict[str, Any]] = []
    region = ec2_client.meta.region_name
    now = datetime.now(timezone.utc)
    fourteen_days_ago = now - timedelta(days=14)
    seven_days_ago = now - timedelta(days=7)
    ninety_days_ago = now - timedelta(days=90)

    # --- 1. Unattached EBS Volumes ---
    try:
        for page in _safe_paginator(
            require_paginator(ec2_client, "describe_volumes").paginate,
            account=alias,
            Filters=[{"Name": "status", "Values": ["available"]}],
        ):
            for vol in page.get("Volumes", []):
                cost = get_ebs_volume_cost(vol["VolumeType"], vol["Size"], region, session)
                out.append(
                {
                    "ResourceType": "EBS Volume",
                    "ResourceId": vol["VolumeId"],
                    "Reason": "Unattached (Available)",
                    "Details": f"Size: {vol['Size']} GiB, Type: {vol['VolumeType']}",
                    "EstimatedMonthlySavings": (
                        f"${cost:.2f}" if cost is not None else "N/A"
                    ),
                }
            )
    except Exception as e:
        log(
            "warning",
            f"Could not check for unattached EBS volumes in {region}: {e}",
            account=alias,
        )

    # --- 2. Unassociated Elastic IPs ---
    try:
        eip_monthly_cost = get_eip_cost(region, session)
        for page in require_paginator(ec2_client, "describe_addresses").paginate():
            for addr in page.get("Addresses", []):
                if "AssociationId" not in addr:
                    out.append(
                        {
                            "ResourceType": "Elastic IP",
                            "ResourceId": addr["PublicIp"],
                            "Reason": "Unassociated",
                            "Details": f"AllocationId: {addr['AllocationId']}",
                            "EstimatedMonthlySavings": (
                                f"${eip_monthly_cost:.2f}"
                                if eip_monthly_cost is not None
                                else "N/A"
                            ),
                        }
                    )
    except Exception as e:
        log(
            "warning",
            f"Could not check for unassociated Elastic IPs in {region}: {e}",
            account=alias,
        )

    # --- 3. Idle Load Balancers ---
    try:
        for lb_page in require_paginator(
            elbv2_client, "describe_load_balancers"
        ).paginate():
            for lb in lb_page.get("LoadBalancers", []):
                target_groups: List[Dict[str, Any]] = []
                for tg_page in require_paginator(
                    elbv2_client, "describe_target_groups"
                ).paginate(LoadBalancerArn=lb["LoadBalancerArn"]):
                    target_groups.extend(tg_page.get("TargetGroups", []))

                is_idle = not target_groups or all(
                    not elbv2_client.describe_target_health(
                        TargetGroupArn=tg["TargetGroupArn"]
                    ).get("TargetHealthDescriptions")
                    for tg in target_groups
                )
            if is_idle:
                out.append(
                    {
                        "ResourceType": "Load Balancer",
                        "ResourceId": lb["DNSName"],
                        "Reason": "Idle (No Healthy Targets)",
                        "Details": f"Name: {lb['LoadBalancerName']}",
                        "EstimatedMonthlySavings": "N/A",  # Cost is complex, depends on LCUs.
                    }
                )
    except Exception as e:
        log(
            "warning",
            f"Could not check for idle Load Balancers in {region}: {e}",
            account=alias,
        )

    # --- 4. Old EBS Snapshots ---
    try:
        snapshot_cost_per_gb = get_ebs_volume_cost("gp2.snapshot", 1, region, session)
        for page in require_paginator(ec2_client, "describe_snapshots").paginate(
            OwnerIds=["self"]
        ):
            for snap in page.get("Snapshots", []):
                if snap["StartTime"] < ninety_days_ago:
                    cost = (
                        snapshot_cost_per_gb * snap["VolumeSize"]
                        if snapshot_cost_per_gb
                        else None
                    )
                    out.append(
                        {
                            "ResourceType": "EBS Snapshot",
                            "ResourceId": snap["SnapshotId"],
                            "Reason": "Old Snapshot (>90 days)",
                            "Details": f"Created: {snap['StartTime']:%Y-%m-%d}, Size: {snap['VolumeSize']} GiB",
                            "EstimatedMonthlySavings": (
                                f"${cost:.2f}" if cost is not None else "N/A"
                            ),
                        }
                    )
    except Exception as e:
        log(
            "warning",
            f"Could not check for old EBS snapshots in {region}: {e}",
            account=alias,
        )

    # --- 5. S3 Incomplete Multipart Uploads ---
    try:
        for bucket in s3_client.list_buckets().get("Buckets", []):
            for mpu_page in require_paginator(s3_client, "list_multipart_uploads").paginate(
                Bucket=bucket["Name"]
            ):
                for upload in mpu_page.get("Uploads", []):
                    if upload["Initiated"] < seven_days_ago:
                        out.append(
                            {
                                "ResourceType": "S3 Incomplete Upload",
                                "ResourceId": f"{bucket['Name']}/{upload['Key']}",
                                "Reason": "Incomplete Multipart Upload (>7 days)",
                                "Details": f"UploadId: {upload['UploadId']}, Initiated: {upload['Initiated']:%Y-%m-%d}",
                                "EstimatedMonthlySavings": "N/A",  # Cost is minor but deletion is best practice
                            }
                        )
    except Exception as e:
        log(
            "warning",
            f"Could not check for S3 incomplete uploads in {region}: {e}",
            account=alias,
        )

    # --- 6. Underutilized EC2 Instances ---
    try:
        for page in _safe_paginator(
            require_paginator(ec2_client, "describe_instances").paginate,
            account=alias,
            Filters=[{"Name": "instance-state-name", "Values": ["running"]}],
        ):
            for reservation in page.get("Reservations", []):
                for instance in reservation.get("Instances", []):
                    metrics = cw_client.get_metric_statistics(
                        Namespace="AWS/EC2",
                        MetricName="CPUUtilization",
                    Dimensions=[
                        {"Name": "InstanceId", "Value": instance["InstanceId"]}
                    ],
                    StartTime=fourteen_days_ago,
                    EndTime=now,
                    Period=86400,
                    Statistics=["Maximum"],
                )
                if (
                    metrics["Datapoints"]
                    and max(dp["Maximum"] for dp in metrics["Datapoints"]) < 5
                ):
                    cost = get_ec2_instance_cost(instance["InstanceId"], ce_client)
                    out.append(
                        {
                            "ResourceType": "EC2 Instance",
                            "ResourceId": instance["InstanceId"],
                            "Reason": "Underutilized (CPU < 5%)",
                            "Details": f"Type: {instance.get('InstanceType')}, Max CPU in 14 days: {max(dp['Maximum'] for dp in metrics['Datapoints']):.2f}%",
                            "EstimatedMonthlySavings": (
                                f"${cost:.2f}" if cost is not None else "N/A"
                            ),
                        }
                    )
    except Exception as e:
        log(
            "warning",
            f"Could not check for underutilized EC2 instances in {region}: {e}",
            account=alias,
        )

    # Add common fields to all collected opportunities
    for item in out:
        item.update({"AccountAlias": alias, "Region": region})
    return {"CostOpportunities": out}


def get_ec2_details(
    ec2_client: BaseClient,
    backup_client: BaseClient,
    alias: str,
    session: boto3.Session,
) -> List[Dict[str, Any]]:
    """Collects detailed information for EC2 instances."""
    out: List[Dict[str, Any]] = []

    instances = [
        inst
        for page in _safe_paginator(
            require_paginator(ec2_client, "describe_instances").paginate, account=alias
        )
        for r in page.get("Reservations", [])
        for inst in r.get("Instances", [])
    ]
    if not instances:
        return out

    # Batch fetch additional data to improve performance
    inst_ids = {i["InstanceId"] for i in instances}

    volume_info = {
        v["VolumeId"]: {"Size": v.get("Size"), "Type": v.get("VolumeType")}
        for chunk in chunked(sorted(inst_ids), 500)
        for v in _safe_aws_call(
            ec2_client.describe_volumes,
            default={"Volumes": []},
            account=alias,
            Filters=[{"Name": "attachment.instance-id", "Values": list(chunk)}],
        ).get("Volumes", [])
    }

    specs = {
        t: s
        for chunk in chunked(tuple({i["InstanceType"] for i in instances}), 200)
        for t, s in fetch_instance_type_specs(
            tuple(chunk), ec2_client.meta.region_name, session
        ).items()
    }

    eips = {
        addr.get("PublicIp")
        for addr in _safe_aws_call(
            ec2_client.describe_addresses, default={"Addresses": []}, account=alias
        ).get("Addresses", [])
        if addr.get("PublicIp")
    }

    # Check AWS Backup coverage
    prot_i, prot_v, plan_i, plan_v = set(), set(), set(), set()
    all_i = all_v = False
    for page in _safe_paginator(
        require_paginator(backup_client, "list_protected_resources").paginate,
        account=alias,
    ):
        for r in page.get("Results", []):
            arn = r.get("ResourceArn", "")
            if ":instance/" in arn:
                prot_i.add(arn.rsplit("/", 1)[-1])
            elif ":volume/" in arn:
                prot_v.add(arn.rsplit("/", 1)[-1])
            elif arn.endswith("/instance/*"):
                all_i = True
            elif arn.endswith("/volume/*"):
                all_v = True

    for p in _safe_aws_call(
        backup_client.list_backup_plans, default={"BackupPlansList": []}, account=alias
    )["BackupPlansList"]:
        for sel in _safe_aws_call(
            backup_client.list_backup_selections,
            default={"BackupSelectionsList": []},
            account=alias,
            BackupPlanId=p["BackupPlanId"],
        )["BackupSelectionsList"]:
            cfg = _safe_aws_call(
                backup_client.get_backup_selection,
                default={"BackupSelection": {}},
                account=alias,
                BackupPlanId=p["BackupPlanId"],
                SelectionId=sel["SelectionId"],
            )["BackupSelection"]
            for arn in cfg.get("Resources", []):
                if ":instance/" in arn:
                    plan_i.add(arn.rsplit("/", 1)[-1])
                elif ":volume/" in arn:
                    plan_v.add(arn.rsplit("/", 1)[-1])
                elif arn.endswith("/instance/*"):
                    all_i = True
                elif arn.endswith("/volume/*"):
                    all_v = True

    # Assemble the final data for each instance
    for inst in instances:
        spec = specs.get(inst["InstanceType"], {})
        attached_vols = [
            bd["Ebs"]["VolumeId"]
            for bd in inst.get("BlockDeviceMappings", [])
            if bd.get("Ebs", {}).get("VolumeId")
        ]

        is_instance_covered = (
            all_i or inst["InstanceId"] in prot_i or inst["InstanceId"] in plan_i
        )
        is_volume_covered = (
            all_v
            and attached_vols
            or any(v in prot_v or v in plan_v for v in attached_vols)
        )
        covered = is_instance_covered or is_volume_covered

        out.append(
            {
                "Name": next(
                    (
                        t["Value"]
                        for t in inst.get("Tags", [])
                        if t["Key"].lower() == "name"
                    ),
                    "",
                ),
                "InstanceId": inst["InstanceId"],
                "InstanceType": inst["InstanceType"],
                "vCPUs": spec.get("vCPUs"),
                "Memory": spec.get("Memory"),
                "OS": inst.get("PlatformDetails", "Linux/UNIX"),
                "State": inst.get("State", {}).get("Name"),
                "LaunchTime": to_local(
                    inst.get("LaunchTime"), ec2_client.meta.region_name
                ),
                "InstanceLifecycle": inst.get("InstanceLifecycle", "on-demand"),
                "PublicIP": inst.get("PublicIpAddress", ""),
                "PrivateIP": inst.get("PrivateIpAddress", ""),
                "IPType": (
                    "Elastic"
                    if inst.get("PublicIpAddress") in eips
                    else ("Ephemeral" if inst.get("PublicIpAddress") else "")
                ),
                "AvailabilityZone": inst.get("Placement", {}).get(
                    "AvailabilityZone", ""
                ),
                "EBSVolumes": ", ".join(
                    f"{v}:{volume_info.get(v,{}).get('Size','?')}GB"
                    for v in attached_vols
                ),
                "KeyPair": inst.get("KeyName", ""),
                "NetworkPerformance": spec.get("NetworkPerformance"),
                "AWSBackup": "Covered" if covered else "Not Covered",
                "Region": ec2_client.meta.region_name,
                "AccountAlias": alias,
            }
        )
    return out


def get_ec2_reserved_instances(
    ec2_client: BaseClient, alias: str
) -> list[dict[str, Any]]:
    """Collects details for active EC2 Reserved Instances."""
    out = []

    # describe_reserved_instances is not a paginated operation.
    resp = _safe_aws_call(
        ec2_client.describe_reserved_instances,
        default={"ReservedInstances": []},
        account=alias,
        Filters=[{"Name": "state", "Values": ["active"]}],
    )

    for ri in resp.get("ReservedInstances", []):
        # Enrich the response with formatted data and common fields.
        ri.update(
            {
                "Duration": seconds_to_years(ri.get("Duration")),
                "StartTime": to_local(ri.get("Start"), ec2_client.meta.region_name),
                "Region": ec2_client.meta.region_name,
                "AccountAlias": alias,
            }
        )
        out.append(ri)

    return out


def get_eventbridge_details(
    events_client: BaseClient, alias: str
) -> List[Dict[str, Any]]:
    """Collects details for EventBridge (CloudWatch Events) rules."""
    out: List[Dict[str, Any]] = []

    # list_event_buses is not paginated.
    buses = _safe_aws_call(
        events_client.list_event_buses, default={"EventBuses": []}, account=alias
    ).get("EventBuses", [])

    for bus in buses:
        bus_name = bus.get("Name", "")
        # The list_rules operation *is* paginated.
        rules = get_aws_resource_details(
            client=events_client,
            paginator_name="list_rules",
            list_key="Rules",
            alias=alias,
            EventBusName=bus_name,
        )
        for rule in rules:
            expr = rule.get("ScheduleExpression", "")
            if not expr:  # Skip rules that are not on a schedule
                continue

            # Correctly call list_targets_by_rule to get target details.
            targets = _safe_aws_call(
                events_client.list_targets_by_rule,
                default={"Targets": []},
                account=alias,
                Rule=rule["Name"],
                EventBusName=bus_name,
            )["Targets"]
            first_target = targets[0] if targets else {}

            freq, details = humanise_schedule(expr)

            out.append(
                {
                    "ScheduleName": rule.get("Name"),
                    "GroupName": bus_name,
                    "State": rule.get("State"),
                    "Frequency": freq,
                    "Expression": expr,
                    "Details": details,
                    "Timezone": _tz_for(events_client.meta.region_name),
                    "TargetArn": first_target.get("Arn", ""),
                    "Input": first_target.get("Input", ""),
                    "Region": events_client.meta.region_name,
                    "AccountAlias": alias,
                }
            )
    return out


def get_eventbridge_scheduler_details(
    scheduler_client: BaseClient, alias: str
) -> List[Dict[str, Any]]:
    """Collects details for EventBridge Scheduler schedules."""
    out: List[Dict[str, Any]] = []
    groups = {"default"} | {
        g["Name"]
        for page in _safe_paginator(
            require_paginator(scheduler_client, "list_schedule_groups").paginate,
            account=alias,
        )
        for g in page.get("ScheduleGroups", [])
    }

    for grp in groups:
        for schedule_summary in get_aws_resource_details(
            client=scheduler_client,
            paginator_name="list_schedules",
            list_key="Schedules",
            alias=alias,
            GroupName=grp,
        ):
            schedule_name = schedule_summary.get("Name", "")

            # Correctly call get_schedule to get full details including the target.
            details = _safe_aws_call(
                scheduler_client.get_schedule,
                default={},
                account=alias,
                Name=schedule_name,
                GroupName=grp,
            )
            if not details:
                continue

            expr = details.get("ScheduleExpression", "")
            freq, desc = humanise_schedule(expr)
            target = details.get("Target", {})

            out.append(
                {
                    "ScheduleName": schedule_name,
                    "GroupName": grp,
                    "State": details.get("State", ""),
                    "Frequency": freq,
                    "Expression": expr,
                    "Details": desc,
                    "Timezone": details.get(
                        "ScheduleExpressionTimezone",
                        REGION_TZ.get(scheduler_client.meta.region_name, "UTC"),
                    ),
                    "TargetArn": target.get("Arn", ""),
                    "Input": target.get("Input", ""),
                    "Region": scheduler_client.meta.region_name,
                    "AccountAlias": alias,
                }
            )
    return out


def get_governance_global_details(
    sess: boto3.Session, alias: str
) -> List[Dict[str, Any]]:
    """Checks the status of key global governance and security services."""
    global_out: List[Dict[str, Any]] = []
    # Check for CloudTrail
    try:
        ct_client = aws_client("cloudtrail", "us-east-1", sess)
        trails = ct_client.describe_trails().get("trailList", [])
        if not trails:
            global_out.append(
                {
                    "Service": "CloudTrail",
                    "Status": "No Trails Found",
                    "Details": "A multi-region trail is a security best practice.",
                }
            )
        else:
            status = (
                "Enabled (Multi-Region)"
                if any(t.get("IsMultiRegionTrail") for t in trails)
                else "Enabled (Single Region Only)"
            )
            global_out.append(
                {
                    "Service": "CloudTrail",
                    "Status": status,
                    "Details": f"{len(trails)} trail(s) found.",
                }
            )
    except Exception as e:
        global_out.append(
            {"Service": "CloudTrail", "Status": "Error", "Details": str(e)}
        )

    # Check for S3 Block Public Access
    try:
        s3_control_client = aws_client("s3control", "us-east-1", sess)
        acct_id = sess.client("sts").get_caller_identity()["Account"]
        config = s3_control_client.get_public_access_block(AccountId=acct_id).get(
            "PublicAccessBlockConfiguration", {}
        )
        status = "Enabled" if all(config.values()) else "Not Fully Enabled"
        global_out.append(
            {
                "Service": "S3 Block Public Access",
                "Status": status,
                "Details": "Account-level setting for all S3 buckets.",
            }
        )
    except ClientError as e:
        # This error means the setting has never been configured, which is a valid state.
        if e.response["Error"]["Code"] == "NoSuchPublicAccessBlockConfiguration":
            status = "Not Configured"
            details = "Account-level setting has not been applied."
            global_out.append(
                {
                    "Service": "S3 Block Public Access",
                    "Status": status,
                    "Details": details,
                }
            )
        else:
            # Handle other potential client errors
            global_out.append(
                {
                    "Service": "S3 Block Public Access",
                    "Status": "Error",
                    "Details": str(e),
                }
            )
    except Exception as e:
        global_out.append(
            {
                "Service": "S3 Block Public Access",
                "Status": "Error",
                "Details": str(e),
            }
        )

    # Add common fields for global checks
    for item in global_out:
        item.update({"AccountAlias": alias, "Region": "global"})
    return global_out


def get_governance_details(
    sess: boto3.Session, alias: str, region: str
) -> Dict[str, List[Dict[str, Any]]]:
    """Checks the status of key regional governance and security services."""
    out: List[Dict[str, Any]] = []

    regional_services_to_check = {
        "Security Hub": "securityhub",
        "GuardDuty": "guardduty",
        "AWS Config": "config",
        "Inspector": "inspector2",
        "Amazon Macie": "macie2",
        "IAM Access Analyzer": "accessanalyzer",
        "Amazon Detective": "detective",
    }

    for service_name, client_name in regional_services_to_check.items():
        try:
            client = aws_client(client_name, region, sess)
            status = "Unknown"
            details = ""
            if service_name == "Security Hub":
                client.describe_hub()
                status = "Enabled"
            elif service_name == "GuardDuty":
                detectors = client.list_detectors().get("DetectorIds", [])
                status = "Not Found"
                if detectors:
                    status = (
                        client.get_detector(DetectorId=detectors[0])
                        .get("Status", "UNKNOWN")
                        .title()
                    )
                    details = f"DetectorId: {detectors[0]}"
            elif service_name == "AWS Config":
                recorders = client.describe_configuration_recorder_status().get(
                    "ConfigurationRecordersStatus", []
                )
                status = "Not Enabled"
                if recorders:
                    status = "Recording" if recorders[0].get("recording") else "Stopped"
                    details = f"Recorder: {recorders[0].get('name')}"

            elif service_name == "Inspector":
                account_status = client.batch_get_account_status().get("accounts", [])
                if account_status:
                    state = account_status[0].get("state", {})
                    status = state.get("status", "UNKNOWN").title()
                    details = f"Resource State: {state.get('resourceState', 'N/A')}"
                else:
                    status = "Not Enabled"

            elif service_name == "Amazon Macie":
                session_info = client.get_macie_session()
                status = session_info.get("status", "UNKNOWN").title()
                details = f"Service Role: {session_info.get('serviceRole')}"
            elif service_name == "IAM Access Analyzer":
                analyzers = client.list_analyzers(type="ACCOUNT").get("analyzers", [])
                status = "Enabled" if analyzers else "Not Enabled"
                if analyzers:
                    details = f"Analyzers found: {[a.get('name') for a in analyzers]}"
            elif service_name == "Amazon Detective":
                graphs = client.list_graphs().get("GraphList", [])
                status = "Enabled" if graphs else "Not Enabled"
                if graphs:
                    details = f"Graph ARN: {graphs[0].get('Arn')}"

            out.append({"Service": service_name, "Status": status, "Details": details})

        # Gracefully handle regions where a service endpoint does not exist.
        except EndpointConnectionError:
            out.append(
                {
                    "Service": service_name,
                    "Status": "Not Available",
                    "Details": f"Service endpoint does not exist in region {region}.",
                }
            )

        except ClientError as e:
            if e.response["Error"]["Code"] in [
                "InvalidAccessException",
                "AccessDeniedException",
                "ResourceNotFoundException",
            ]:
                out.append(
                    {
                        "Service": service_name,
                        "Status": "Not Enabled",
                        "Details": "Service not subscribed or enabled in this region.",
                    }
                )
            else:
                out.append(
                    {"Service": service_name, "Status": "Error", "Details": str(e)}
                )
        except Exception as e:
            out.append(
                {
                    "Service": service_name,
                    "Status": "Error",
                    "Details": f"Could not check status: {e}",
                }
            )

    for item in out:
        item.update({"AccountAlias": alias, "Region": region})
    return {"Governance": out}


def get_iam_details(
    iam_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """
    Unified collector for all IAM resources. This function calls sub-collectors
    for users, roles, groups, and policies to provide a comprehensive view.
    """
    return {
        "IAMUsers": get_iam_users_details(iam_client, alias),
        "IAMRoles": get_iam_roles_details(iam_client, alias),
        "IAMGroups": get_iam_groups_details(iam_client, alias),
        "IAMPolicies": get_iam_policies_details(iam_client, alias),
    }


def get_iam_users_details(iam_client: BaseClient, alias: str) -> List[Dict[str, Any]]:
    """Collects details for IAM Users, including their groups and policies."""
    out = []
    users = get_aws_resource_details(
        client=iam_client, paginator_name="list_users", list_key="Users", alias=alias
    )
    for user in users:
        user_name = user["UserName"]

        # Enrich with associated groups and policies
        user["Groups"] = [
            g["GroupName"]
            for page in _safe_paginator(
                iam_client.get_paginator("list_groups_for_user").paginate,
                account=alias,
                UserName=user_name,
            )
            for g in page.get("Groups", [])
        ]
        user["AttachedPolicies"] = [
            p["PolicyArn"]
            for page in _safe_paginator(
                iam_client.get_paginator("list_attached_user_policies").paginate,
                account=alias,
                UserName=user_name,
            )
            for p in page.get("AttachedPolicies", [])
        ]
        inline_policies: List[str] = []
        for page in _safe_paginator(
            require_paginator(iam_client, "list_user_policies").paginate,
            account=alias,
            UserName=user_name,
        ):
            inline_policies.extend(page.get("PolicyNames", []))
        user["InlinePolicies"] = sorted(set(inline_policies))

        # Format timestamps
        user["CreateDate"] = to_local(
            user.get("CreateDate"), iam_client.meta.region_name
        )
        user["PasswordLastUsed"] = to_local(
            user.get("PasswordLastUsed"), iam_client.meta.region_name
        )

        out.append(user)
    return out


def get_iam_roles_details(iam_client: BaseClient, alias: str) -> List[Dict[str, Any]]:
    """Collects details for IAM Roles, including their trust policies and attached policies."""
    out = []
    roles = get_aws_resource_details(
        client=iam_client, paginator_name="list_roles", list_key="Roles", alias=alias
    )
    for role in roles:
        role_name = role["RoleName"]

        # Correctly parse the trust policy and collect all principals
        trust_policy = role.get("AssumeRolePolicyDocument", {})
        statements = trust_policy.get("Statement", [])
        if isinstance(statements, dict):
            statements = [statements]

        service_principals: Set[str] = set()
        account_principals: Set[str] = set()
        federated_principals: Set[str] = set()

        for stmt in statements:
            principals = stmt.get("Principal", {})
            if principals == "*":
                account_principals.add("*")
                continue
            if isinstance(principals, str):
                account_principals.add(principals)
                continue
            services = principals.get("Service", [])
            aws_accounts = principals.get("AWS", [])
            federated = principals.get("Federated", [])

            if isinstance(services, str):
                service_principals.add(services)
            else:
                service_principals.update(services)

            if isinstance(aws_accounts, str):
                account_principals.add(aws_accounts)
            else:
                account_principals.update(aws_accounts)

            if isinstance(federated, str):
                federated_principals.add(federated)
            else:
                federated_principals.update(federated)

        all_principals = service_principals | account_principals | federated_principals

        role["ServicePrincipals"] = sorted(all_principals)
        role["AccountPrincipals"] = sorted(account_principals)
        role["FederatedPrincipals"] = sorted(federated_principals)

        # Enrich with attached and inline policies
        role["AttachedPolicies"] = [
            p["PolicyArn"]
            for page in _safe_paginator(
                iam_client.get_paginator("list_attached_role_policies").paginate,
                account=alias,
                RoleName=role_name,
            )
            for p in page.get("AttachedPolicies", [])
        ]
        inline_policies = []
        for page in _safe_paginator(
            require_paginator(iam_client, "list_role_policies").paginate,
            account=alias,
            RoleName=role_name,
        ):
            inline_policies.extend(page.get("PolicyNames", []))
        role["InlinePolicies"] = sorted(set(inline_policies))

        # Format timestamp
        role["CreateDate"] = to_local(
            role.get("CreateDate"), "us-east-1"
        )  # IAM is global

        out.append(role)
    return out


def get_iam_groups_details(iam_client: BaseClient, alias: str) -> List[Dict[str, Any]]:
    """Collects details for IAM Groups, including their members and policies."""
    out = []
    groups = get_aws_resource_details(
        client=iam_client, paginator_name="list_groups", list_key="Groups", alias=alias
    )
    for group in groups:
        group_name = group["GroupName"]

        # Enrich with members and policies
        group["Members"] = [
            u["UserName"]
            for u in _safe_aws_call(
                iam_client.get_group, default={"Users": []}, GroupName=group_name
            ).get("Users", [])
        ]
        group["AttachedPolicies"] = [
            p["PolicyArn"]
            for page in _safe_paginator(
                iam_client.get_paginator("list_attached_group_policies").paginate,
                account=alias,
                GroupName=group_name,
            )
            for p in page.get("AttachedPolicies", [])
        ]
        inline_policies = []
        for page in _safe_paginator(
            require_paginator(iam_client, "list_group_policies").paginate,
            account=alias,
            GroupName=group_name,
        ):
            inline_policies.extend(page.get("PolicyNames", []))
        group["InlinePolicies"] = sorted(set(inline_policies))

        # Format timestamp
        group["CreateDate"] = to_local(
            group.get("CreateDate"), iam_client.meta.region_name
        )

        out.append(group)
    return out


def get_iam_policies_details(
    iam_client: BaseClient, alias: str
) -> List[Dict[str, Any]]:
    """Collects details for both AWS Managed and Customer Managed IAM Policies."""
    out = []
    # Check both 'Local' (Customer-Managed) and 'AWS' (AWS-Managed) policies.
    for scope in ["Local", "AWS"]:
        policies = get_aws_resource_details(
            client=iam_client,
            paginator_name="list_policies",
            list_key="Policies",
            alias=alias,
            Scope=scope,
        )
        for policy in policies:
            # To reduce noise, only show AWS-managed policies that are actually in use.
            if scope == "AWS" and policy.get("AttachmentCount", 0) == 0:
                continue

            # Get the policy document (the JSON defining permissions).
            try:
                version_id = policy["DefaultVersionId"]
                ver = iam_client.get_policy_version(
                    PolicyArn=policy["Arn"], VersionId=version_id
                )
                policy["PolicyDocument"] = ver["PolicyVersion"]["Document"]
            except Exception:
                policy["PolicyDocument"] = (
                    {}
                )  # Handle cases where version info is not available

            # Get all entities (users, roles, groups) the policy is attached to.
            entities = {"PolicyUsers": [], "PolicyRoles": [], "PolicyGroups": []}
            for page in _safe_paginator(
                iam_client.get_paginator("list_entities_for_policy").paginate,
                account=alias,
                PolicyArn=policy["Arn"],
            ):
                entities["PolicyUsers"].extend(page.get("PolicyUsers", []))
                entities["PolicyRoles"].extend(page.get("PolicyRoles", []))
                entities["PolicyGroups"].extend(page.get("PolicyGroups", []))
            policy["AttachmentEntities"] = (
                [
                    {"Type": "User", "Name": e["UserName"]}
                    for e in entities.get("PolicyUsers", [])
                ]
                + [
                    {"Type": "Role", "Name": e["RoleName"]}
                    for e in entities.get("PolicyRoles", [])
                ]
                + [
                    {"Type": "Group", "Name": e["GroupName"]}
                    for e in entities.get("PolicyGroups", [])
                ]
            )

            # Add final metadata
            policy["PolicyType"] = (
                "Customer Managed" if scope == "Local" else "AWS Managed"
            )
            policy["CreateDate"] = to_local(policy.get("CreateDate"), "us-east-1")
            policy["UpdateDate"] = to_local(policy.get("UpdateDate"), "us-east-1")

            out.append(policy)
    return out


def get_iam_summary_details(
    users: List[Dict[str, Any]],
    roles: List[Dict[str, Any]],
    groups: List[Dict[str, Any]],
    alias: str,
) -> List[Dict[str, Any]]:
    """
    Creates a consolidated summary of all IAM principals from pre-fetched data.
    """
    out: List[Dict[str, Any]] = []

    for user in users:
        out.append(
            {
                "PrincipalName": user.get("UserName"),
                "PrincipalType": "User",
                "Arn": user.get("Arn"),
                "CreateDate": user.get("CreateDate"),
                "AttachedPoliciesCount": len(user.get("AttachedPolicies", [])),
                "InlinePoliciesCount": len(user.get("InlinePolicies", [])),
                "Details": f"Password Last Used: {user.get('PasswordLastUsed', 'N/A')}",
                "AccountAlias": alias,
            }
        )

    for role in roles:
        out.append(
            {
                "PrincipalName": role.get("RoleName"),
                "PrincipalType": "Role",
                "Arn": role.get("Arn"),
                "CreateDate": role.get("CreateDate"),
                "AttachedPoliciesCount": len(role.get("AttachedPolicies", [])),
                "InlinePoliciesCount": len(role.get("InlinePolicies", [])),
                "Details": (
                    f"Service Principals: {role.get('ServicePrincipals', [])}; "
                    f"Account Principals: {role.get('AccountPrincipals', [])}; "
                    f"Federated Principals: {role.get('FederatedPrincipals', [])}"
                ),
                "AccountAlias": alias,
            }
        )

    for group in groups:
        out.append(
            {
                "PrincipalName": group.get("GroupName"),
                "PrincipalType": "Group",
                "Arn": group.get("Arn"),
                "CreateDate": group.get("CreateDate"),
                "AttachedPoliciesCount": len(group.get("AttachedPolicies", [])),
                "InlinePoliciesCount": len(group.get("InlinePolicies", [])),
                "Details": f"Members: {len(group.get('Members', []))}",
                "AccountAlias": alias,
            }
        )

    return {"IAM_Summary": out}


def get_kms_details(
    kms_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """Collects details for KMS keys, including aliases, tags, and rotation status."""
    out: List[Dict[str, Any]] = []

    # List all key summaries first.
    keys = get_aws_resource_details(
        client=kms_client, paginator_name="list_keys", list_key="Keys", alias=alias
    )

    for key_summary in keys:
        key_id = key_summary["KeyId"]

        # Get the core metadata for the key.
        meta = _safe_aws_call(
            kms_client.describe_key, default={}, account=alias, KeyId=key_id
        ).get("KeyMetadata", {})

        if not meta:
            continue

        # For customer-managed keys, get additional details.
        if meta.get("KeyManager") != "AWS":
            meta["RotationEnabled"] = _safe_aws_call(
                kms_client.get_key_rotation_status,
                default={"KeyRotationEnabled": False},
                account=alias,
                KeyId=key_id,
            ).get("KeyRotationEnabled", False)

            # Use the standardized format_tags helper
            tags_list = [
                t
                for page in _safe_paginator(
                    require_paginator(kms_client, "list_resource_tags").paginate,
                    account=alias,
                    KeyId=key_id,
                )
                for t in page.get("Tags", [])
            ]
            meta["Tags"] = format_tags(tags_list)
        else:
            meta["RotationEnabled"] = False
            meta["Tags"] = {}

        # Get aliases and grant information for all key types.
        meta["AliasNames"] = [
            a["AliasName"]
            for page in _safe_paginator(
                require_paginator(kms_client, "list_aliases").paginate,
                account=alias,
                KeyId=key_id,
            )
            for a in page.get("Aliases", [])
            if a.get("AliasName", "").startswith("alias/")
        ]

        meta["GrantsCount"] = sum(
            len(p.get("Grants", []))
            for p in _safe_paginator(
                require_paginator(kms_client, "list_grants").paginate,
                account=alias,
                KeyId=key_id,
            )
        )

        # Format timestamps and add common fields.
        meta["CreationDate"] = to_local(
            meta.get("CreationDate"), kms_client.meta.region_name
        )
        meta["DeletionDate"] = to_local(
            meta.get("DeletionDate"), kms_client.meta.region_name
        )
        meta["ValidTo"] = to_local(meta.get("ValidTo"), kms_client.meta.region_name)

        out.append(meta)

    return {"KMS": out}


def get_lightsail_inventory(
    ls_client: BaseClient,
    alias: str,
    session: boto3.Session,
    supported_regions: set[str],
) -> Dict[str, List[Dict[str, Any]]]:
    """Collects details for all regional Lightsail resources."""
    out: List[Dict[str, Any]] = []
    current_region = ls_client.meta.region_name

    if supported_regions and current_region not in supported_regions:
        return {"Lightsail": []}

    # --- Pre-fetch Bundle Specifications for the region ---
    bundle_specs = {}
    try:
        # Fetch Instance and Database Bundles
        for page in _safe_paginator(
            require_paginator(ls_client, "get_bundles").paginate, account=alias
        ):
            for bundle in page.get("bundles", []):
                bundle_specs[bundle["bundleId"]] = bundle
        for page in _safe_paginator(
            require_paginator(ls_client, "get_relational_database_bundles").paginate,
            account=alias,
        ):
            for bundle in page.get("bundles", []):
                bundle_specs[bundle["bundleId"]] = bundle
    except Exception as e:
        log(
            "warning",
            f"Could not fetch all Lightsail bundle specs in {current_region}: {e}",
            account=alias,
        )

    # --- Regional Resources ---
    # Instances
    for inst in get_aws_resource_details(
        client=ls_client,
        paginator_name="get_instances",
        list_key="instances",
        alias=alias,
    ):
        specs = bundle_specs.get(inst.get("bundleId"), {})
        inst.update(
            {
                "ResourceType": "Instance",
                "vCPUs": specs.get("cpuCount"),
                "MemoryInGB": specs.get("ramSizeInGb"),
                "DiskSizeGB": specs.get("diskSizeInGb"),
                "DataTransferGB": specs.get("transferPerMonthInGb"),
                "CreatedAt": to_local(inst.get("createdAt"), current_region),
            }
        )
        out.append(inst)

    # Databases
    for db in get_aws_resource_details(
        client=ls_client,
        paginator_name="get_relational_databases",
        list_key="relationalDatabases",
        alias=alias,
    ):
        specs = bundle_specs.get(db.get("relationalDatabaseBundleId"), {})
        db.update(
            {
                "ResourceType": "Database",
                "vCPUs": specs.get("cpuCount"),
                "MemoryInGB": specs.get("ramSizeInGb"),
                "DiskSizeGB": specs.get("diskSizeInGb"),
                "CreatedAt": to_local(db.get("createdAt"), current_region),
            }
        )
        out.append(db)

    # Disks and Static IPs
    for disk in get_aws_resource_details(
        client=ls_client, paginator_name="get_disks", list_key="disks", alias=alias
    ):
        disk.update(
            {
                "ResourceType": "Disk",
                "CreatedAt": to_local(disk.get("createdAt"), current_region),
            }
        )
        out.append(disk)
    for ip in get_aws_resource_details(
        client=ls_client,
        paginator_name="get_static_ips",
        list_key="staticIps",
        alias=alias,
    ):
        ip.update(
            {
                "ResourceType": "StaticIp",
                "CreatedAt": to_local(ip.get("createdAt"), current_region),
            }
        )
        out.append(ip)

    # Certificates are regional
    for cert in _safe_aws_call(
        ls_client.get_certificates, account=alias, default={}
    ).get("certificates", []):
        cert.update(
            {
                "ResourceType": "Certificate",
                "CreatedAt": to_local(cert.get("createdAt"), current_region),
                "ExpiresAt": to_local(cert.get("notAfter"), current_region),
                "Region": current_region,
                "AccountAlias": alias,
            }
        )
        out.append(cert)
    return {"Lightsail": out}


def get_lightsail_global_details(
    ls_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """Collects details for global Lightsail resources like domains."""
    out: List[Dict[str, Any]] = []
    try:
        # Note: This client should be for us-east-1 as it's a global service endpoint
        for domain in get_aws_resource_details(
            client=ls_client,
            paginator_name="get_domains",
            list_key="domains",
            alias=alias,
        ):
            domain.update(
                {
                    "ResourceType": "Domain",
                    "CreatedAt": to_local(domain.get("createdAt"), "us-east-1"),
                    "Region": "global",
                    "AccountAlias": alias,
                }
            )
            out.append(domain)
    except Exception as e:
        log("error", f"Failed to fetch Lightsail domains: {e}", account=alias)

    # Return a dictionary consistent with the standardized design
    return {"Lightsail": out}


def get_rds_details(
    rds_client: BaseClient, alias: str, session: boto3.Session
) -> List[Dict[str, Any]]:
    """Collects details for RDS database instances."""
    out: List[Dict[str, Any]] = []

    # First, get a list of all DB instances.
    all_dbs = get_aws_resource_details(
        client=rds_client,
        paginator_name="describe_db_instances",
        list_key="DBInstances",
        alias=alias,
    )
    if not all_dbs:
        return out

    # To improve performance, gather all unique instance classes and
    # fetch their specifications in a single batch.
    instance_classes = {db["DBInstanceClass"].removeprefix("db.") for db in all_dbs}
    specs: Dict[str, Any] = {}
    if instance_classes:
        for chunk in chunked(sorted(instance_classes), 20):
            specs.update(
                fetch_instance_type_specs(
                    tuple(chunk), rds_client.meta.region_name, session
                )
            )

    for db in all_dbs:
        arn = db["DBInstanceArn"]

        # Look up the pre-fetched instance specifications.
        instance_class = db["DBInstanceClass"]
        spec = specs.get(instance_class.removeprefix("db."), {})

        # Enrich the database details with the gathered information.
        db.update(
            {
                "vCPUs": spec.get("vCPUs"),
                "Memory": spec.get("Memory"),
                # Use the standardized format_tags helper
                "Tags": format_tags(
                    _safe_aws_call(
                        rds_client.list_tags_for_resource,
                        default={"TagList": []},
                        account=alias,
                        ResourceName=arn,
                    )["TagList"]
                ),
                "EndpointAddress": db.get("Endpoint", {}).get("Address", ""),
                "EndpointPort": db.get("Endpoint", {}).get("Port", ""),
                "InstanceCreateTime": to_local(
                    db.get("InstanceCreateTime"), rds_client.meta.region_name
                ),
            }
        )
        out.append(db)

    return out


def get_rds_reserved_instances(
    rds_client: BaseClient, alias: str
) -> List[Dict[str, Any]]:
    """Collects details for active RDS Reserved Instances."""
    out = []

    # Use the generic helper to list all reserved DB instances.
    instances = get_aws_resource_details(
        client=rds_client,
        paginator_name="describe_reserved_db_instances",
        list_key="ReservedDBInstances",
        alias=alias,
    )

    for ri in instances:
        # Enrich the response with formatted data.
        ri["Duration"] = seconds_to_years(ri.get("Duration"))
        ri["StartTime"] = to_local(ri.get("StartTime"), rds_client.meta.region_name)
        out.append(ri)

    return out


def get_route53_details(
    r53_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """Collects details for Route 53 Hosted Zones."""
    out: List[Dict[str, Any]] = []

    # list_hosted_zones uses manual pagination with a Marker and does not accept a Region parameter.
    next_marker = None
    while True:
        params = {"Marker": next_marker} if next_marker else {}
        resp = _safe_aws_call(
            r53_client.list_hosted_zones,
            default={"HostedZones": []},
            account=alias,
            **params,
        )

        for z in resp.get("HostedZones", []):
            zid = z["Id"].split("/")[-1]

            # Enrich with additional details for each zone
            det = _safe_aws_call(
                r53_client.get_hosted_zone, default={}, account=alias, Id=zid
            )
            z["VPCAssociations"] = det.get("VPCs", [])
            z["DelegationSet"] = det.get("DelegationSet", {}).get("NameServers", [])

            sec_status = _safe_aws_call(
                r53_client.get_dnssec, default={}, account=alias, HostedZoneId=zid
            ).get("Status", {})
            z["DNSSECStatus"] = (
                f'{sec_status.get("Status","")}/{sec_status.get("ServeSignature","")}'.strip(
                    "/"
                )
            )

            z["Tags"] = {
                t["Key"]: t["Value"]
                for t in _safe_aws_call(
                    r53_client.list_tags_for_resource,
                    default={"ResourceTagSet": {"Tags": []}},
                    account=alias,
                    ResourceType="hostedzone",
                    ResourceId=zid,
                )["ResourceTagSet"]["Tags"]
            }

            record_types, health_checks = set(), set()
            for page in _safe_paginator(
                require_paginator(r53_client, "list_resource_record_sets").paginate,
                account=alias,
                HostedZoneId=zid,
            ):
                for rr in page.get("ResourceRecordSets", []):
                    record_types.add(rr.get("Type", ""))
                    if "HealthCheckId" in rr:
                        health_checks.add(rr["HealthCheckId"])

            z["RecordTypes"] = sorted(record_types)
            z["HealthChecks"] = sorted(health_checks)
            z["Name"] = z.get("Name", "").rstrip(".")
            z["Region"] = "global"  # Set region manually for global services
            z["AccountAlias"] = alias
            out.append(z)

        next_marker = resp.get("NextMarker")
        if not next_marker:
            break

    return {"Route53": out}


def get_s3_details(
    s3_client: BaseClient,
    all_buckets: List[Dict[str, Any]],
    alias: str,
    region: str,
    session: boto3.Session,
) -> Dict[str, List[Dict[str, Any]]]:
    """Collects detailed information for S3 buckets in a specific region."""
    out: List[Dict[str, Any]] = []
    now = datetime.now(timezone.utc)

    # Filter the global list of buckets to only those in the current region.
    buckets_in_region = [
        b for b in all_buckets if bucket_region(b["Name"], region, session) == region
    ]
    if not buckets_in_region:
        return {"S3": []}

    cw_client = aws_client("cloudwatch", region, session)

    # --- Metric Gathering Strategy ---
    # 1. Prepare CloudWatch Metric Queries
    metric_queries, id_map = [], {}
    for idx, b in enumerate(buckets_in_region):
        for metric, stype, key in [
            ("BucketSizeBytes", "StandardStorage", "size"),
            ("NumberOfObjects", "AllStorageTypes", "count"),
        ]:
            qid = f"b{idx:05}{key[0]}"
            id_map[qid] = (b["Name"], key)
            metric_queries.append(
                {
                    "Id": qid,
                    "MetricStat": {
                        "Metric": {
                            "Namespace": "AWS/S3",
                            "MetricName": metric,
                            "Dimensions": [
                                {"Name": "BucketName", "Value": b["Name"]},
                                {"Name": "StorageType", "Value": stype},
                            ],
                        },
                        "Period": _METRIC_PERIOD,
                        "Stat": "Average",
                    },
                    "ReturnData": True,
                }
            )

    # 2. Execute CloudWatch queries
    metric_results: List[Dict[str, Any]] = []
    for chunk in chunked(metric_queries, 500):
        for page in _safe_paginator(
            require_paginator(cw_client, "get_metric_data").paginate,
            account=alias,
            MetricDataQueries=chunk,
            StartTime=(now - timedelta(hours=_METRIC_OFFSET + 2)),
            EndTime=now,
            ScanBy="TimestampDescending",
        ):
            metric_results.extend(page.get("MetricDataResults", []))

    cw_data = {b["Name"]: {"size": None, "count": None} for b in buckets_in_region}
    for r in metric_results:
        if vals := r.get("Values"):
            name, key = id_map[r["Id"]]
            cw_data[name][key] = int(vals[0])

    # --- Bucket Details Collection ---
    for b in buckets_in_region:
        name = b["Name"]
        size, count = cw_data[name]["size"], cw_data[name]["count"]
        method = "CloudWatch" if size is not None and count is not None else None

        # 3. Fallback Metric Strategies
        if method is None:
            size, count = storage_lens_metrics(name, region, session)
            method = "StorageLens" if size is not None or count is not None else None
        if method is None:
            size, count = inventory_metrics(name, region, alias, session)
            method = "Inventory" if size is not None or count is not None else None

        if method is None:
            try:
                partial_scan = False
                temp_size, temp_count = 0, 0
                paginator = require_paginator(s3_client, "list_objects_v2")
                for page in _safe_paginator(
                    paginator.paginate, account=alias, Bucket=name
                ):
                    for obj in page.get("Contents", []):
                        temp_size += obj.get("Size", 0)
                        temp_count += 1
                        if temp_count >= MAX_KEYS_FOR_FULL_SCAN:
                            partial_scan = True
                            break
                    if partial_scan:
                        break

                size, count = temp_size, temp_count
                if count > 0:
                    method = (
                        "Partial Scan (ListObjects)"
                        if partial_scan
                        else "Full Scan (ListObjects)"
                    )

            except Exception as e:
                log(
                    "warning",
                    f"list_objects_v2 fallback for {name} failed: {e}",
                    account=alias,
                )

        # 4. Gather other bucket configurations
        versioning = _safe_aws_call(
            s3_client.get_bucket_versioning, default={}, account=alias, Bucket=name
        ).get("Status", "Disabled")
        encryption = (
            "Enabled"
            if "ServerSideEncryptionConfiguration"
            in _safe_aws_call(
                s3_client.get_bucket_encryption, default={}, account=alias, Bucket=name
            )
            else "Not enabled"
        )
        pab_config = _safe_aws_call(
            s3_client.get_public_access_block, default={}, account=alias, Bucket=name
        ).get("PublicAccessBlockConfiguration", {})
        public_access = "Blocked" if all(pab_config.values()) else "Not fully blocked"
        policy_status = (
            "Public"
            if _safe_aws_call(
                s3_client.get_bucket_policy_status,
                default={"PolicyStatus": {"IsPublic": False}},
                account=alias,
                Bucket=name,
            )["PolicyStatus"].get("IsPublic")
            else "Not Public"
        )
        lifecycle_rules = _safe_aws_call(
            s3_client.get_bucket_lifecycle_configuration,
            default={"Rules": []},
            account=alias,
            Bucket=name,
        ).get("Rules", [])

        tags = format_tags(
            _safe_aws_call(
                s3_client.get_bucket_tagging,
                default={"TagSet": []},
                account=alias,
                Bucket=name,
            )["TagSet"]
        )

        out.append(
            {
                "BucketName": name,
                "Region": region,
                "CreationDate": to_local(b["CreationDate"], region),
                "Size": human_size(size),
                "ObjectCount": count,
                "LastMetricsUpdate": to_local(now, region) if method else "",
                "MetricsCalculationMethod": method or "N/A",
                "Versioning": versioning,
                "Encryption": encryption,
                "PublicAccess": public_access,
                "PolicyStatus": policy_status,
                "LifecycleRules": lifecycle_rules,
                "Tags": tags,
                "AccountAlias": alias,
            }
        )
    return {"S3": out}


def get_lambda_details(
    lambda_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """Collects details for Lambda functions, including their configuration and tags."""
    out: List[Dict[str, Any]] = []

    # List all functions in the region.
    functions = get_aws_resource_details(
        client=lambda_client,
        paginator_name="list_functions",
        list_key="Functions",
        alias=alias,
    )

    for fn in functions:
        name, arn = fn["FunctionName"], fn["FunctionArn"]

        # Get additional configuration details for each function.
        cfg = _safe_aws_call(
            lambda_client.get_function_configuration,
            default={},
            account=alias,
            FunctionName=name,
        )

        # Get the tags for each function.
        tag_resp = _safe_aws_call(
            lambda_client.list_tags,
            default={"Tags": {}},
            account=alias,
            Resource=arn,
        )

        out.append(
            {
                "FunctionName": name,
                "FunctionArn": arn,
                "State": cfg.get("State", ""),
                "LastUpdateStatus": cfg.get("LastUpdateStatus", ""),
                "Runtime": fn.get("Runtime", ""),
                "Handler": fn.get("Handler", ""),
                "Role": fn.get("Role", ""),
                "Description": fn.get("Description", ""),
                "MemorySize": fn.get("MemorySize"),
                "Timeout": fn.get("Timeout"),
                "PackageType": fn.get("PackageType", ""),
                "Architectures": fn.get("Architectures", []),
                "TracingMode": cfg.get("TracingConfig", {}).get("Mode", ""),
                "LastModified": to_local(
                    fn.get("LastModified"), lambda_client.meta.region_name
                ),
                "KMSKeyArn": cfg.get("KMSKeyArn", ""),
                "CodeSize": human_size(fn.get("CodeSize", 0)),
                "VpcSecurityGroupIds": fn.get("VpcConfig", {}).get(
                    "SecurityGroupIds", []
                ),
                "VpcSubnetIds": fn.get("VpcConfig", {}).get("SubnetIds", []),
                "EnvironmentVars": fn.get("Environment", {}).get("Variables", {}),
                "Tags": tag_resp.get("Tags", {}),
                "Region": lambda_client.meta.region_name,
                "AccountAlias": alias,
            }
        )

    return {"Lambda": out}


def get_savings_plan_details(
    sp_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """Collects details for AWS Savings Plans across all relevant states."""
    out: List[Dict[str, Any]] = []

    # Explicitly request plans in all states to ensure none are missed.
    # The API defaults to 'active' only. We include 'retired' to get a complete history.
    savings_plan_ids = [
        sp["savingsPlanId"]
        for page in require_paginator(sp_client, "describe_savings_plans").paginate()
        for sp in page.get("savingsPlans", [])
    ]

    if not savings_plan_ids:
        return {"SavingsPlans": []}

    # Describe the plans in chunks to avoid overly large API requests.
    for chunk in chunked(savings_plan_ids, 10):
        resp = _safe_aws_call(
            sp_client.describe_savings_plans,
            savingsPlanIds=chunk,
            default={},
            account=alias,
        )

        for plan in resp.get("savingsPlans", []):
            start_str = plan.get("start")
            end_str = plan.get("end")

            # Format timestamps and calculate the term duration.
            plan["Start"] = to_local(start_str, "us-east-1")  # Savings Plans are global
            plan["End"] = to_local(end_str, "us-east-1")

            if start_str and end_str:
                try:
                    start_dt = datetime.fromisoformat(start_str.replace("Z", "+00:00"))
                    end_dt = datetime.fromisoformat(end_str.replace("Z", "+00:00"))
                    plan["Term"] = seconds_to_years((end_dt - start_dt).total_seconds())
                except (ValueError, TypeError):
                    plan["Term"] = "N/A"
            else:
                plan["Term"] = seconds_to_years(plan.get("termDurationInSeconds"))

            plan["Region"] = "global"
            plan["AccountAlias"] = alias
            out.append(plan)

    return {"SavingsPlans": out}


def get_ses_details(
    ses_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """Collects details for SES identities and their verification status."""
    out: List[Dict[str, Any]] = []

    # List all identities across pages.
    identities: List[str] = []
    for page in _safe_paginator(
        require_paginator(ses_client, "list_identities").paginate,
        account=alias,
    ):
        identities.extend(page.get("Identities", []))

    # Deduplicate and sort before fetching verification status.
    identities = sorted(set(identities))

    if not identities:
        return {"SES": []}

    # Get verification attributes for all identities in a single batch call.
    verification_attributes = _safe_aws_call(
        ses_client.get_identity_verification_attributes,
        default={"VerificationAttributes": {}},
        account=alias,
        Identities=identities,
    ).get("VerificationAttributes", {})

    for identity in identities:
        status = verification_attributes.get(identity, {}).get(
            "VerificationStatus", "Unknown"
        )
        out.append(
            {
                "Identity": identity,
                "VerificationStatus": status,
                "Region": ses_client.meta.region_name,
                "AccountAlias": alias,
            }
        )

    return {"SES": out}


def get_sns_details(
    sns_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """Collects details for SNS topics and their subscriptions."""
    out: List[Dict[str, Any]] = []

    # List all topics in the region.
    topics = get_aws_resource_details(
        client=sns_client, paginator_name="list_topics", list_key="Topics", alias=alias
    )

    for topic in topics:
        arn = topic["TopicArn"]

        # For each topic, list all of its subscriptions.
        subscriptions = [
            f"{s.get('Protocol','')}://{s.get('Endpoint','')}"
            for page in _safe_paginator(
                require_paginator(sns_client, "list_subscriptions_by_topic").paginate,
                account=alias,
                TopicArn=arn,
            )
            for s in page.get("Subscriptions", [])
        ]

        # Assemble the final record for the report.
        out.append(
            {
                "TopicArn": arn,
                "TopicName": arn.split(":")[-1],
                "SubscriptionCount": len(subscriptions),
                "Subscriptions": subscriptions,
                "Region": sns_client.meta.region_name,
                "AccountAlias": alias,
            }
        )

    return {"SNS": out}


def get_vpc_inventory(
    ec2_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """
    Unified collector for all core VPC resources, including VPCs, Subnets,
    Route Tables, and Security Groups.
    """

    # --- VPCs ---
    vpcs = get_aws_resource_details(
        client=ec2_client, paginator_name="describe_vpcs", list_key="Vpcs", alias=alias
    )
    for vpc in vpcs:
        vpc["Tags"] = format_tags(vpc.get("Tags", []))

    # --- Subnets ---
    subnets = get_aws_resource_details(
        client=ec2_client,
        paginator_name="describe_subnets",
        list_key="Subnets",
        alias=alias,
    )
    for subnet in subnets:
        subnet["Tags"] = format_tags(subnet.get("Tags", []))

    # --- Route Tables ---
    route_tables = []
    for rt in get_aws_resource_details(
        client=ec2_client,
        paginator_name="describe_route_tables",
        list_key="RouteTables",
        alias=alias,
    ):
        rt["IsMain"] = any(a.get("Main") for a in rt.get("Associations", []))
        rt["Tags"] = format_tags(rt.get("Tags", []))
        route_tables.append(rt)

    # --- Security Groups ---
    security_groups = []
    for sg in get_aws_resource_details(
        client=ec2_client,
        paginator_name="describe_security_groups",
        list_key="SecurityGroups",
        alias=alias,
    ):
        sg["Tags"] = format_tags(sg.get("Tags", []))
        security_groups.append(sg)

    return {
        "VPCs": vpcs,
        "Subnets": subnets,
        "RouteTables": route_tables,
        "SecurityGroups": security_groups,
    }


def get_vpn_details(
    ec2_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """Collects details for AWS Site-to-Site VPN connections, including tunnel status."""
    out: List[Dict[str, Any]] = []

    # describe_vpn_connections uses manual pagination.
    next_token = None
    while True:
        params = {"NextToken": next_token} if next_token else {}
        resp = _safe_aws_call(
            ec2_client.describe_vpn_connections,
            default={"VpnConnections": []},
            account=alias,
            **params,
        )

        for conn in resp.get("VpnConnections", []):
            # The customer gateway configuration is an XML string that needs to be parsed.
            xml_config = conn.get("CustomerGatewayConfiguration", "")

            # The parse_vpn_config helper extracts the gateway IP from the XML.
            customer_gw_ip, _ = (
                parse_vpn_config(xml_config) if xml_config else ("N/A", [])
            )

            # Extract enhanced tunnel details including status, message, and last change time.
            tunnels_info = []
            for telemetry in conn.get("VgwTelemetry", []):
                tunnels_info.append(
                    {
                        "OutsideIpAddress": telemetry.get("OutsideIpAddress"),
                        "Status": telemetry.get("Status", "UNKNOWN").upper(),
                        "StatusMessage": telemetry.get("StatusMessage", ""),
                        "LastStatusChange": to_local(
                            telemetry.get("LastStatusChange"),
                            ec2_client.meta.region_name,
                        ),
                    }
                )

            out.append(
                {
                    "VpnConnectionId": conn.get("VpnConnectionId"),
                    "Name": next(
                        (
                            t["Value"]
                            for t in conn.get("Tags", [])
                            if t["Key"].lower() == "name"
                        ),
                        "",
                    ),
                    "State": conn.get("State"),
                    "CustomerGatewayId": conn.get("CustomerGatewayId"),
                    "CustomerGatewaySource": customer_gw_ip,
                    "TunnelDetails": tunnels_info,
                    "Region": ec2_client.meta.region_name,
                    "AccountAlias": alias,
                }
            )

        next_token = resp.get("NextToken")
        if not next_token:
            break

    return {"VPN": out}


def get_waf_classic_details(
    wafc_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """Collects details for WAF Classic Web ACLs."""
    out: List[Dict[str, Any]] = []

    # list_web_acls for WAF Classic requires manual pagination using NextMarker.
    next_marker = None
    while True:
        params = {"NextMarker": next_marker} if next_marker else {}
        resp = _safe_aws_call(
            wafc_client.list_web_acls, default={"WebACLs": []}, account=alias, **params
        )

        for acl in resp.get("WebACLs", []):
            details = _safe_aws_call(
                wafc_client.get_web_acl,
                default={"WebACL": {}},
                account=alias,
                WebACLId=acl["WebACLId"],
            ).get("WebACL", {})

            if details:
                out.append(
                    {
                        "Name": details.get("Name", ""),
                        "WebACLId": acl.get("WebACLId"),
                        "RuleCount": len(details.get("Rules", [])),
                        "Rules": [
                            r.get("RuleId", "") for r in details.get("Rules", [])
                        ],
                        "Region": wafc_client.meta.region_name,
                        "AccountAlias": alias,
                    }
                )

        next_marker = resp.get("NextMarker")
        if not next_marker:
            break

    return {"WAFClassic": out}


def get_waf_v2_details(
    wafv2_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """Collects details for WAFv2 Web ACLs."""
    out = []

    # list_web_acls for WAFv2 also requires manual pagination using NextMarker.
    next_marker = None
    while True:
        params = (
            {
                "Scope": "REGIONAL",  # This scope is required for the API call
                "NextMarker": next_marker,
            }
            if next_marker
            else {"Scope": "REGIONAL"}
        )

        resp = _safe_aws_call(
            wafv2_client.list_web_acls, default={"WebACLs": []}, account=alias, **params
        )

        for summary in resp.get("WebACLs", []):
            details = _safe_aws_call(
                wafv2_client.get_web_acl,
                default={"WebACL": {}},
                account=alias,
                Name=summary["Name"],
                Scope="REGIONAL",
                Id=summary["Id"],
            ).get("WebACL", {})

            if details:
                out.append(
                    {
                        "Name": details.get("Name"),
                        "WebACLId": details.get("Id"),
                        "RuleCount": len(details.get("Rules", [])),
                        "Rules": [r.get("Name") for r in details.get("Rules", [])],
                        "Region": wafv2_client.meta.region_name,
                        "AccountAlias": alias,
                    }
                )

        next_marker = resp.get("NextMarker")
        if not next_marker:
            break

    return {"WAFv2": out}


# =================================================================================================
# UNIFIED COLLECTORS AND MAIN SCAN LOGIC
# =================================================================================================
def get_ec2_inventory(
    ec2_client: BaseClient,
    backup_client: BaseClient,
    alias: str,
    session: boto3.Session,
) -> Dict[str, List[Dict[str, Any]]]:
    """Unified collector for all EC2-related resources."""
    log(
        "debug",
        f"Fetching EC2 and Reserved Instances in {ec2_client.meta.region_name}",
        account=alias,
    )
    return {
        "EC2": get_ec2_details(ec2_client, backup_client, alias, session),
        "EC2ReservedInstances": get_ec2_reserved_instances(ec2_client, alias),
    }


def get_rds_inventory(
    rds_client: BaseClient, alias: str, session: boto3.Session
) -> Dict[str, List[Dict[str, Any]]]:
    """Unified collector for all RDS-related resources."""
    log(
        "debug",
        f"Fetching RDS and Reserved Instances in {rds_client.meta.region_name}",
        account=alias,
    )
    return {
        "RDS": get_rds_details(rds_client, alias, session),
        "RDSReservedInstances": get_rds_reserved_instances(rds_client, alias),
    }


def get_eventbridge_inventory(
    events_client: BaseClient, scheduler_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """Unified collector for all EventBridge-related resources."""
    log(
        "debug",
        f"Fetching EventBridge and Scheduler rules in {events_client.meta.region_name}",
        account=alias,
    )
    return {
        "EventBridge": get_eventbridge_details(events_client, alias),
        "EventBridgeScheduler": get_eventbridge_scheduler_details(
            scheduler_client, alias
        ),
    }


def get_vpc_inventory(
    ec2_client: BaseClient, alias: str
) -> Dict[str, List[Dict[str, Any]]]:
    """Unified collector for all core VPC resources."""
    log(
        "debug",
        f"Fetching VPC resources in {ec2_client.meta.region_name}",
        account=alias,
    )

    vpcs = get_aws_resource_details(
        client=ec2_client, paginator_name="describe_vpcs", list_key="Vpcs", alias=alias
    )
    for vpc in vpcs:
        vpc["Tags"] = format_tags(vpc.get("Tags", []))

    subnets = get_aws_resource_details(
        client=ec2_client,
        paginator_name="describe_subnets",
        list_key="Subnets",
        alias=alias,
    )
    for subnet in subnets:
        subnet["Tags"] = format_tags(subnet.get("Tags", []))

    route_tables = []
    for rt in get_aws_resource_details(
        client=ec2_client,
        paginator_name="describe_route_tables",
        list_key="RouteTables",
        alias=alias,
    ):
        rt["IsMain"] = any(a.get("Main") for a in rt.get("Associations", []))
        rt["Tags"] = format_tags(rt.get("Tags", []))
        route_tables.append(rt)

    security_groups = []
    for sg in get_aws_resource_details(
        client=ec2_client,
        paginator_name="describe_security_groups",
        list_key="SecurityGroups",
        alias=alias,
    ):
        sg["Tags"] = format_tags(sg.get("Tags", []))
        security_groups.append(sg)

    return {
        "VPCs": vpcs,
        "Subnets": subnets,
        "RouteTables": route_tables,
        "SecurityGroups": security_groups,
    }


def scan_account(
    acct_id: str,
    sess: boto3.Session,
    regions: list[str],
    include_collectors: Optional[set[str]] = None,
    exclude_collectors: Optional[set[str]] = None,
) -> tuple[dict[str, Any], str]:
    """
    Scans a single AWS account for resources across specified regions, orchestrating
    global and regional collectors.
    """
    try:
        aliases = (
            sess.client("iam", region_name="us-east-1")
            .list_account_aliases()
            .get("AccountAliases", [])
        )
        alias = aliases[0] if aliases else acct_id
    except Exception as e:
        log(
            "warning",
            f"Could not list account aliases for {acct_id}: {e}",
            account=acct_id,
        )
        alias = acct_id

    primary_region = regions[0] if regions else "us-east-1"

    s3_global = aws_client("s3", "us-east-1", sess)
    all_buckets = _safe_aws_call(s3_global.list_buckets, default={"Buckets": []}).get(
        "Buckets", []
    )

    buckets_by_region: dict[str, list[dict[str, Any]]] = {r: [] for r in regions}
    for b in all_buckets:
        # This line now safely uses the pre-defined primary_region variable
        b_region = bucket_region(b["Name"], primary_region, sess)
        if b_region in buckets_by_region:
            buckets_by_region[b_region].append(b)

    lightsail_supported_regions = get_lightsail_supported_regions(sess, acct_id)

    global_block: dict[str, Any] = {}

    # --- Global Services Scan ---
    iam_data = get_iam_details(aws_client("iam", "us-east-1", sess), alias)

    global_collectors = {
        "IAM": lambda: iam_data,
        "IAM_Summary": lambda: get_iam_summary_details(
            iam_data.get("IAMUsers", []),
            iam_data.get("IAMRoles", []),
            iam_data.get("IAMGroups", []),
            alias,
        ),
        "Governance": lambda: {
            "Governance": get_governance_global_details(sess, alias)
        },
        "Lightsail": lambda: get_lightsail_global_details(
            aws_client("lightsail", "us-east-1", sess), alias
        ),
        "Route53": lambda: get_route53_details(
            aws_client("route53", "us-east-1", sess), alias
        ),
        "SavingsPlans": lambda: get_savings_plan_details(
            aws_client("savingsplans", "us-east-1", sess), alias
        ),
    }

    for name, fn in global_collectors.items():
        if (not include_collectors or name in include_collectors) and (
            not exclude_collectors or name not in exclude_collectors
        ):
            try:
                result = fn()
                global_block.update(result)
            except Exception as exc:
                log("error", f"Global collector {name} failed: {exc}", account=acct_id)

    # --- Regional Services Scan ---
    def region_worker(region: str) -> tuple[str, dict[str, Any]]:
        # This variable is now only used for the Governance collector call
        is_primary = region == primary_region

        def cost_collector_factory(session_obj):
            return lambda: get_cost_opportunities(
                aws_client("ec2", region, session_obj),
                aws_client("s3", region, session_obj),
                aws_client("elbv2", region, session_obj),
                aws_client("cloudwatch", region, session_obj),
                aws_client("ce", "us-east-1", session_obj),
                alias,
                session_obj,
            )

        collectors: dict[str, Callable[[], Any]] = {
            "ACM": lambda: get_acm_details(aws_client("acm", region, sess), alias),
            "ALB": lambda: get_alb_details(aws_client("elbv2", region, sess), alias),
            "Backup": lambda: get_backup_details(
                aws_client("backup", region, sess), alias
            ),
            "CloudWatchAlarms": lambda: get_cloudwatch_alarms_details(
                aws_client("cloudwatch", region, sess), alias
            ),
            "CloudWatchLogs": lambda: get_cloudwatch_logs_details(
                aws_client("logs", region, sess), alias
            ),
            "CostOpportunities": cost_collector_factory(sess),
            "EC2": lambda: get_ec2_inventory(
                aws_client("ec2", region, sess),
                aws_client("backup", region, sess),
                alias,
                sess,
            ),
            "EventBridge": lambda: get_eventbridge_inventory(
                aws_client("events", region, sess),
                aws_client("scheduler", region, sess),
                alias,
            ),
            "Governance": lambda: get_governance_details(sess, alias, region),
            "KMS": lambda: get_kms_details(aws_client("kms", region, sess), alias),
            "Lambda": lambda: get_lambda_details(
                aws_client("lambda", region, sess), alias
            ),
            "Lightsail": lambda: get_lightsail_inventory(
                aws_client("lightsail", region, sess),
                alias,
                sess,
                lightsail_supported_regions,
            ),
            "RDS": lambda: get_rds_inventory(
                aws_client("rds", region, sess), alias, sess
            ),
            "S3": lambda: get_s3_details(
                aws_client("s3", region, sess),
                buckets_by_region.get(region, []),
                alias,
                region,
                sess,
            ),
            "SNS": lambda: get_sns_details(aws_client("sns", region, sess), alias),
            "SES": lambda: get_ses_details(aws_client("ses", region, sess), alias),
            "VPC": lambda: get_vpc_inventory(aws_client("ec2", region, sess), alias),
            "VPN": lambda: get_vpn_details(aws_client("ec2", region, sess), alias),
            "WAFClassic": lambda: get_waf_classic_details(
                aws_client("waf-regional", region, sess), alias
            ),
            "WAFv2": lambda: get_waf_v2_details(
                aws_client("wafv2", region, sess), alias
            ),
        }

        collectors_to_run = {
            k: v
            for k, v in collectors.items()
            if (not include_collectors or k in include_collectors)
            and (not exclude_collectors or k not in exclude_collectors)
        }

        region_block: dict[str, Any] = {}
        with ThreadPoolExecutor(
            max_workers=MAX_TASKS_IN_REGION, thread_name_prefix=f"{region}_worker"
        ) as pool:
            future_to_name = {
                pool.submit(fn): name for name, fn in collectors_to_run.items()
            }
            for future in as_completed(future_to_name):
                sheet_name = future_to_name[future]
                try:
                    region_block.update(future.result())
                except Exception as exc:
                    log(
                        "error",
                        f"Collector {sheet_name} in {region} failed: {exc}",
                        account=acct_id,
                    )
                    region_block[sheet_name] = {}
        return region, {
            k: v
            for k, v in region_block.items()
            if v is not None and v != [] and v != {}
        }

    all_regions_data: dict[str, Any] = {}
    with ThreadPoolExecutor(
        max_workers=min(len(regions), MAX_REGIONS_IN_FLIGHT),
        thread_name_prefix="region_master",
    ) as pool:
        future_to_region = {pool.submit(region_worker, r): r for r in regions}
        for future in as_completed(future_to_region):
            region_name, service_data = future.result()
            all_regions_data[region_name] = service_data

    all_regions_data["global"] = global_block
    return all_regions_data, alias


# =================================================================================================
# EXCEL WRITER
# =================================================================================================
class StreamingExcelWriter:
    """
    A class to write AWS inventory data to a multi-sheet Excel workbook.
    This class streams data to the workbook to keep memory usage low and
    formats the output into structured tables for easy analysis.
    """

    def __init__(self, filename: str, export_tz: str):
        self._acct_ids: dict[str, str] = {}
        self.filename = filename
        self.export_tz = export_tz
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove the default sheet
        self.sheets: Dict[str, Worksheet] = {}
        # seen[(account)][sheet] = set of (unique_key, region)
        self._seen: Dict[str, Dict[str, set[Tuple[str, str]]]] = {}
        self._table_names: set[str] = set()

    @staticmethod
    def _excel_str(cell: Any, *, limit: int = 120) -> str:
        """
        Serialise a cell value for Excel, truncating long strings.
        """
        if cell is None:
            return ""
        s = str(cell)
        return (s[: limit - 3] + "...") if len(s) > limit else s

    @staticmethod
    def _logical_len(items: Union[Set[Any], Tuple[Any, ...]]) -> int:
        """
        Return the count of *logical* resources in *items*.
        """
        if not items:
            return 0
        return len({v[0] if isinstance(v, tuple) else v for v in set(items)})

    def record_account(self, account_id: str, alias: str) -> None:
        """Stores the AccountId-to-Alias mapping."""
        self._acct_ids[account_id] = alias

    def _safe_table_name(self, sheet_name: str) -> str:
        """Generate a unique, Excel-compliant table name."""
        stem = "".join(c for c in sheet_name if c.isalnum())
        stem = stem[:_MAX_EXCEL_NAME_LEN]

        name, idx = stem, 1
        while name in self._table_names or name in self.wb.sheetnames:
            suffix = f"_{idx}"
            name = f"{stem[:_MAX_EXCEL_NAME_LEN - len(suffix)]}{suffix}"
            idx += 1

        self._table_names.add(name)
        return name

    def _get_sheet(self, name: str) -> Worksheet:
        """Creates a new worksheet with headers if it doesn't exist."""
        if name not in self.sheets:
            ws = self.wb.create_sheet(name[:_MAX_EXCEL_NAME_LEN])
            if name in SERVICE_COLUMNS:
                ws.append(SERVICE_COLUMNS[name])
            self.sheets[name] = ws
        return self.sheets[name]

    def _append_serialized(
        self, ws: Worksheet, sheet_name: str, row: Dict[str, Any]
    ) -> None:
        """JSON-encodes complex cells and appends the row."""
        serialized_row = []
        for col in SERVICE_COLUMNS.get(sheet_name, []):
            cell_value = row.get(col)
            if isinstance(cell_value, (dict, list, set, tuple)):
                serialized_row.append(json.dumps(cell_value, separators=(",", ":")))
            else:
                serialized_row.append(cell_value if cell_value is not None else "")
        ws.append(serialized_row)

    def write_row(self, sheet: str, row: Dict[str, Any]) -> None:
        """
        Appends a single row to the specified sheet, handling de-duplication.
        """
        ws = self._get_sheet(sheet)
        acct_id = row.get("AccountId")
        uniq_key_def = UNIQUE_KEYS.get(sheet)

        if not (uniq_key_def and acct_id):
            self._append_serialized(ws, sheet, row)
            return

        if isinstance(uniq_key_def, list):
            base_id = tuple(str(row.get(k, "")) for k in uniq_key_def)
        else:
            base_id = row.get(uniq_key_def)

        identifier = (
            (base_id, row.get("Region")) if sheet in _REGION_SCOPED_SHEETS else base_id
        )

        seen_in_acct = self._seen.setdefault(acct_id, {}).setdefault(sheet, set())
        if identifier in seen_in_acct:
            return

        seen_in_acct.add(identifier)
        self._append_serialized(ws, sheet, row)

    def close(self) -> None:
        """Finalizes the workbook by formatting sheets and saving the file."""
        # --- Create Summary Sheet ---
        ordered_sheets = sorted(self.sheets.keys())
        summary_ws = self.wb.create_sheet("Summary", 0)
        summary_ws.append(
            [
                "Exported At:",
                datetime.now(ZoneInfo(self.export_tz)).strftime("%Y-%m-%d %H:%M:%S %Z"),
            ]
        )
        summary_ws.append([])

        summary_header = [
            "AccountId",
            "AccountAlias",
            "TotalResources",
        ] + ordered_sheets
        summary_ws.append(summary_header)

        for acct_id, svc_map in sorted(self._seen.items()):
            alias = self._acct_ids.get(acct_id, acct_id)
            total = sum(
                self._logical_len(svc_map.get(s, set())) for s in ordered_sheets
            )
            row_data = [acct_id, alias, total] + [
                self._logical_len(svc_map.get(s, set())) for s in ordered_sheets
            ]
            summary_ws.append(row_data)

        # --- Format All Sheets with Centered Alignment and Correct Sorting  ---
        all_sheets_to_format = {"Summary": summary_ws, **self.sheets}
        center_align = Alignment(horizontal="center", vertical="center")

        for name in ["Summary"] + ordered_sheets:
            ws = all_sheets_to_format[name]
            if ws.max_row <= 1:
                continue

            # Set column widths first
            for col_idx in range(1, ws.max_column + 1):
                letter = get_column_letter(col_idx)
                max_len = max(
                    (len(self._excel_str(c.value)) for c in ws[letter] if c.value),
                    default=10,
                )
                ws.column_dimensions[letter].width = max_len + 2

            # Apply centered alignment to all cells with content
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = center_align

            # Create and style a table
            header_row = 3 if name == "Summary" else 1
            table_ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
            table = Table(displayName=self._safe_table_name(name), ref=table_ref)
            style_name = (
                "TableStyleMedium9" if name == "Summary" else "TableStyleLight9"
            )
            table.tableStyleInfo = TableStyleInfo(
                name=style_name,
                showRowStripes=True,
                showFirstColumn=False,
                showLastColumn=False,
            )
            ws.add_table(table)

            # Freeze panes
            ws.freeze_panes = ws[f"A{header_row + 1}"]

        # Reorder sheets to ensure Summary is first, followed by others alphabetically
        desired_order = ["Summary"] + sorted(
            [title for title in self.wb.sheetnames if title != "Summary"]
        )
        for idx, title in enumerate(desired_order):
            ws = self.wb[title]
            current_idx = self.wb.sheetnames.index(title)
            offset = idx - current_idx
            if offset:
                self.wb.move_sheet(ws, offset)
        self.wb.active = 0

        try:
            self.wb.save(self.filename)
            log("info", f"Successfully saved report to {self.filename}", account="-")
        except PermissionError as exc:
            log(
                "error",
                f"Failed to save {self.filename} ({exc}). Is the file open?",
                account="-",
            )
            raise


# =================================================================================================
# MAIN EXECUTION BLOCK
# =================================================================================================
def parse_cli() -> argparse.Namespace:
    """Parses command-line arguments for the AWS Scan Report Generator."""
    parser = argparse.ArgumentParser(
        description="AWS Scan Report Generator",
        formatter_class=argparse.RawTextHelpFormatter,
    )

    # --- Arguments for Controlling the Scan ---
    parser.add_argument(
        "--master",
        action="store_true",
        help="Run from a management/master account to scan member accounts.",
    )
    parser.add_argument(
        "--exclude-accounts",
        default="",
        help="Comma-separated list of account IDs to exclude from the scan.\n(e.g., 111122223333,444455556666)",
    )
    parser.add_argument(
        "--include-accounts",
        default="",
        help="Comma-separated list of account IDs to exclusively scan.\n(e.g., 123456789012,987654321098)",
    )
    parser.add_argument(
        "--role-name",
        default=DEFAULT_ROLE,
        help=f"IAM role to assume in member accounts (default: {DEFAULT_ROLE}).",
    )
    parser.add_argument(
        "-r",
        "--regions",
        help="Comma-separated list of regions to scan.\n(e.g., us-east-1,us-west-2). Default: all enabled regions.",
    )
    parser.add_argument(
        "--no-excel",
        action="store_true",
        help="Disable the generation of the Excel report.",
    )
    parser.add_argument(
        "--include",
        default="",
        help="Comma-separated list of collectors to include (overrides scan mode defaults).\n(e.g., EC2,S3,IAM)",
    )
    parser.add_argument(
        "--exclude",
        default="",
        help="Comma-separated list of collectors to exclude.\n(e.g., WAFClassic,ALB)",
    )
    parser.add_argument(
        "--scan-mode",
        choices=["full", "inventory", "security", "cost"],
        # Updated to make 'inventory' the default choice
        default="inventory",
        help="Specify the scan mode:\n'full' for a complete resource scan.\n'inventory' for a basic asset scan (default).\n'security' for a security-focused scan.\n'cost' for a cost-optimization scan.",
    )

    return parser.parse_args()


def main() -> None:
    """Main function to orchestrate the AWS resource scan."""
    args = parse_cli()

    # --- Validate AWS Credentials ---
    try:
        frozen = boto3.Session().get_credentials().get_frozen_credentials()
        if not (frozen.access_key and frozen.secret_key):
            raise RuntimeError("No AWS credentials found.")
    except Exception as e:
        sys.exit(
            f"ERROR: No AWS credentials - set AWS_PROFILE or configure them via environment variables. Details: {e}"
        )

    base_session = boto3.Session()

    # --- Determine Regions to Scan ---
    try:
        ec2_client = base_session.client("ec2", config=RETRY_CONFIG)
        enabled_regions = [
            r["RegionName"]
            for r in ec2_client.describe_regions(
                Filters=[
                    {
                        "Name": "opt-in-status",
                        "Values": ["opt-in-not-required", "opted-in"],
                    }
                ]
            ).get("Regions", [])
        ]
    except (ClientError, NoCredentialsError):
        enabled_regions = base_session.get_available_regions("ec2")

    regions_to_scan = (
        [r.strip() for r in args.regions.split(",")]
        if args.regions
        else enabled_regions
    )
    if unknown_regions := set(regions_to_scan) - set(enabled_regions):
        log(
            "error",
            f"Region(s) requested but not enabled in this account: {', '.join(sorted(unknown_regions))}",
            account="-",
        )
        sys.exit(2)

    # --- Determine Accounts to Scan ---
    include_accounts = {
        a.strip() for a in args.include_accounts.split(",") if a.strip()
    }
    exclude_accounts = {
        a.strip() for a in args.exclude_accounts.split(",") if a.strip()
    }

    sessions_to_scan: Dict[str, boto3.Session] = {}
    if args.master:
        org_client = base_session.client("organizations", config=RETRY_CONFIG)
        accounts = [
            a
            for page in require_paginator(org_client, "list_accounts").paginate()
            for a in page.get("Accounts", [])
            if a.get("Status") == "ACTIVE"
        ]
        for acct in accounts:
            aid = acct["Id"]
            if (include_accounts and aid not in include_accounts) or (
                aid in exclude_accounts
            ):
                continue
            if session := assume_role(aid, args.role_name, regions_to_scan[0]):
                sessions_to_scan[aid] = session
    else:
        my_account_id = base_session.client(
            "sts", config=RETRY_CONFIG
        ).get_caller_identity()["Account"]
        if not (
            (include_accounts and my_account_id not in include_accounts)
            or (my_account_id in exclude_accounts)
        ):
            sessions_to_scan[my_account_id] = base_session

    # --- Determine Collectors to Run ---
    user_includes = {c.strip() for c in args.include.split(",") if c.strip()}
    user_excludes = {c.strip() for c in args.exclude.split(",") if c.strip()}

    # Categorize all your collectors
    CORE_INVENTORY = {
        "EC2",
        "RDS",
        "VPC",
        "S3",
        "ALB",
        "IAM",
        "IAM_Summary",
        "Lambda",
        "KMS",
        "Route53",
        "EventBridge",
        "EventBridgeScheduler",
        "Backup",
        "ACM",
        "VPN",
        "SNS",
        "SES",
        "Lightsail",
    }
    SECURITY_FINDINGS = {
        "Governance",
        "WAFv2",
        "WAFClassic",
    }
    COST_FINDINGS = {
        "CostOpportunities",
        "SavingsPlans",
    }

    # Define scan modes based on categories
    if user_includes:
        collectors_to_run = user_includes
    elif args.scan_mode == "security":
        log("info", "Running in SECURITY scan mode.", account="-")
        collectors_to_run = CORE_INVENTORY | SECURITY_FINDINGS
    elif args.scan_mode == "cost":
        log("info", "Running in COST scan mode.", account="-")
        collectors_to_run = CORE_INVENTORY | COST_FINDINGS
    elif args.scan_mode == "inventory":
        log("info", "Running in lean INVENTORY scan mode.", account="-")
        collectors_to_run = CORE_INVENTORY
    else:  # This can be a new 'full' mode, or the default
        log("info", "Running in FULL scan mode.", account="-")
        collectors_to_run = CORE_INVENTORY | SECURITY_FINDINGS | COST_FINDINGS

    collectors_to_run -= user_excludes

    # --- Initialize Excel Writer with Dynamic Filename ---
    if not args.no_excel:
        current_date = datetime.now().strftime("%Y-%m-%d")
        dynamic_filename = f"aws_scan_report_{args.scan_mode}_{current_date}.xlsx"
        log("info", f"Output will be saved to: {dynamic_filename}", account="-")
        writer = StreamingExcelWriter(dynamic_filename, str(get_localzone() or "UTC"))
    else:
        writer = None

    # --- Execute Scan and Write Report ---
    for acct_id, sess in sessions_to_scan.items():
        log("info", f"Starting scan for account: {acct_id}", account=acct_id)
        account_data, alias = scan_account(
            acct_id,
            sess,
            regions_to_scan,
            include_collectors=collectors_to_run,
            exclude_collectors=user_excludes,
        )

        if writer:
            writer.record_account(acct_id, alias)
            for region_name, svc_block in account_data.items():
                for sheet, rows in svc_block.items():
                    if rows:
                        for row in rows:
                            row["AccountId"] = acct_id
                            writer.write_row(sheet, row)

    if writer:
        writer.close()

    log("info", "AWS scan completed.", account="-")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit("\nScan interrupted by user.")
    except Exception as e:
        log(
            "critical",
            f"An unexpected error occurred during the scan: {e}",
            account="-",
        )
        sys.exit(1)
